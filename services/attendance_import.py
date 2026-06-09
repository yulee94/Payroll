"""
attendance_import.py - normalize biometric/work attendance files into payroll rows.

The current payroll engine already knows how to calculate from invoice-like
rows. This module turns uploaded attendance records into that shape so invoice,
attendance, and mixed payroll inputs can share the same downstream engine.
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path
from typing import Any, Iterable

from roster_constants import norm_name_key
from services.payroll_policy_store import resolve_payroll_operation_policy
from utils import safe_number

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "name": ("성명", "이름", "사원명", "직원명", "근로자명", "직원성명", "근무자", "name", "employee"),
    "dept": ("부서", "소속", "팀", "근무부서", "dept", "department"),
    "workplace": ("사업장", "근무지", "현장", "현장명", "site", "workplace"),
    "date": ("일자", "날짜", "근무일", "근태일자", "date"),
    "clock_in": ("출근", "출근시간", "출근시각", "입실", "clockin", "clock_in", "start"),
    "clock_out": ("퇴근", "퇴근시간", "퇴근시각", "퇴실", "clockout", "clock_out", "end"),
    "work_hours": ("근무시간", "실근로", "실근로시간", "총근무", "총근로시간", "workhours", "hours"),
    "break_minutes": ("휴게", "휴게시간", "휴게분", "break", "breakminutes"),
    "late_minutes": ("지각", "지각시간", "지각분", "late", "lateminutes"),
    "early_leave_minutes": ("조퇴", "조퇴시간", "조퇴분", "지조외", "earlyleave"),
    "overtime_hours": ("연장", "연장시간", "연장근로", "잔업", "ot", "overtime"),
    "night_hours": ("야간", "야간시간", "심야", "심야시간", "night"),
    "special_hours": ("특근", "특근시간", "휴일", "휴일근로", "special", "holidaywork"),
    "leave_days": ("연차", "연차일수", "휴가", "휴가일수", "leave"),
    "unpaid_days": ("결근", "결근일수", "무급", "무급일", "absence", "unpaid"),
}

DEFAULT_SYMBOL_WORK_HOURS = 8.0
HALF_DAY_WORK_HOURS = 4.0

WORK_DAY_MARKERS = {
    "●",
    "○",
    "◎",
    "◉",
    "■",
    "□",
    "✓",
    "✔",
    "o",
    "ok",
    "출",
    "근",
    "출근",
    "정상",
    "근무",
}
HALF_WORK_DAY_MARKERS = {
    "◐",
    "◑",
    "△",
    "▲",
    "반근",
    "반일",
    "반근무",
}
LEAVE_DAY_MARKERS = {
    "연",
    "연차",
    "월차",
    "휴가",
    "유급",
    "유급휴가",
}
HALF_LEAVE_DAY_MARKERS = {
    "반",
    "반차",
    "오전반차",
    "오후반차",
    "반휴",
    "반휴가",
}
UNPAID_DAY_MARKERS = {
    "결",
    "결근",
    "무",
    "무급",
    "무단",
    "무단결근",
    "x",
    "×",
    "✕",
    "✖",
}
HALF_UNPAID_DAY_MARKERS = {
    "◔",
    "반결",
    "반결근",
    "반무",
    "반무급",
    "무급반",
}
BLANK_ATTENDANCE_MARKERS = {
    "-",
    "–",
    "—",
    "·",
    ".",
    "휴",
    "휴무",
    "휴일",
    "공휴",
    "공휴일",
    "비번",
    "off",
}

_DAY_HEADER_RE = re.compile(r"^0?(\d{1,2})(?:일|day)?(?:\([^)]*\))?$", re.IGNORECASE)
_DATE_DAY_HEADER_RE = re.compile(
    r"^(?:(?:\d{4}년)?\d{1,2}월|(?:\d{4}[-/.])?\d{1,2}[-/.])0?(\d{1,2})일?(?:\([^)]*\))?$",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class AttendanceImportResult:
    source_path: Path
    rows: list[dict[str, Any]]
    invoice_rows: list[dict[str, Any]]
    warnings: list[str]

    @property
    def count(self) -> int:
        return len(self.invoice_rows)


@dataclass(frozen=True)
class AttendanceMarkerParts:
    work_hours: float = 0.0
    leave_days: float = 0.0
    unpaid_days: float = 0.0
    marked: bool = False


def _compact(text: Any) -> str:
    return str(text or "").strip().replace(" ", "").replace("\n", "").lower()


def _canonical_header(text: Any) -> str | None:
    c = _compact(text)
    if not c:
        return None
    for key, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if c == _compact(alias) or _compact(alias) in c:
                return key
    return None


def _parse_day_header(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.day
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if float(value).is_integer():
            day = int(value)
            return day if 1 <= day <= 31 else None
        return None
    text = str(value or "").strip()
    if not text:
        return None
    compact = _compact(text)
    if compact.endswith(".0"):
        compact = compact[:-2]
    match = _DAY_HEADER_RE.match(compact)
    if not match:
        match = _DATE_DAY_HEADER_RE.match(compact)
    if not match:
        return None
    day = int(match.group(1))
    return day if 1 <= day <= 31 else None


def _calendar_day_columns(header_row: list[Any], mapping: dict[str, int]) -> list[int]:
    mapped = set(mapping.values())
    cols: list[int] = []
    for col, value in enumerate(header_row):
        if col in mapped:
            continue
        if _parse_day_header(value) is not None:
            cols.append(col)
    return cols


def _read_csv_rows(path: Path) -> list[list[Any]]:
    for encoding in ("utf-8-sig", "cp949", "euc-kr"):
        try:
            with path.open("r", encoding=encoding, newline="") as f:
                return [list(row) for row in csv.reader(f)]
        except UnicodeDecodeError:
            continue
    with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
        return [list(row) for row in csv.reader(f)]


def _read_xlsx_rows(path: Path) -> list[list[Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _read_table_rows(path: Path) -> list[list[Any]]:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return _read_xlsx_rows(path)
    if suffix in (".csv", ".txt"):
        return _read_csv_rows(path)
    raise ValueError("근태 파일은 .xlsx, .xlsm, .csv, .txt 형식만 지원합니다.")


def _find_header(rows: list[list[Any]]) -> tuple[int, dict[str, int]]:
    best: tuple[int, int, dict[str, int]] | None = None
    for idx, row in enumerate(rows[:12]):
        mapping: dict[str, int] = {}
        for col, value in enumerate(row):
            key = _canonical_header(value)
            if key and key not in mapping:
                mapping[key] = col
        score = len(mapping)
        day_cols = _calendar_day_columns(row, mapping)
        if "name" in mapping and (score >= 2 or day_cols):
            rank = score * 10 + min(len(day_cols), 31)
            if best is None or rank > best[1]:
                best = (idx, rank, mapping)
    if best:
        return best[0], best[2]
    raise ValueError("근태 파일에서 성명/출퇴근 헤더를 찾지 못했습니다.")


def _value(row: list[Any], mapping: dict[str, int], key: str) -> Any:
    idx = mapping.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _parse_time(value: Any) -> time | None:
    if _is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    if isinstance(value, (int, float)):
        # Excel stores time as a day fraction when the cell is formatted as time.
        if 0 <= float(value) < 1:
            total_minutes = round(float(value) * 24 * 60)
            return time(total_minutes // 60 % 24, total_minutes % 60)
        return None
    raw = str(value).strip()
    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%p %I:%M"):
        try:
            return datetime.strptime(raw, fmt).time()
        except ValueError:
            continue
    return None


def _hours_from_clock(clock_in: Any, clock_out: Any, break_minutes: Any = None) -> float:
    start = _parse_time(clock_in)
    end = _parse_time(clock_out)
    if not start or not end:
        return 0.0
    start_minutes = start.hour * 60 + start.minute + start.second / 60
    end_minutes = end.hour * 60 + end.minute + end.second / 60
    if end_minutes < start_minutes:
        end_minutes += 24 * 60
    minutes = max(0.0, end_minutes - start_minutes - safe_number(break_minutes, 0.0))
    return minutes / 60.0


def _number_as_hours(value: Any) -> float:
    if _is_blank(value):
        return 0.0
    raw = str(value).strip()
    if ":" in raw:
        parts = raw.split(":")
        try:
            hours = float(parts[0])
            minutes = float(parts[1]) if len(parts) > 1 else 0.0
            return max(0.0, hours + minutes / 60.0)
        except ValueError:
            return 0.0
    n = safe_number(value, 0.0)
    if n > 24:
        return n / 60.0
    return n


def _marker_key(value: Any) -> str:
    return _compact(value).replace(" ", "")


def _marker_attendance_parts(value: Any) -> AttendanceMarkerParts:
    if _is_blank(value):
        return AttendanceMarkerParts()
    key = _marker_key(value)
    if not key or key in BLANK_ATTENDANCE_MARKERS:
        return AttendanceMarkerParts()
    if key in WORK_DAY_MARKERS:
        return AttendanceMarkerParts(work_hours=DEFAULT_SYMBOL_WORK_HOURS, marked=True)
    if key in HALF_WORK_DAY_MARKERS:
        return AttendanceMarkerParts(work_hours=HALF_DAY_WORK_HOURS, marked=True)
    if key in HALF_LEAVE_DAY_MARKERS:
        return AttendanceMarkerParts(
            work_hours=HALF_DAY_WORK_HOURS,
            leave_days=0.5,
            marked=True,
        )
    if key in LEAVE_DAY_MARKERS:
        return AttendanceMarkerParts(leave_days=1.0, marked=True)
    if key in HALF_UNPAID_DAY_MARKERS:
        return AttendanceMarkerParts(
            work_hours=HALF_DAY_WORK_HOURS,
            unpaid_days=0.5,
            marked=True,
        )
    if key in UNPAID_DAY_MARKERS:
        return AttendanceMarkerParts(unpaid_days=1.0, marked=True)
    return AttendanceMarkerParts()


def _work_hours_from_value(value: Any) -> float:
    hours = _number_as_hours(value)
    if hours > 0:
        return hours
    return _marker_attendance_parts(value).work_hours


def _days_from_value(value: Any, *, field: str) -> float:
    if _is_blank(value):
        return 0.0
    days = safe_number(value, 0.0)
    if days > 0:
        return days
    key = _marker_key(value)
    if field == "leave":
        if key in HALF_LEAVE_DAY_MARKERS:
            return 0.5
        if key in LEAVE_DAY_MARKERS:
            return 1.0
    elif field == "unpaid":
        if key in HALF_UNPAID_DAY_MARKERS:
            return 0.5
        if key in UNPAID_DAY_MARKERS:
            return 1.0
    if key in WORK_DAY_MARKERS:
        return 1.0
    if key in HALF_WORK_DAY_MARKERS:
        return 0.5
    return 0.0


def _calendar_marker_totals(row: list[Any], day_cols: list[int]) -> AttendanceMarkerParts:
    work_hours = 0.0
    leave_days = 0.0
    unpaid_days = 0.0
    marked_days = 0
    for col in day_cols:
        if col >= len(row):
            continue
        value = row[col]
        hours = _number_as_hours(value)
        if hours > 0:
            work_hours += hours
            marked_days += 1
            continue
        parts = _marker_attendance_parts(value)
        if not parts.marked:
            continue
        work_hours += parts.work_hours
        leave_days += parts.leave_days
        unpaid_days += parts.unpaid_days
        marked_days += 1
    return AttendanceMarkerParts(
        work_hours=work_hours,
        leave_days=leave_days,
        unpaid_days=unpaid_days,
        marked=marked_days > 0,
    )


def _calendar_marked_days(row: list[Any], day_cols: list[int]) -> int:
    count = 0
    for col in day_cols:
        if col >= len(row):
            continue
        value = row[col]
        if _number_as_hours(value) > 0 or _marker_attendance_parts(value).marked:
            count += 1
    return count


def _minutes_as_hours(value: Any) -> float:
    if _is_blank(value):
        return 0.0
    raw = str(value).strip()
    if ":" in raw:
        return _number_as_hours(value)
    return max(0.0, safe_number(value, 0.0) / 60.0)


def _round_hours(hours: float, rounding_minutes: int) -> float:
    if hours <= 0:
        return 0.0
    unit = max(1, int(rounding_minutes))
    minutes = round(hours * 60 / unit) * unit
    return round(minutes / 60.0, 4)


def _iter_record_rows(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    rows = _read_table_rows(path)
    header_idx, mapping = _find_header(rows)
    day_cols = _calendar_day_columns(rows[header_idx], mapping)
    out: list[dict[str, Any]] = []
    warnings: list[str] = []
    for source_row_no, row in enumerate(rows[header_idx + 1 :], header_idx + 2):
        name = str(_value(row, mapping, "name") or "").strip()
        if not name or name in ("합계", "소계", "계"):
            continue
        marker_totals = _calendar_marker_totals(row, day_cols)
        marked_days = _calendar_marked_days(row, day_cols)
        work_hours = _work_hours_from_value(_value(row, mapping, "work_hours"))
        if work_hours <= 0:
            work_hours = _hours_from_clock(
                _value(row, mapping, "clock_in"),
                _value(row, mapping, "clock_out"),
                _value(row, mapping, "break_minutes"),
            )
        if marker_totals.marked and work_hours <= 0:
            work_hours = marker_totals.work_hours
        explicit_leave_days = _days_from_value(
            _value(row, mapping, "leave_days"),
            field="leave",
        )
        explicit_unpaid_days = _days_from_value(
            _value(row, mapping, "unpaid_days"),
            field="unpaid",
        )
        record = {
            "source_row": source_row_no,
            "name": name,
            "name_key": norm_name_key(name),
            "dept": str(_value(row, mapping, "dept") or "").strip(),
            "workplace": str(_value(row, mapping, "workplace") or "").strip(),
            "date": str(_value(row, mapping, "date") or "").strip(),
            "work_hours": max(0.0, work_hours),
            "late_hours": _minutes_as_hours(_value(row, mapping, "late_minutes")),
            "early_leave_hours": _minutes_as_hours(_value(row, mapping, "early_leave_minutes")),
            "overtime_hours": _number_as_hours(_value(row, mapping, "overtime_hours")),
            "night_hours": _number_as_hours(_value(row, mapping, "night_hours")),
            "special_hours": _number_as_hours(_value(row, mapping, "special_hours")),
            "leave_days": explicit_leave_days
            if explicit_leave_days > 0
            else marker_totals.leave_days,
            "unpaid_days": explicit_unpaid_days
            if explicit_unpaid_days > 0
            else marker_totals.unpaid_days,
            "_attendance_days": marked_days or 1,
        }
        if record["work_hours"] <= 0 and record["leave_days"] <= 0 and record["unpaid_days"] <= 0:
            warnings.append(f"{source_row_no}행 {name}: 근무시간/휴가/결근 값이 없어 제외했습니다.")
            continue
        out.append(record)
    return out, warnings


def _aggregate_records(
    records: Iterable[dict[str, Any]],
    *,
    workplace: str,
    rounding_minutes: int,
    late_grace_minutes: int,
    early_leave_grace_minutes: int,
) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for rec in records:
        key = rec["name_key"]
        if not key:
            continue
        item = grouped.setdefault(
            key,
            {
                "name": rec["name"],
                "dept": rec.get("dept", ""),
                "workplace": rec.get("workplace") or workplace,
                "work_hours": 0.0,
                "late_hours": 0.0,
                "early_leave_hours": 0.0,
                "overtime_hours": 0.0,
                "night_hours": 0.0,
                "special_hours": 0.0,
                "leave_days": 0.0,
                "unpaid_days": 0.0,
                "_attendance_days": 0,
            },
        )
        item["work_hours"] += rec.get("work_hours", 0.0)
        late_h = rec.get("late_hours", 0.0)
        early_h = rec.get("early_leave_hours", 0.0)
        if late_h * 60 > late_grace_minutes:
            item["late_hours"] += max(0.0, late_h - late_grace_minutes / 60.0)
        if early_h * 60 > early_leave_grace_minutes:
            item["early_leave_hours"] += max(0.0, early_h - early_leave_grace_minutes / 60.0)
        item["overtime_hours"] += rec.get("overtime_hours", 0.0)
        item["night_hours"] += rec.get("night_hours", 0.0)
        item["special_hours"] += rec.get("special_hours", 0.0)
        item["leave_days"] += rec.get("leave_days", 0.0)
        item["unpaid_days"] += rec.get("unpaid_days", 0.0)
        item["_attendance_days"] += max(1, int(rec.get("_attendance_days") or 1))

    rows: list[dict[str, Any]] = []
    for item in grouped.values():
        work_h = _round_hours(item["work_hours"], rounding_minutes)
        early_h = _round_hours(
            item["early_leave_hours"] + item["late_hours"],
            rounding_minutes,
        )
        rows.append(
            {
                "row": 0,
                "name": item["name"],
                "dept": item.get("dept", ""),
                "hire_date": "",
                "workplace": item.get("workplace") or workplace,
                "base_hourly": 0.0,
                "ordinary_hourly": 0.0,
                "base_days": work_h,
                "work_days": work_h,
                "unpaid_days": item["unpaid_days"],
                "leave_days": item["leave_days"],
                "ot_hours": _round_hours(item["overtime_hours"], rounding_minutes),
                "shift_hours": 0.0,
                "night_hours": _round_hours(item["night_hours"], rounding_minutes),
                "special_hours": _round_hours(item["special_hours"], rounding_minutes),
                "special_ext_hours": 0.0,
                "early_leave_hours": early_h,
                "base_salary": 0,
                "base_deduction": 0,
                "ot_pay": 0,
                "night_pay": 0,
                "special_pay": 0,
                "special_ext_pay": 0,
                "position_pay": 0,
                "shift_pay": 0,
                "workers_day_pay": 0,
                "annual_pay": 0,
                "transport": 0,
                "subtotal": 0,
                "gross_pay": 0,
                "health_insurance": 0,
                "long_term_care": 0,
                "national_pension": 0,
                "employment_insurance": 0,
                "insurance_total": 0,
                "_attendance_days": item["_attendance_days"],
                "_attendance_input": True,
            }
        )
    rows.sort(key=lambda r: str(r["name"]))
    return rows


def extract_attendance_invoice_rows(
    attendance_path: Path | str,
    *,
    workplace: str = "",
    tenant_id: str | None = None,
) -> AttendanceImportResult:
    """Parse attendance upload and return invoice-compatible payroll rows."""
    path = Path(attendance_path)
    if not path.is_file():
        raise FileNotFoundError(str(path))

    resolved = resolve_payroll_operation_policy(workplace, tenant_id=tenant_id)
    attendance_policy = resolved["policy"]["attendance"]
    records, warnings = _iter_record_rows(path)
    invoice_rows = _aggregate_records(
        records,
        workplace=workplace,
        rounding_minutes=int(attendance_policy["rounding_minutes"]),
        late_grace_minutes=int(attendance_policy["late_grace_minutes"]),
        early_leave_grace_minutes=int(attendance_policy["early_leave_grace_minutes"]),
    )
    if not invoice_rows:
        warnings.append("급여 산출에 사용할 근태 행이 없습니다.")
    return AttendanceImportResult(
        source_path=path,
        rows=records,
        invoice_rows=invoice_rows,
        warnings=warnings,
    )
