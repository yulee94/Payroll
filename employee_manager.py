"""
직원 마스터(직원정보.xlsx) 로드 및 청구서 데이터와 병합.

사번·이름으로 매칭하며, 마스터에 있는 4대보험·소득세·통상시급 등을
청구서 추출값과 합칩니다.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

from utils import safe_number

# 직원정보.xlsx 필수/권장 열 이름
EMPLOYEE_COLUMNS = [
    "사번",
    "이름",
    "부서",
    "통상시급",
    "국민연금",
    "건강보험",
    "소득세",
    "계좌번호",
]

# 추가로 읽을 수 있는 열
OPTIONAL_COLUMNS = ["고정수당", "교통비", "기본급"]


def _normalize_key(text: Any) -> str:
    """이름/사번 비교용 정규화 (공백 제거)."""
    if text is None:
        return ""
    return str(text).strip().replace(" ", "")


def load_employee_data(employee_file: Path) -> Dict[str, Dict[str, Any]]:
    """
    직원정보.xlsx를 읽어 {키: 직원dict} 형태로 반환합니다.

    키는 '이름'과 '사번' 두 가지로 등록되어 이름/사번 어느 쪽으로도
    조회할 수 있습니다.

    Parameters
    ----------
    employee_file : Path
        employees/직원정보.xlsx 경로

    Returns
    -------
    dict
        정규화된 이름 또는 사번 → 직원 정보
    """
    if not employee_file.exists():
        return {}

    wb = openpyxl.load_workbook(employee_file, data_only=True)
    ws = wb.active

    # 헤더 행 찾기 (1행 또는 '사번'/'이름' 포함 행)
    header_row = 1
    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val:
            headers[str(val).strip()] = col

    employees_by_key: Dict[str, Dict[str, Any]] = {}

    for row in range(2, ws.max_row + 1):
        record: Dict[str, Any] = {}
        for col_name in EMPLOYEE_COLUMNS + OPTIONAL_COLUMNS:
            col_idx = headers.get(col_name)
            if col_idx:
                record[col_name] = ws.cell(row, col_idx).value

        name = _normalize_key(record.get("이름"))
        emp_no = _normalize_key(record.get("사번"))
        if not name and not emp_no:
            continue

        # 숫자 필드 변환
        for num_field in ("통상시급", "국민연금", "건강보험", "소득세", "고정수당", "교통비", "기본급"):
            if num_field in record and record[num_field] is not None:
                record[num_field] = safe_number(record[num_field])

        if name:
            employees_by_key[name] = record
        if emp_no:
            employees_by_key[emp_no] = record

    wb.close()
    return employees_by_key


def find_employee(
    master: Dict[str, Dict[str, Any]],
    name: str,
    emp_no: Optional[str] = None,
) -> Optional[Dict[str, Any]]:
    """
    이름 또는 사번으로 직원 마스터 레코드를 찾습니다.
    """
    key_name = _normalize_key(name)
    if key_name and key_name in master:
        return master[key_name]
    if emp_no:
        key_no = _normalize_key(emp_no)
        if key_no in master:
            return master[key_no]
    return None


def merge_employee_data(
    invoice_rows: List[Dict[str, Any]],
    master: Dict[str, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    청구서 추출 데이터와 직원 마스터를 병합합니다.

    마스터에 통상시급·교통비·4대보험 preset 등이 있으면
    청구서 값보다 우선하거나 보충합니다.
    """
    merged: List[Dict[str, Any]] = []
    for inv in invoice_rows:
        emp = find_employee(master, inv.get("name", ""), inv.get("emp_no"))
        row = dict(inv)

        if emp:
            row["emp_no"] = emp.get("사번", "")
            row["department"] = emp.get("부서", "")
            row["account_no"] = emp.get("계좌번호", "")
            row["preset_national_pension"] = emp.get("국민연금")
            row["preset_health"] = emp.get("건강보험")
            row["preset_income_tax"] = emp.get("소득세")
            row["fixed_allowance"] = safe_number(emp.get("고정수당", 0))

            # 통상시급: 마스터 우선
            master_oh = safe_number(emp.get("통상시급", 0))
            if master_oh > 0:
                row["ordinary_hourly"] = master_oh

            # 교통비: 청구서 0이면 마스터 값 사용
            master_transport = safe_number(emp.get("교통비", 0))
            if safe_number(row.get("transport_allowance", 0)) <= 0 and master_transport > 0:
                row["transport_allowance"] = master_transport

            # 기본급: 청구서 없으면 마스터
            master_base = safe_number(emp.get("기본급", 0))
            if safe_number(row.get("base_salary", 0)) <= 0 and master_base > 0:
                row["base_salary"] = master_base
        else:
            row.setdefault("emp_no", "")
            row.setdefault("department", "")
            row.setdefault("account_no", "")
            row.setdefault("fixed_allowance", 0)
            row.setdefault("preset_national_pension", None)
            row.setdefault("preset_health", None)
            row.setdefault("preset_income_tax", None)

        merged.append(row)
    return merged
