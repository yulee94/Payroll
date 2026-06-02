"""
샘플 양식·직원정보·테스트 청구서 생성 스크립트.

templates/ 또는 employees/ 파일이 없을 때 main.py·test_e2e.py에서 호출합니다.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font


def _create_ledger_template(path: Path) -> None:
    """급여대장 양식 생성."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "급여대장"
    ws["A1"] = "급여대장"
    ws["A1"].font = Font(bold=True, size=14)
    headers = [
        "사번", "이름", "부서", "기본급", "연장", "야간", "휴일", "주휴",
        "식대", "교통비", "기타수당", "총지급", "국민연금", "건강보험",
        "장기요양", "고용보험", "소득세", "지방세", "공제합계", "실수령",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(4, i, h).font = Font(bold=True)
    wb.save(path)
    wb.close()


def _create_payslip_template(path: Path) -> None:
    """급여명세서 양식 (직원당 1시트 복사용)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "명세서양식"
    ws["A1"] = "급여명세서"
    labels = [
        (7, "기본급"), (8, "연장수당"), (9, "야간수당"), (10, "휴일수당"),
        (11, "주휴수당"), (12, "식대"), (13, "교통비"), (14, "기타"),
        (15, "추가수당"), (16, "중복가산"), (18, "총지급액"),
        (21, "국민연금"), (22, "건강보험"), (23, "장기요양"), (24, "고용보험"),
        (25, "소득세"), (26, "지방소득세"), (27, "공제합계"), (29, "실수령액"),
    ]
    for row, label in labels:
        ws[f"C{row}"] = label
    wb.save(path)
    wb.close()


def _create_payment_template(path: Path) -> None:
    """지급내역 양식."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "지급내역"
    ws["A1"] = "지급내역"
    for i, h in enumerate(["사번", "이름", "계좌번호", "실수령액"], 1):
        ws.cell(2, i, h).font = Font(bold=True)
    wb.save(path)
    wb.close()


def _create_employee_file(path: Path) -> None:
    """직원정보.xlsx 샘플."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "직원정보"
    headers = [
        "사번",
        "이름",
        "부서",
        "기본시급",
        "통상시급",
        "잔여연차",
        "발생연차",
        "사용연차",
        "국민연금",
        "건강보험",
        "소득세",
        "계좌번호",
        "고정수당",
        "교통비",
        "기본급",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h).font = Font(bold=True)

    sample = [
        ("E001", "김철수", "생산", 10030, 15300, 10, 15, 5, 120000, 95000, 85000, "110-123-456789", 200000, 100000, 2800000),
        ("E002", "이영희", "인사", 10030, 16700, 8, 15, 7, 130000, 102000, 92000, "110-987-654321", 150000, 80000, 3100000),
        ("E003", "박민수", "IT", 10030, 18200, 5, 15, 10, 140000, 110000, 105000, "333-12-3456789", 100000, 120000, 3400000),
    ]
    for r, row in enumerate(sample, 2):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    wb.save(path)
    wb.close()


def _create_sample_invoice(path: Path) -> None:
    """테스트용 청구서 (5행부터 데이터)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "청구서"

    # 헤더 (4행)
    headers = {
        3: "이름", 4: "직무", 5: "성별", 7: "기본시급", 8: "통상시급",
        10: "식대일수", 13: "연장", 14: "야간", 15: "휴일",
        16: "기타", 17: "추가", 19: "기본급", 20: "교통비",
    }
    for col, label in headers.items():
        ws.cell(4, col, label)

    # 연장시간 열 (K=11)
    ws.cell(4, 11, "연장시간")

    rows = [
        # name, job, gender, base_h, ord_h, meal, ot_h(col11), ot(M), night(N), hol(O), other, add, base_s, transport
        ("김철수", "조립", "남", 12000, 15300, 22, 10, 10, 5, 0, 0, 0, 2800000, 0),
        ("이영희", "경리", "여", 13000, 16700, 20, 5, 5, 0, 8, 50000, 0, 3100000, 0),
        ("박민수", "개발", "남", 14000, 18200, 21, 8, 8, 4, 4, 0, 100000, 3400000, 0),
    ]
    for i, row in enumerate(rows, 5):
        ws.cell(i, 3, row[0])
        ws.cell(i, 4, row[1])
        ws.cell(i, 5, row[2])
        ws.cell(i, 7, row[3])
        ws.cell(i, 8, row[4])
        ws.cell(i, 10, row[5])
        ws.cell(i, 11, row[6])   # 연장시간
        ws.cell(i, 13, row[7])   # M 연장 (시간과 동일하게)
        ws.cell(i, 14, row[8])
        ws.cell(i, 15, row[9])
        ws.cell(i, 16, row[10])
        ws.cell(i, 17, row[11])
        ws.cell(i, 19, row[12])
        ws.cell(i, 20, row[13])

    wb.save(path)
    wb.close()


def ensure_sample_files(base_dir: Path) -> None:
    """
    templates/, employees/ 아래 필수 파일이 없으면 샘플을 생성합니다.
    """
    templates = base_dir / "templates"
    employees = base_dir / "employees"
    templates.mkdir(parents=True, exist_ok=True)
    employees.mkdir(parents=True, exist_ok=True)
    (base_dir / "output").mkdir(parents=True, exist_ok=True)

    files = {
        templates / "급여대장양식.xlsx": _create_ledger_template,
        templates / "급여명세서양식.xlsx": _create_payslip_template,
        templates / "지급내역양식.xlsx": _create_payment_template,
        templates / "근로자명부.xlsx": _create_employee_file,
        employees / "직원정보.xlsx": _create_employee_file,
    }
    for path, creator in files.items():
        if not path.exists():
            creator(path)


def create_test_invoice(base_dir: Path) -> Path:
    """E2E 테스트용 샘플 청구서 경로 반환 (없으면 생성)."""
    test_dir = base_dir / "test_data"
    test_dir.mkdir(parents=True, exist_ok=True)
    inv = test_dir / "샘플_청구서.xlsx"
    if not inv.exists():
        _create_sample_invoice(inv)
    return inv
