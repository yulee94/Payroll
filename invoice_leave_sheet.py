"""
invoice_leave_sheet.py - 도급비 청구서 내 「연차」 시트 파싱 → 연차대장 양식용 데이터
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from annual_leave_manager import calc_absence_occurrence_count, format_absence_usage_memo, format_usage_date_memos
from invoice_parser import _cell_str, _is_valid_name, _header_matches_name
from roster_constants import find_fuzzy_name_key, norm_name_key
from utils import safe_number

_PERIOD_TITLE_RE = re.compile(r"(20\d{2})\D{0,3}(\d{1,2})")
_PERIOD_SHORT_RE = re.compile(r"(\d{2})\s*년\s*(\d{1,2})\s*월")

_LEAVE_MARKERS = ("연", "휴", "차", "반")
_ABSENCE_MARKERS = ("무", "결", "무급", "결근", "결과")


@dataclass
class InvoiceLeaveRow:
    name: str
    name_key: str
    leave_days: float = 0.0
    unpaid_days: float = 0.0
    leave_dates: list[int] = field(default_factory=list)
    absence_dates: list[int] = field(default_factory=list)
    leave_memo: str = ""
    absence_memo: str = ""
    display: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "leave_days": self.leave_days,
            "unpaid_days": self.unpaid_days,
            "leave_memo": self.leave_memo,
            "absence_memo": self.absence_memo,
            "display": self.display,
            "leave_dates": list(self.leave_dates),
            "absence_dates": list(self.absence_dates),
        }


def infer_period_from_workbook(wb: Workbook) -> str | None:
    for name in wb.sheetnames:
        m = _PERIOD_TITLE_RE.search(name.replace(" ", ""))
        if m:
            return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
        m2 = _PERIOD_SHORT_RE.search(name)
        if m2:
            y = 2000 + int(m2.group(1))
            return f"{y:04d}-{int(m2.group(2)):02d}"
    return None


def find_leave_worksheet(wb: Workbook) -> Worksheet | None:
    """청구서 내 연차 시트 (청구내역 시트 제외)."""
    for name in wb.sheetnames:
        compact = name.replace(" ", "")
        if "연차" in compact and "청구" not in compact:
            return wb[name]
    if len(wb.sheetnames) >= 2:
        second = wb[wb.sheetnames[1]]
        first = wb[wb.sheetnames[0]]
        if second.title != first.title and "연차" in second.title.replace(" ", ""):
            return second
    return None


def _parse_day_header(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and 1 <= int(value) <= 31:
        return int(value)
    text = str(value).strip().replace("일", "")
    if text.isdigit():
        d = int(text)
        if 1 <= d <= 31:
            return d
    return None


def _scan_leave_sheet_layout(ws: Worksheet) -> tuple[int, int, list[int]] | None:
    """헤더 행, 성명 열, 일자(1~31) 열 목록."""
    name_col: int | None = None
    header_row = 1
    day_cols: dict[int, int] = {}

    for r in range(1, min(8, ws.max_row + 1)):
        for c in range(1, min(ws.max_column + 1, 40)):
            raw = ws.cell(r, c).value
            if raw is None:
                continue
            text = str(raw).strip()
            if _header_matches_name(text):
                name_col = c
                header_row = r
            day = _parse_day_header(raw)
            if day is not None:
                day_cols[day] = c
        if name_col and day_cols:
            days_sorted = sorted(day_cols.keys())
            return header_row, name_col, [day_cols[d] for d in days_sorted]

    if name_col is None:
        return None
    # 성명 열은 있으나 일 헤더가 없으면 연속 열을 일자로 추정
    if not day_cols and name_col:
        start = name_col + 1
        cols = list(range(start, min(start + 31, ws.max_column + 1)))
        if cols:
            return header_row, name_col, cols
    return None


def _classify_cell(value: Any) -> tuple[float, float, int | None, str]:
    """
    Returns:
        (leave_days_add, unpaid_days_add, day_hint, kind)
        kind: 'leave' | 'absence' | ''
    """
    if value is None:
        return 0.0, 0.0, None, ""
    if isinstance(value, (int, float)):
        n = float(value)
        if n <= 0:
            return 0.0, 0.0, None, ""
        if abs(n - 0.5) < 1e-9:
            return 0.5, 0.0, None, "leave"
        if n < 1:
            return n, 0.0, None, "leave"
        return n, 0.0, None, "leave"

    text = str(value).strip().replace(" ", "")
    if not text or text in ("-", "0", "○", "O", "o"):
        return 0.0, 0.0, None, ""

    upper = text.upper()
    for mk in _ABSENCE_MARKERS:
        if mk in text or mk in upper:
            if "반" in text:
                return 0.0, 0.5, None, "absence"
            return 0.0, 1.0, None, "absence"

    for mk in _LEAVE_MARKERS:
        if mk in text:
            if "반" in text:
                return 0.5, 0.0, None, "leave"
            return 1.0, 0.0, None, "leave"

    try:
        n = float(text)
        if n > 0:
            return n, 0.0, None, "leave"
    except ValueError:
        pass
    return 0.0, 0.0, None, ""


def load_invoice_leave_sheet(
    wb: Workbook,
    period_label: str | None = None,
) -> tuple[dict[str, InvoiceLeaveRow], str | None]:
    """
  연차 시트에서 성명별 연차·결근 일수 및 메모를 읽습니다.

    Returns:
        ({name_key: InvoiceLeaveRow}, period YYYY-MM)
    """
    ws = find_leave_worksheet(wb)
    period = period_label or infer_period_from_workbook(wb)
    if ws is None:
        return {}, period

    layout = _scan_leave_sheet_layout(ws)
    if not layout:
        return {}, period

    header_row, name_col, day_cols = layout
    month = 1
    if period and "-" in period:
        try:
            month = int(period.split("-")[1])
        except ValueError:
            pass

    out: dict[str, InvoiceLeaveRow] = {}
    for row in range(header_row + 1, ws.max_row + 1):
        name = _cell_str(ws.cell(row, name_col).value)
        if not _is_valid_name(name):
            continue

        leave_total = 0.0
        unpaid_total = 0.0
        leave_dates: list[int] = []
        absence_dates: list[int] = []

        for day_idx, col in enumerate(day_cols, start=1):
            val = ws.cell(row, col).value
            ld, ud, _, kind = _classify_cell(val)
            if ld <= 0 and ud <= 0:
                continue
            day_num = day_idx if day_idx <= 31 else day_idx
            if kind == "absence":
                unpaid_total += ud if ud > 0 else ld
                absence_dates.append(day_num)
            else:
                leave_total += ld
                leave_dates.append(day_num)

        if leave_total <= 0 and unpaid_total <= 0:
            continue

        key = norm_name_key(name)
        leave_memo = ""
        if leave_total > 0 and period:
            leave_memo = format_usage_date_memos(period, leave_total)
            if leave_dates and leave_total == len(leave_dates):
                parts = [f"{month:02d}월 {d:02d}일 사용" for d in sorted(set(leave_dates))]
                leave_memo = ", ".join(parts)
        absence_memo = ""
        if unpaid_total > 0 and period:
            absence_memo = format_absence_usage_memo(period, unpaid_total)

        display_parts: list[str] = []
        if leave_total > 0:
            display_parts.append(leave_memo or f"{leave_total:g}일")
        if unpaid_total > 0:
            display_parts.append(absence_memo or f"결근/무급 {unpaid_total:g}일")

        out[key] = InvoiceLeaveRow(
            name=name,
            name_key=key,
            leave_days=leave_total,
            unpaid_days=unpaid_total,
            leave_dates=leave_dates,
            absence_dates=absence_dates,
            leave_memo=leave_memo,
            absence_memo=absence_memo,
            display=" / ".join(display_parts),
        )
    return out, period


def load_invoice_leave_sheet_from_path(
    invoice_path: Path,
    period_label: str | None = None,
) -> tuple[dict[str, InvoiceLeaveRow], str | None]:
    if not invoice_path.is_file():
        return {}, period_label
    wb = openpyxl.load_workbook(invoice_path, data_only=True)
    try:
        return load_invoice_leave_sheet(wb, period_label)
    finally:
        wb.close()


def apply_leave_sheet_to_invoice_rows(
    invoice_rows: list[dict[str, Any]],
    leave_by_key: dict[str, InvoiceLeaveRow],
    *,
    prefer_sheet: bool = True,
) -> None:
    """청구서 연차 시트 값을 청구내역 행에 병합."""
    if not leave_by_key:
        return

    for inv in invoice_rows:
        key = norm_name_key(inv.get("name"))
        if not key:
            continue
        row = leave_by_key.get(key)
        if row is None:
            fuzzy = find_fuzzy_name_key(key, leave_by_key.keys())
            if fuzzy:
                row = leave_by_key[fuzzy]
        if row is None:
            continue

        if prefer_sheet or safe_number(inv.get("leave_days"), 0) <= 0:
            if row.leave_days > 0:
                inv["leave_days"] = row.leave_days
        if prefer_sheet or safe_number(inv.get("unpaid_days"), 0) <= 0:
            if row.unpaid_days > 0:
                inv["unpaid_days"] = row.unpaid_days

        inv["leave_sheet_leave_days"] = row.leave_days
        inv["leave_sheet_unpaid_days"] = row.unpaid_days
        inv["leave_sheet_memo"] = row.leave_memo
        inv["leave_sheet_absence_memo"] = row.absence_memo
        inv["leave_sheet_leave_dates"] = list(row.leave_dates)
        inv["leave_sheet_absence_dates"] = list(row.absence_dates)
        if row.display:
            inv["leave_usage_display"] = row.display
        inv["_from_leave_sheet"] = True


def leave_rows_to_monthly_summaries(
    leave_by_key: dict[str, InvoiceLeaveRow],
    period: str,
    roster: dict[str, dict[str, Any]] | None = None,
) -> list:
    """연차대장 월별현황 시트 양식(MonthlyLeaveSummary)으로 변환."""
    from roster_workbook import MonthlyLeaveSummary

    summaries = []
    for row in leave_by_key.values():
        emp_no = ""
        accrued = used_total = remaining = 0.0
        if roster:
            from annual_leave_manager import _match_roster_record

            emp = _match_roster_record(row.name, roster)
            if emp:
                emp_no = emp.get("사번") or ""
                accrued = safe_number(emp.get("발생연차"), 0)
                used_total = safe_number(emp.get("사용연차"), 0)
                remaining = safe_number(emp.get("잔여연차"), 0)

        summaries.append(
            MonthlyLeaveSummary(
                name=row.name,
                emp_no=emp_no,
                period_label=period,
                accrued=accrued,
                month_leave_used=row.leave_days,
                used_total=used_total,
                remaining=remaining,
                absence_days=row.unpaid_days,
                absence_count=calc_absence_occurrence_count(row.unpaid_days),
                leave_memo=row.leave_memo,
                absence_memo=row.absence_memo,
            )
        )
    return sorted(summaries, key=lambda s: s.name)
