"""
services/monthly_leave_manager.py - 월별 연차 소진·결근 현황 (스냅샷 ↔ 연차대장)
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from annual_leave_manager import calc_absence_occurrence_count, format_absence_usage_memo
from core.config import APP_CONFIG
from leave_usage_ledger import (
    MONTHLY_SHEET_NAME,
    SHEET_NAME,
    get_leave_usage_ledger_path,
    normalize_period_label,
    save_leave_usage_ledger_entries,
)
from payroll_archive import format_period_display
from leave_calendar_marks import (
    aggregate_leave_usage_rows,
    build_marks_from_row,
    format_usage_dates_summary,
)
from roster_constants import norm_name_key
from roster_workbook import (
    LEAVE_LEDGER_ALIASES,
    MONTHLY_LEAVE_ALIASES,
    MonthlyLeaveSummary,
    build_header_map,
)
from services.org_registry import filter_records, OrgSelection


def _split_usage_display(display: str) -> tuple[str, str]:
    text = (display or "").strip()
    if not text:
        return "", ""
    if " / " in text:
        parts = [p.strip() for p in text.split(" / ", 1)]
        leave_part = parts[0]
        abs_part = parts[1] if len(parts) > 1 else ""
        if "결근" in abs_part or "무급" in abs_part:
            return leave_part, abs_part
        return leave_part, ""
    if "결근" in text or "무급" in text:
        return "", text
    return text, ""


def _build_summaries_for_sync(
    period: str,
    records: list[dict[str, Any]],
    *,
    invoice_path: Path | None = None,
) -> list[MonthlyLeaveSummary]:
    """연차대장 월별현황 행 — 청구서 연차 시트 우선, 없으면 스냅샷."""
    period_norm = normalize_period_label(period) or period

    if invoice_path and invoice_path.is_file():
        from invoice_leave_sheet import leave_rows_to_monthly_summaries, load_invoice_leave_sheet_from_path
        from payroll_builder import get_templates_roster_path
        from roster_workbook import load_employee_roster_from_workbook

        leave_by_key, sheet_period = load_invoice_leave_sheet_from_path(invoice_path, period_norm)
        if leave_by_key:
            roster_path = get_templates_roster_path()
            roster = (
                load_employee_roster_from_workbook(roster_path)
                if roster_path and roster_path.is_file()
                else {}
            )
            return leave_rows_to_monthly_summaries(
                leave_by_key,
                sheet_period or period_norm,
                roster,
            )

    return build_summaries_from_records(period_norm, records)


def build_summaries_from_records(period: str, records: list[dict[str, Any]]) -> list[MonthlyLeaveSummary]:
    """급여 스냅샷 → 월별현황 행 (1인 1행)."""
    period_norm = normalize_period_label(period) or period
    out: list[MonthlyLeaveSummary] = []
    for r in records:
        name = str(r.get("name") or "").strip()
        if not name:
            continue
        leave_days = float(r.get("leave_days") or 0)
        unpaid = float(r.get("unpaid_days") or 0)
        if r.get("leave_sheet_leave_days") is not None:
            leave_days = float(r.get("leave_sheet_leave_days") or leave_days)
        if r.get("leave_sheet_unpaid_days") is not None:
            unpaid = float(r.get("leave_sheet_unpaid_days") or unpaid)
        display = str(r.get("leave_usage_display") or "").strip()
        sheet_leave = str(r.get("leave_sheet_memo") or "").strip()
        sheet_absence = str(r.get("leave_sheet_absence_memo") or "").strip()
        leave_memo, absence_memo = _split_usage_display(display)
        if sheet_leave:
            leave_memo = sheet_leave
        elif leave_days > 0 and not leave_memo:
            leave_memo = display or f"{leave_days:g}일"
        if sheet_absence:
            absence_memo = sheet_absence
        elif unpaid > 0 and not absence_memo:
            absence_memo = format_absence_usage_memo(period_norm, unpaid)

        accrued = float(r.get("accrued_leave") or r.get("annual_leave_accrued") or 0)
        used_total = float(r.get("used_leave") or r.get("annual_leave_used") or 0)
        remaining = float(r.get("remaining_leave") or r.get("annual_leave_remaining") or 0)

        out.append(
            MonthlyLeaveSummary(
                name=name,
                emp_no=r.get("emp_no") or r.get("사번") or "",
                period_label=period_norm,
                accrued=accrued,
                month_leave_used=leave_days,
                used_total=used_total,
                remaining=remaining,
                absence_days=unpaid,
                absence_count=calc_absence_occurrence_count(unpaid),
                leave_memo=leave_memo,
                absence_memo=absence_memo,
            )
        )
    return sorted(out, key=lambda s: (s.name,))


def load_monthly_leave_rows(period: str) -> list[dict[str, Any]]:
    """연차대장 「월별현황」 시트에서 해당 처리월 행 로드."""
    path = get_leave_usage_ledger_path()
    period_norm = normalize_period_label(period) or period
    if not path.is_file():
        return []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if MONTHLY_SHEET_NAME not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[MONTHLY_SHEET_NAME]
    headers = build_header_map(ws, MONTHLY_LEAVE_ALIASES)
    period_col = headers.get("처리월")
    name_col = headers.get("성명")
    if not period_col or not name_col:
        wb.close()
        return []

    rows: list[dict[str, Any]] = []
    for r in range(2, (ws.max_row or 1) + 1):
        nm = ws.cell(r, name_col).value
        if nm is None or str(nm).strip() == "":
            continue
        pv = normalize_period_label(ws.cell(r, period_col).value)
        if pv != period_norm:
            continue
        row: dict[str, Any] = {"name": str(nm).strip()}
        for key in (
            "사번",
            "처리월",
            "발생연차",
            "당월연차",
            "누적사용연차",
            "잔여연차",
            "무급일수",
            "무급횟수",
            "연차내역",
            "무급내역",
            "최종갱신",
        ):
            col = headers.get(key)
            if col:
                row[key] = ws.cell(r, col).value
        rows.append(row)
    wb.close()
    return sorted(rows, key=lambda x: str(x.get("name") or ""))


def load_ledger_detail_rows(period: str) -> list[dict[str, Any]]:
    """연차사용대장 상세 시트에서 해당 처리월 행 로드 (사용 건별 1행)."""
    path = get_leave_usage_ledger_path()
    period_norm = normalize_period_label(period) or period
    if not path.is_file():
        return []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[SHEET_NAME]
    headers = build_header_map(ws, LEAVE_LEDGER_ALIASES)
    period_col = headers.get("처리월") or headers.get("사용월")
    name_col = headers.get("성명")
    if not period_col or not name_col:
        wb.close()
        return []

    rows: list[dict[str, Any]] = []
    for r in range(2, (ws.max_row or 1) + 1):
        nm = ws.cell(r, name_col).value
        if nm is None or str(nm).strip() == "":
            continue
        pv = normalize_period_label(ws.cell(r, period_col).value)
        if pv != period_norm:
            continue
        kind = str(ws.cell(r, headers.get("구분", 0)).value or "연차").strip() if headers.get("구분") else "연차"
        days_col = headers.get("사용일수")
        days = float(ws.cell(r, days_col).value or 0) if days_col else 0.0
        memo_col = headers.get("사용내역")
        memo = str(ws.cell(r, memo_col).value or "").strip() if memo_col else ""
        is_absence = kind == "결근" or ":결근:" in memo or "결근/무급" in memo
        row: dict[str, Any] = {
            "name": str(nm).strip(),
            "emp_no": ws.cell(r, headers["사번"]).value if headers.get("사번") else "",
            "month_leave": 0.0 if is_absence else days,
            "absence_days": days if is_absence else 0.0,
            "leave_memo": "" if is_absence else memo,
            "absence_memo": memo if is_absence else "",
            "source": "대장상세",
        }
        for key in ("발생연차", "누적사용연차", "잔여연차", "무급횟수"):
            col = headers.get(key)
            if col:
                val = ws.cell(r, col).value
                field_map = {
                    "발생연차": "accrued",
                    "누적사용연차": "used_total",
                    "잔여연차": "remaining",
                    "무급횟수": "absence_count",
                }
                row[field_map[key]] = val
        rows.append(row)
    wb.close()
    return rows


def _upsert_leave_view(
    by_key: dict[str, dict[str, Any]],
    row: dict[str, Any],
    *,
    prefer: str = "",
) -> None:
    """직원별 연차·결근 행 병합 (norm_name_key 기준)."""
    name = str(row.get("name") or row.get("성명") or "").strip()
    if not name:
        return
    key = norm_name_key(name)
    if not key:
        return

    if key not in by_key:
        entry = dict(row)
        entry["name"] = name
        if prefer:
            entry["source"] = prefer
        by_key[key] = entry
        return

    existing = by_key[key]
    for field in ("month_leave", "absence_days"):
        ev = float(existing.get(field) or 0)
        nv = float(row.get(field) or 0)
        if prefer == "급여" and existing.get("source") == "급여":
            existing[field] = nv if nv > 0 else ev
        elif prefer == "대장" and existing.get("source") in ("", "대장", "대장상세", "병합"):
            if nv > 0 and ev <= 0:
                existing[field] = nv
            elif nv > 0 and ev > 0 and abs(nv - ev) > 1e-9:
                existing[field] = max(ev, nv)
        else:
            existing[field] = ev + nv

    for memo_field in ("leave_memo", "absence_memo"):
        em = str(existing.get(memo_field) or "").strip()
        nm = str(row.get(memo_field) or "").strip()
        if nm and (not em or prefer == "급여"):
            if nm not in em:
                existing[memo_field] = f"{em}, {nm}".strip(", ") if em else nm
        elif nm and nm not in em:
            existing[memo_field] = f"{em}, {nm}".strip(", ") if em else nm

    for date_field in ("leave_sheet_leave_dates", "leave_sheet_absence_dates"):
        if row.get(date_field):
            ed = existing.get(date_field) or []
            nd = row.get(date_field) or []
            existing[date_field] = sorted({int(d) for d in list(ed) + list(nd) if d is not None})

    if prefer == "급여" or not existing.get("workplace"):
        existing["workplace"] = row.get("workplace") or existing.get("workplace") or ""
        existing["affiliate"] = row.get("affiliate") or existing.get("affiliate") or ""
    existing["emp_no"] = existing.get("emp_no") or row.get("emp_no") or row.get("사번") or ""

    for field in ("accrued", "used_total", "remaining"):
        nv = row.get(field)
        if nv not in (None, "", 0, 0.0):
            if prefer == "급여" or not existing.get(field):
                existing[field] = nv

    ac = row.get("absence_count")
    if ac:
        existing["absence_count"] = int(existing.get("absence_count") or 0) + int(ac)

    if prefer:
        if existing.get("source") == "급여" and prefer == "대장":
            existing["source"] = "병합"
        elif not existing.get("source"):
            existing["source"] = prefer


def merge_leave_views(
    period: str,
    records: list[dict[str, Any]],
    ledger_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """스냅샷(최신 급여) + 연차대장(잔여·누적) 병합."""
    period_norm = normalize_period_label(period) or period
    by_key: dict[str, dict[str, Any]] = {}

    payroll_rows: list[dict[str, Any]] = []
    for r in records:
        name = str(r.get("name") or "").strip()
        if not name:
            continue
        leave_days = float(r.get("leave_days") or 0)
        unpaid = float(r.get("unpaid_days") or 0)
        if r.get("leave_sheet_leave_days") is not None:
            leave_days = float(r.get("leave_sheet_leave_days") or leave_days)
        if r.get("leave_sheet_unpaid_days") is not None:
            unpaid = float(r.get("leave_sheet_unpaid_days") or unpaid)
        display = str(r.get("leave_usage_display") or "").strip()
        sheet_leave = str(r.get("leave_sheet_memo") or "").strip()
        sheet_absence = str(r.get("leave_sheet_absence_memo") or "").strip()
        leave_memo, absence_memo = _split_usage_display(display)
        if sheet_leave:
            leave_memo = sheet_leave
        elif leave_days > 0 and not leave_memo:
            leave_memo = display or f"{leave_days:g}일"
        if sheet_absence:
            absence_memo = sheet_absence
        elif unpaid > 0 and not absence_memo:
            absence_memo = format_absence_usage_memo(period_norm, unpaid)
        payroll_rows.append(
            {
                "name": name,
                "affiliate": r.get("affiliate", ""),
                "workplace": r.get("workplace", "") or r.get("dept", ""),
                "emp_no": r.get("emp_no") or "",
                "month_leave": leave_days,
                "absence_days": unpaid,
                "absence_count": calc_absence_occurrence_count(unpaid),
                "leave_memo": leave_memo,
                "absence_memo": absence_memo,
                "leave_sheet_leave_dates": r.get("leave_sheet_leave_dates"),
                "leave_sheet_absence_dates": r.get("leave_sheet_absence_dates"),
                "accrued": float(r.get("accrued_leave") or 0),
                "used_total": float(r.get("used_leave") or 0),
                "remaining": float(r.get("remaining_leave") or 0),
            }
        )

    ledger_std: list[dict[str, Any]] = []
    for row in ledger_rows:
        name = str(row.get("name") or row.get("성명") or "").strip()
        if not name:
            continue
        ledger_std.append(
            {
                "name": name,
                "emp_no": row.get("사번") or "",
                "month_leave": float(row.get("당월연차") or 0),
                "absence_days": float(row.get("무급일수") or 0),
                "absence_count": int(row.get("무급횟수") or 0),
                "leave_memo": str(row.get("연차내역") or ""),
                "absence_memo": str(row.get("무급내역") or ""),
                "accrued": float(row.get("발생연차") or 0),
                "used_total": float(row.get("누적사용연차") or 0),
                "remaining": float(row.get("잔여연차") or 0),
            }
        )

    for r in aggregate_leave_usage_rows(payroll_rows, period_norm):
        _upsert_leave_view(by_key, r, prefer="급여")
    for r in aggregate_leave_usage_rows(ledger_std, period_norm):
        _upsert_leave_view(by_key, r, prefer="대장")

    return aggregate_leave_usage_rows(list(by_key.values()), period_norm)


def sync_monthly_leave_to_ledger(
    period: str,
    records: list[dict[str, Any]],
    *,
    scopes: list | None = None,
    invoice_path: Path | None = None,
) -> dict[str, Any]:
    """청구서 연차 시트(우선) + 스냅샷으로 연차대장 월별현황 갱신."""
    from services.payroll_scope import PayrollScope, resolve_output_dir

    period_norm = normalize_period_label(period) or period
    summaries = _build_summaries_for_sync(period_norm, records, invoice_path=invoice_path)
    info = save_leave_usage_ledger_entries([], period_norm, monthly_summaries=summaries)

    exported: list[str] = []
    scope_list = scopes or []
    if not scope_list:
        seen: set[tuple[str, str]] = set()
        from core.org_config import canonical_scope_workplace

        for r in records:
            aff = str(r.get("affiliate") or "").strip()
            wp = canonical_scope_workplace(
                str(r.get("_scope_workplace") or r.get("workplace") or "").strip()
            )
            if aff and wp and (aff, wp) not in seen:
                seen.add((aff, wp))
                scope_list.append(PayrollScope(aff, wp, period_norm))

    for scope in scope_list:
        if not isinstance(scope, PayrollScope):
            continue
        if scope.period != period_norm:
            continue
        from core.org_config import scope_workplaces_match

        scope_records = [
            r
            for r in records
            if str(r.get("affiliate") or "").strip() == scope.affiliate
            and scope_workplaces_match(
                scope.workplace,
                str(r.get("_scope_workplace") or r.get("workplace") or ""),
            )
        ]
        if not scope_records:
            continue
        path = export_scope_leave_workbook(scope, period_norm, scope_records)
        exported.append(str(path))

    info["exported"] = exported
    info["summaries"] = len(summaries)
    return info


def export_scope_leave_workbook(
    scope,
    period: str,
    records: list[dict[str, Any]],
) -> Path:
    """사업장·월 폴더에 연차·결근 현황 Excel 저장."""
    from services.payroll_scope import resolve_output_dir

    out_dir = resolve_output_dir(scope)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"{period}_연차결근현황.xlsx"

    brand = APP_CONFIG.brand
    navy = brand.primary_navy.lstrip("#")
    fill_hdr = PatternFill("solid", fgColor=navy)
    font_hdr = Font(name="맑은 고딕", bold=True, color="FFFFFF")
    font_title = Font(name="맑은 고딕", size=12, bold=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "연차결근"
    ws["A1"] = f"{format_period_display(period)} — {scope.workplace} 연차·결근 현황"
    ws["A1"].font = font_title
    ws.merge_cells("A1:K1")

    # 연차사용대장 「월별현황」 시트와 동일한 열 구성
    headers = [
        "성명",
        "사번",
        "처리월",
        "발생연차",
        "당월연차",
        "누적사용연차",
        "잔여연차",
        "무급일수",
        "무급횟수",
        "연차내역",
        "무급내역",
    ]
    row = 3
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.fill = fill_hdr
        cell.font = font_hdr
        cell.alignment = Alignment(horizontal="center")

    summaries = build_summaries_from_records(period, records)
    for s in summaries:
        row += 1
        ws.cell(row, 1, s.name)
        ws.cell(row, 2, s.emp_no)
        ws.cell(row, 3, period)
        ws.cell(row, 4, s.accrued)
        ws.cell(row, 5, s.month_leave_used)
        ws.cell(row, 6, s.used_total)
        ws.cell(row, 7, s.remaining)
        ws.cell(row, 8, s.absence_days)
        ws.cell(row, 9, s.absence_count)
        ws.cell(row, 10, s.leave_memo)
        ws.cell(row, 11, s.absence_memo)

    for col, w in zip("ABCDEFGHIJK", [10, 10, 12, 10, 10, 10, 10, 10, 10, 28, 28]):
        ws.column_dimensions[col].width = w

    from core.file_save import save_workbook

    save_workbook(wb, path)
    wb.close()
    return path


def scope_leave_export_path(scope, period: str) -> Path:
    from services.payroll_scope import resolve_output_dir

    return resolve_output_dir(scope) / f"{period}_연차결근현황.xlsx"


def filtered_leave_rows(
    period: str,
    records: list[dict[str, Any]],
    selection: OrgSelection | None = None,
    *,
    only_with_usage: bool = False,
) -> list[dict[str, Any]]:
    filtered = filter_records(records, selection or OrgSelection())
    ledger_rows = load_monthly_leave_rows(period)
    detail_rows = load_ledger_detail_rows(period)
    merged = merge_leave_views(period, filtered, ledger_rows)

    if detail_rows:
        detail_agg = aggregate_leave_usage_rows(detail_rows, period)
        detail_by_key = {norm_name_key(r.get("name", "")): r for r in detail_agg}
        for row in merged:
            key = norm_name_key(row.get("name", ""))
            detail = detail_by_key.get(key)
            if not detail:
                continue
            if float(row.get("month_leave") or 0) <= 0 and float(detail.get("month_leave") or 0) > 0:
                row["month_leave"] = detail["month_leave"]
            if float(row.get("absence_days") or 0) <= 0 and float(detail.get("absence_days") or 0) > 0:
                row["absence_days"] = detail["absence_days"]
            if not row.get("leave_memo") and detail.get("leave_memo"):
                row["leave_memo"] = detail["leave_memo"]
            if not row.get("absence_memo") and detail.get("absence_memo"):
                row["absence_memo"] = detail["absence_memo"]
            if not row.get("remaining") and detail.get("remaining"):
                row["remaining"] = detail["remaining"]
            marks = build_marks_from_row(row, period)
            row["dates_summary"] = format_usage_dates_summary(period, marks)

        merged_keys = {norm_name_key(r.get("name", "")) for r in merged}
        for key, detail in detail_by_key.items():
            if key and key not in merged_keys:
                merged.append(detail)

    merged = aggregate_leave_usage_rows(merged, period)

    if only_with_usage:
        merged = [
            r
            for r in merged
            if float(r.get("month_leave") or 0) > 0
            or float(r.get("absence_days") or 0) > 0
            or r.get("leave_memo")
            or r.get("absence_memo")
        ]
    return merged
