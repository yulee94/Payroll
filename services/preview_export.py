"""
services/preview_export.py - 파일 미리보기 · Excel 내려받기 (그리드 레이아웃)
"""

from __future__ import annotations

import csv
import json
import re
import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

MAX_PREVIEW_ROWS = 120
MAX_PREVIEW_COLS = 64

# 같은 파일을 UI에서 여러 번 클릭/전환할 때 openpyxl 재로딩이 누적되어 렉이 발생할 수 있어
# (경로, mtime, sheet) 기준으로 간단 캐시를 둡니다.
_PREVIEW_CACHE: dict[tuple[str, float, str], PreviewData] = {}
_PREVIEW_CACHE_MAX = 32

RowRole = Literal["title", "subtitle", "section", "header", "data", "empty"]


@dataclass
class GridRow:
    """Excel 한 행 — 역할(제목/헤더/데이터)과 정렬 정보 포함."""

    index: int
    role: RowRole
    cells: list[str]
    aligns: list[str]  # w | e | center


@dataclass
class PreviewData:
    title: str
    kind: str  # excel | csv | json | text | unsupported
    headers: list[str]
    rows: list[list[Any]]
    text: str = ""
    truncated: bool = False
    truncated_cols: bool = False
    grid_rows: list[GridRow] = field(default_factory=list)
    column_count: int = 0
    column_letters: list[str] = field(default_factory=list)
    sheet_names: list[str] = field(default_factory=list)
    active_sheet: str = ""


def copy_file_to(dest: Path, source: Path) -> Path:
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, dest)
    return dest


def _is_number(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    s = str(value).replace(",", "").replace(" ", "").strip()
    if not s or s in ("-", "—"):
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def _format_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "예" if value else "아니오"
    if isinstance(value, float):
        if value == int(value):
            return f"{int(value):,}"
        text = f"{value:,.2f}".rstrip("0").rstrip(".")
        return text
    if isinstance(value, int):
        return f"{value:,}"
    text = str(value).strip()
    if _is_number(text):
        s = text.replace(",", "")
        if "." in s:
            f = float(s)
            if f == int(f):
                return f"{int(f):,}"
            return f"{f:,.2f}".rstrip("0").rstrip(".")
        return f"{int(float(s)):,}"
    return text


def _non_empty_count(row: list[str]) -> int:
    return sum(1 for c in row if c.strip())


def _row_has_numbers(row: list[str]) -> int:
    return sum(1 for c in row if c.strip() and _is_number(c))


def _classify_rows(raw_rows: list[list[str]]) -> list[RowRole]:
    """보고서·대장 양식의 제목·섹션·헤더·데이터 행을 추정."""
    roles: list[RowRole] = []
    n = len(raw_rows)

    for i, row in enumerate(raw_rows):
        filled = _non_empty_count(row)
        texts = [c for c in row if c.strip()]

        if filled == 0:
            roles.append("empty")
            continue

        joined = " ".join(texts)

        if filled <= 3 and re.match(r"^\d+\.\s", texts[0]):
            roles.append("section")
            continue

        if filled <= 3 and texts[0].startswith(("■", "▶", "[", "※")):
            roles.append("section")
            continue

        if filled == 1 and (len(texts[0]) >= 8 or len(row) >= 4):
            roles.append("title")
            continue

        if filled <= 2 and any(k in joined for k in ("보고", "요약", "작성", "대장", "현황")):
            roles.append("title" if i == 0 or roles[-1] == "empty" else "subtitle")
            continue

        # 다음 행에 숫자가 많으면 현재 행을 헤더로
        if i + 1 < n:
            next_nums = _row_has_numbers(raw_rows[i + 1])
            next_filled = _non_empty_count(raw_rows[i + 1])
            cur_nums = _row_has_numbers(row)
            if filled >= 2 and next_nums >= 1 and cur_nums <= 1 and next_filled >= 2:
                if all(len(t) <= 18 for t in texts):
                    roles.append("header")
                    continue

        # 직전이 section/title 이고 짧은 텍스트만 있으면 헤더
        if i > 0 and roles[i - 1] in ("section", "title", "subtitle", "empty"):
            if filled >= 2 and _row_has_numbers(row) == 0 and all(len(t) <= 20 for t in texts):
                roles.append("header")
                continue

        roles.append("data")

    return roles


def _cell_align(value: str, role: RowRole) -> str:
    if role in ("title", "subtitle", "section"):
        return "w"
    if role == "header":
        return "center"
    if value.strip() and _is_number(value):
        return "e"
    return "w"


def _trim_trailing_empty_cols(rows: list[list[str]]) -> tuple[list[list[str]], int]:
    if not rows:
        return rows, 0
    last_col = 0
    for row in rows:
        for i, cell in enumerate(row):
            if cell.strip():
                last_col = max(last_col, i + 1)
    last_col = min(max(last_col, 1), MAX_PREVIEW_COLS)
    trimmed = [row[:last_col] + [""] * (last_col - len(row)) for row in rows]
    return trimmed, last_col


def _build_grid(raw_rows: list[list[str]], start_index: int = 1) -> tuple[list[GridRow], int]:
    trimmed, col_count = _trim_trailing_empty_cols(raw_rows)
    roles = _classify_rows(trimmed)
    grid: list[GridRow] = []

    for i, (row, role) in enumerate(zip(trimmed, roles)):
        cells = [_format_cell(v) for v in row]
        aligns = [_cell_align(c, role) for c in cells]
        grid.append(GridRow(index=start_index + i, role=role, cells=cells, aligns=aligns))

    return grid, col_count


def preview_excel(path: Path, sheet_name: str | None = None) -> PreviewData:
    try:
        mtime = path.stat().st_mtime
    except OSError:
        mtime = 0.0
    cache_key = (str(path), float(mtime), sheet_name or "")
    cached = _PREVIEW_CACHE.get(cache_key)
    if cached is not None:
        return cached

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = list(wb.sheetnames)
    ws = wb[sheet_name] if sheet_name and sheet_name in sheet_names else wb.active
    active = ws.title

    raw_rows: list[list[str]] = []
    truncated = False
    truncated_cols = False
    max_row = ws.max_row or 1
    sheet_cols = ws.max_column or 1
    if sheet_cols > MAX_PREVIEW_COLS:
        truncated_cols = True
    max_col = min(sheet_cols, MAX_PREVIEW_COLS)
    if max_row > MAX_PREVIEW_ROWS:
        truncated = True
    limit = min(max_row, MAX_PREVIEW_ROWS)

    for row in ws.iter_rows(min_row=1, max_row=limit, max_col=max_col, values_only=True):
        raw_rows.append([_format_cell(v) for v in row])

    wb.close()

    grid, col_count = _build_grid(raw_rows)
    letters = [get_column_letter(i + 1) for i in range(col_count)]

    # 레거시 Treeview 호환 (헤더 행 자동 탐색)
    header_idx = next((i for i, g in enumerate(grid) if g.role == "header"), None)
    if header_idx is not None:
        headers = grid[header_idx].cells
        body = [g.cells for g in grid[header_idx + 1 :] if g.role == "data"]
    elif grid and grid[0].role == "data":
        headers = [f"열{i + 1}" for i in range(col_count)]
        body = [g.cells for g in grid if g.role == "data"]
    else:
        headers = letters
        body = [g.cells for g in grid if g.role in ("data", "header")]

    data = PreviewData(
        title=f"{path.name} — {active}",
        kind="excel",
        headers=headers[:col_count],
        rows=body,
        truncated=truncated or truncated_cols,
        truncated_cols=truncated_cols,
        grid_rows=grid,
        column_count=col_count,
        column_letters=letters,
        sheet_names=sheet_names,
        active_sheet=active,
    )
    _cache_set(cache_key, data)
    return data


def _cache_set(key: tuple[str, float, str], value: PreviewData) -> None:
    _PREVIEW_CACHE[key] = value
    if len(_PREVIEW_CACHE) <= _PREVIEW_CACHE_MAX:
        return
    # 오래된 항목을 대충 제거 (dict insertion order 유지 가정)
    for k in list(_PREVIEW_CACHE.keys())[: max(1, len(_PREVIEW_CACHE) - _PREVIEW_CACHE_MAX)]:
        _PREVIEW_CACHE.pop(k, None)


def preview_csv(path: Path) -> PreviewData:
    rows_raw: list[list[str]] = []
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i >= MAX_PREVIEW_ROWS:
                break
            rows_raw.append(row)

    formatted = [[_format_cell(c) for c in row] for row in rows_raw]
    grid, col_count = _build_grid(formatted)
    if grid and grid[0].role != "header":
        grid[0].role = "header"
        grid[0].aligns = ["center"] * col_count

    headers = grid[0].cells if grid else []
    body = [g.cells for g in grid[1:] if g.role == "data"]

    return PreviewData(
        title=path.name,
        kind="csv",
        headers=headers,
        rows=body,
        truncated=len(rows_raw) >= MAX_PREVIEW_ROWS,
        grid_rows=grid,
        column_count=col_count,
        column_letters=[get_column_letter(i + 1) for i in range(col_count)],
    )


def preview_json(path: Path) -> PreviewData:
    text = path.read_text(encoding="utf-8")
    try:
        obj = json.loads(text)
        pretty = json.dumps(obj, ensure_ascii=False, indent=2)
    except json.JSONDecodeError:
        pretty = text
    if len(pretty) > 12000:
        pretty = pretty[:12000] + "\n\n… (미리보기 일부만 표시)"
    return PreviewData(title=path.name, kind="json", headers=[], rows=[], text=pretty)


def preview_file(path: Path, sheet_name: str | None = None) -> PreviewData:
    if not path.exists():
        return PreviewData(
            title=str(path),
            kind="unsupported",
            headers=[],
            rows=[],
            text="파일을 찾을 수 없습니다.",
        )
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return preview_excel(path, sheet_name=sheet_name)
    if suffix == ".csv":
        return preview_csv(path)
    if suffix == ".json":
        return preview_json(path)
    return PreviewData(
        title=path.name,
        kind="unsupported",
        headers=[],
        rows=[],
        text=f"'{suffix}' 형식은 미리보기를 지원하지 않습니다.\n「열기」로 Excel/뷰어를 사용하세요.",
    )
