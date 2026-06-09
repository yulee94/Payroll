"""
invoice_parser.py - 도급비 청구서(26년05월 형식) 데이터 추출

- 1번 시트(청구내역): 급여·보험 등
- 2번 시트(근태): 지조외 = 조퇴시간(시간)
"""

from __future__ import annotations

from dataclasses import dataclass
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from roster_constants import find_fuzzy_name_key, norm_name_key
from utils import safe_number

from core.payroll.site_benefits import find_workers_day_column

DATA_START_ROW = 5

# 실제 청구서 열 번호 (1-based)
COL = {
    "no": 1,           # A
    "dept": 2,         # B 소속
    "name": 3,         # C 성명
    "hire": 4,         # D 입사일
    "base_days": 9,    # I 기준일
    "work_days": 10,   # J 근무일
    "unpaid": 11,      # K 무급/결근
    "leave": 12,       # L 휴가/연차
    "ot_hours": 13,    # M O/T(150%)
    "shift_hours": 14, # N 교대
    "night_hours": 15, # O 심야(50%)
    "special_hours": 16,   # P 특근(150%)
    "special_ext_hours": 17,  # Q 특근연장(50%)
    "early_leave": 18, # R 지조외 (청구 시트 — 근태 시트가 있으면 근태 우선)
    "base_hourly": 7,  # G
    "ordinary_hourly": 8,  # H
    "base_salary": 19,     # S 기본급
    "base_deduction": 20,  # T 기본공제
    "ot_pay": 22,          # V O/T수당
    "night_pay": 23,       # W 심야수당
    "special_pay": 24,     # X 특근수당
    "special_ext_pay": 25, # Y 특근연장
    "position_pay": 26,    # Z 직책수당
    "shift_pay": 27,       # AA 교대수당
    "subtotal": 28,        # AB 소계
    "annual_pay": 31,      # AE (업무추진비 등 — 소계 포함, 미사용)
    "transport": 47,       # AU 교통비
    "health": 35,          # AI 건강보험
    "long_term_care": 36,  # AJ 장기요양
    "pension": 37,         # AK 국민연금
    "employment": 39,      # AM 고용보험
}


@dataclass(frozen=True)
class BillingColumnLayout:
    """청구내역 시트의 헤더 기반 열 배치."""

    header_row: int
    data_start_row: int
    columns: dict[str, int]
    source: str


def _compact_header(value: Any) -> str:
    """
    헤더 비교용 정규화.

    업로드 양식마다 공백·개행·괄호·슬래시 표기가 달라서, 의미 키워드만 남겨
    `O/T(150%)`, `OT 150`, `ot_150` 같은 표기를 같은 후보로 봅니다.
    """
    text = str(value or "").strip().lower()
    text = text.replace("％", "%")
    return re.sub(r"[\s\r\n\t_/\\()\[\]{}·.,:;%\-]+", "", text)


INVOICE_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "no": ("순번", "번호", "no", "num"),
    "dept": ("소속", "부서", "팀", "부문", "직무", "업무", "dept", "department"),
    "name": ("성명", "이름", "사원명", "직원명", "근로자명", "성명(한글)", "name", "employee"),
    "hire": ("입사일", "입사일자", "채용일", "hiredate", "joindate"),
    "base_hourly": ("기본시급", "기본 시급", "시급", "basehourly", "basepayrate"),
    "ordinary_hourly": ("통상시급", "통상 시급", "ordinaryhourly", "regularhourly"),
    "base_days": ("기준일", "기준일수", "기준시간", "월기준시간", "산정일", "산정시간"),
    "work_days": ("근무일", "근무일수", "근무시간", "실근무", "실근무시간", "출근일수", "workdays", "workhours"),
    "unpaid": ("무급/결근", "무급결근", "결근", "결근일", "무급", "무급일", "absence", "unpaid"),
    "leave": ("휴가/연차", "연차사용", "연차", "휴가", "휴가일", "leave"),
    "ot_hours": ("잔업시간", "연장시간", "연장근로시간", "ot시간", "ot150", "overtimehours"),
    "shift_hours": ("교대시간", "교대근무시간", "shift시간", "shifthours"),
    "night_hours": ("심야시간", "야간시간", "심야근로시간", "night시간", "nighthours"),
    "special_hours": ("특근시간", "휴일근로시간", "휴일시간", "specialhours", "holidayworkhours"),
    "special_ext_hours": ("특근연장시간", "휴일연장시간", "specialexthours", "holidayovertimehours"),
    "early_leave": ("지조외", "조퇴", "조퇴시간", "조퇴분", "조퇴누계", "earlyleave"),
    "base_salary": ("기본급", "기본임금", "월기본급", "basesalary"),
    "base_deduction": ("기본공제", "기본급공제", "결근공제", "무급공제", "basededuction"),
    "ot_pay": ("잔업수당", "연장수당", "ot수당", "ot150수당", "overtimepay"),
    "shift_pay": ("교대수당", "shift수당", "shiftpay"),
    "night_pay": ("심야수당", "야간수당", "nightpay"),
    "special_pay": ("특근수당", "휴일근로수당", "specialpay", "holidayworkpay"),
    "special_ext_pay": ("특근연장수당", "휴일연장수당", "specialextpay", "holidayovertimepay"),
    "position_pay": ("직책수당", "직무수당", "고정수당", "positionpay"),
    "annual_pay": ("연차수당", "업무추진비", "상여", "기타수당", "annualpay"),
    "subtotal": ("소계", "지급소계", "과세소계", "급여소계", "subtotal"),
    "gross_pay": ("총지급액", "총지급", "급여총액", "총급여", "지급총액", "grosspay", "totalpay"),
    "transport": ("교통비", "교통수당", "차량유지비", "transport", "transportation"),
    "health": ("건강보험", "건보", "healthinsurance"),
    "long_term_care": ("장기요양", "장기요양보험", "요양보험", "longtermcare"),
    "pension": ("국민연금", "연금", "nationalpension", "pension"),
    "employment": ("고용보험", "employmentinsurance"),
}


_AMOUNT_HINTS = (
    "수당",
    "급",
    "임금",
    "공제",
    "보험",
    "연금",
    "소계",
    "총",
    "pay",
    "salary",
    "deduction",
    "insurance",
)

_TIME_HINTS = ("시간", "일수", "일", "근무", "연차", "휴가", "결근", "무급", "hours", "days")


def _invoice_header_match_score(header: Any, key: str) -> int:
    compact = _compact_header(header)
    if not compact:
        return 0

    aliases = tuple(sorted(INVOICE_HEADER_ALIASES.get(key, ()), key=lambda a: len(_compact_header(a)), reverse=True))
    for alias in aliases:
        alias_compact = _compact_header(alias)
        if not alias_compact:
            continue
        if compact == alias_compact:
            return 100 + len(alias_compact)
        if alias_compact in compact:
            return 60 + len(alias_compact)

    return 0


def _match_invoice_header(header: Any) -> tuple[str, int] | None:
    compact = _compact_header(header)
    if not compact:
        return None

    candidates: list[tuple[int, str]] = []
    for key in INVOICE_HEADER_ALIASES:
        score = _invoice_header_match_score(header, key)
        if score:
            candidates.append((score, key))
    if not candidates:
        return None

    candidates.sort(reverse=True)
    score, key = candidates[0]

    # "기본시급"이 "시급" 때문에 base_hourly로 잡히는 것은 맞지만,
    # "통상시급"은 반드시 ordinary_hourly가 우선되어야 합니다.
    if _invoice_header_match_score(header, "ordinary_hourly") >= score:
        key = "ordinary_hourly"
        score = _invoice_header_match_score(header, key)

    # 시간/일수 계열과 금액 계열이 같이 걸리면 헤더 단어로 한번 더 가릅니다.
    if key.endswith("_hours") or key in {"base_days", "work_days", "unpaid", "leave", "early_leave"}:
        if any(h in compact for h in _AMOUNT_HINTS) and not any(h in compact for h in _TIME_HINTS):
            amount_partner = {
                "ot_hours": "ot_pay",
                "shift_hours": "shift_pay",
                "night_hours": "night_pay",
                "special_hours": "special_pay",
                "special_ext_hours": "special_ext_pay",
            }.get(key)
            if amount_partner and _invoice_header_match_score(header, amount_partner):
                key = amount_partner
    return key, score


def _scan_billing_header_row(ws: Worksheet, row_no: int) -> dict[str, tuple[int, int]]:
    """한 헤더 행에서 {필드: (열, 점수)}를 찾습니다."""
    mapping: dict[str, tuple[int, int]] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row_no, col).value
        matched = _match_invoice_header(raw)
        if not matched:
            continue
        key, score = matched
        prev = mapping.get(key)
        if prev is None or score > prev[1]:
            mapping[key] = (col, score)
    return mapping


def _billing_layout_score(mapping: dict[str, tuple[int, int]]) -> int:
    keys = set(mapping)
    score = len(keys) * 10
    if "name" in keys:
        score += 60
    if "subtotal" in keys or "gross_pay" in keys:
        score += 35
    if "base_salary" in keys:
        score += 20
    if {"health", "pension", "employment"} & keys:
        score += 10
    return score


def find_billing_column_layout(ws: Worksheet) -> BillingColumnLayout:
    """
    청구내역 컬럼을 헤더 키워드로 인식합니다.

    새 양식은 헤더 위치/열 순서가 달라도 키워드가 있으면 동작합니다. 헤더를 충분히
    찾지 못하면 기존 26년05월 고정 열번호로 fallback합니다.
    """
    best: tuple[int, int, dict[str, tuple[int, int]]] | None = None
    for row_no in range(1, min(15, ws.max_row + 1)):
        mapping = _scan_billing_header_row(ws, row_no)
        score = _billing_layout_score(mapping)
        if score <= 0:
            continue
        if best is None or score > best[0]:
            best = (score, row_no, mapping)

    if best is not None:
        score, header_row, mapping = best
        keys = set(mapping)
        if "name" in keys and ("subtotal" in keys or "gross_pay" in keys or "base_salary" in keys) and score >= 90:
            return BillingColumnLayout(
                header_row=header_row,
                data_start_row=header_row + 1,
                columns={key: col for key, (col, _score) in mapping.items()},
                source="keyword",
            )

    return BillingColumnLayout(
        header_row=DATA_START_ROW - 1,
        data_start_row=DATA_START_ROW,
        columns=dict(COL),
        source="legacy-fixed",
    )


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _header_matches_early_leave(header: str) -> bool:
    h = header.replace(" ", "").replace("\n", "")
    if not h:
        return False
    if "지조외" in h:
        return True
    if h in ("조퇴", "조퇴시간", "조퇴누계", "조퇴합계"):
        return True
    if "조퇴" in h and "공제" not in h:
        return True
    return False


def _header_value_is_minutes(header: Any) -> bool:
    compact = _compact_header(header)
    if not compact:
        return False
    if "분" in compact:
        return True
    if "minute" in compact or compact.endswith("min"):
        return True
    return False


def _header_matches_name(header: str) -> bool:
    h = header.replace(" ", "").strip()
    return _match_invoice_header(h) == ("name", _invoice_header_match_score(h, "name"))


def find_billing_worksheet(wb: Workbook, period_hint: str | None = None) -> Worksheet:
    """
    도급비 청구내역 시트 선택.

    통합 청구서(청구내역·연차·근태 다수 시트)에서 active가 연차 시트인 경우가 있어
    이름·기간 힌트로 청구 시트를 고릅니다.
    """
    from invoice_leave_sheet import infer_period_from_workbook as _infer_period

    hint = period_hint or _infer_period(wb)
    scored: list[tuple[int, str]] = []

    for name in wb.sheetnames:
        compact = name.replace(" ", "")
        lower = name.lower()
        if any(
            skip in compact
            for skip in ("연차", "근태", "Summary", "analysis", "Outsourcing", "작업", "work")
        ):
            continue
        if "청구" not in compact:
            continue
        score = 0
        if "내역" in compact:
            score += 10
        if hint and _PERIOD_IN_NAME_RE.search(name):
            try:
                y, m = hint.split("-")
                yy = int(y)
                mm = int(m)
                if f"{yy % 100:02d}" in compact and (f"{mm:02d}" in compact or f"{mm}월" in name):
                    score += 25
                if f"{yy}년" in name and f"{mm}월" in name:
                    score += 25
            except ValueError:
                pass
        if score > 0 or "내역" in compact:
            scored.append((score, name))

    if scored:
        scored.sort(key=lambda x: (-x[0], x[1]))
        return wb[scored[0][1]]

    for name in wb.sheetnames:
        compact = name.replace(" ", "")
        if "연차" in compact or "근태" in compact:
            continue
        if _sheet_has_billing_headers(wb[name]):
            return wb[name]

    return wb.active


_PERIOD_IN_NAME_RE = re.compile(r"(20\d{2}|26|25|24)\D{0,3}(\d{1,2})")


def _sheet_has_billing_headers(ws: Worksheet) -> bool:
    layout = find_billing_column_layout(ws)
    return layout.source == "keyword"


def find_attendance_worksheet(wb: Workbook) -> Worksheet | None:
    """도급비 청구서 2번째 시트(근태) 또는 이름에 '근태' 포함 시트."""
    for name in wb.sheetnames:
        if "근태" in name.replace(" ", ""):
            return wb[name]
    if len(wb.sheetnames) >= 2:
        second = wb[wb.sheetnames[1]]
        first = wb[wb.sheetnames[0]]
        if second.title != first.title:
            return second
    return None


def _scan_attendance_columns(ws: Worksheet) -> tuple[int, int, int, bool] | None:
    """헤더 행, 성명 열, 지조외(조퇴시간) 열, 조퇴값 분 단위 여부."""
    name_col = early_col = None
    early_is_minutes = False
    header_row = 1
    for r in range(1, min(12, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            raw = ws.cell(r, c).value
            if raw is None:
                continue
            text = str(raw).strip()
            if _header_matches_name(text):
                name_col = c
                header_row = r
            if _header_matches_early_leave(text):
                early_col = c
                early_is_minutes = _header_value_is_minutes(text)
                header_row = r
        if name_col and early_col:
            return header_row, name_col, early_col, early_is_minutes
    return None


def load_attendance_early_leave_hours(
    wb: Workbook,
) -> tuple[dict[str, float], dict[str, str]]:
    """
    근태 시트에서 성명별 지조외(조퇴시간)를 읽습니다.

    Returns:
        ({정규화된 성명 키: 조퇴시간}, {정규화된 성명 키: 표시 이름})
    """
    ws = find_attendance_worksheet(wb)
    if ws is None:
        return {}, {}

    scanned = _scan_attendance_columns(ws)
    if not scanned:
        return {}, {}

    header_row, name_col, early_col, early_is_minutes = scanned
    hours_by_key: dict[str, float] = {}
    display_by_key: dict[str, str] = {}
    for row in range(header_row + 1, ws.max_row + 1):
        name = _cell_str(ws.cell(row, name_col).value)
        if not _is_valid_name(name):
            continue
        hours = safe_number(ws.cell(row, early_col).value, 0.0)
        if early_is_minutes:
            hours = hours / 60.0
        if hours < 0:
            hours = 0.0
        key = norm_name_key(name)
        if key:
            hours_by_key[key] = hours
            display_by_key[key] = name
    return hours_by_key, display_by_key


def _resolve_attendance_hours(
    name_key: str,
    attendance_early: dict[str, float],
) -> tuple[float, str | None]:
    """
    근태 조퇴시간 조회. 정확히 일치하지 않으면 1글자 유사 이름으로 매칭.

    Returns:
        (조퇴시간, 유사매칭된 근태 키 또는 None)
    """
    if name_key in attendance_early:
        return attendance_early[name_key], None
    fuzzy_key = find_fuzzy_name_key(name_key, attendance_early.keys())
    if fuzzy_key:
        return attendance_early[fuzzy_key], fuzzy_key
    return 0.0, None


def attendance_name_warnings(
    invoice_path: Path,
    invoice_rows: list[dict[str, Any]],
) -> list[str]:
    """근태 시트 성명이 청구서와 어긋난 경우(오타·미등록) 경고."""
    wb = openpyxl.load_workbook(invoice_path, data_only=True)
    attendance_early, attendance_display = load_attendance_early_leave_hours(wb)
    wb.close()

    if not attendance_early:
        return []

    invoice_by_key: dict[str, str] = {}
    for row in invoice_rows:
        key = norm_name_key(row.get("name"))
        if key:
            invoice_by_key[key] = str(row["name"]).strip()

    invoice_keys = set(invoice_by_key.keys())
    warnings: list[str] = []
    seen: set[str] = set()

    for att_key, hours in attendance_early.items():
        if att_key in invoice_keys:
            continue

        att_name = attendance_display.get(att_key, att_key)
        fuzzy_inv = find_fuzzy_name_key(att_key, invoice_keys)
        if fuzzy_inv:
            inv_name = invoice_by_key[fuzzy_inv]
            msg = (
                f"근태 시트 '{att_name}' ↔ 청구서 '{inv_name}' 이름이 1글자 다릅니다(오타 가능). "
                f"근태 시트 성명을 '{inv_name}'으로 수정하세요."
            )
            if msg not in seen:
                seen.add(msg)
                warnings.append(msg)
            continue

        if hours > 0:
            msg = f"근태 시트 '{att_name}' 가 청구서에 없어 조퇴시간({hours}h)이 반영되지 않습니다."
            if msg not in seen:
                seen.add(msg)
                warnings.append(msg)

    return warnings


def _is_valid_name(name: str) -> bool:
    """합계·빈 행 등 비직원 이름을 걸러냅니다."""
    if not name or name in ("0", "합계", "소계", "계"):
        return False
    if name.replace(".", "").isdigit():
        return False
    return True


def extract_invoice_data(invoice_path: Path) -> list[dict[str, Any]]:
    """청구서에서 직원별 급여·근태·공제 데이터를 추출합니다."""
    from invoice_leave_sheet import apply_leave_sheet_to_invoice_rows, load_invoice_leave_sheet

    from invoice_leave_sheet import infer_period_from_workbook

    wb = openpyxl.load_workbook(invoice_path, data_only=True)
    period_hint = infer_period_from_workbook(wb)
    ws = find_billing_worksheet(wb, period_hint)
    layout = find_billing_column_layout(ws)
    attendance_early, _attendance_display = load_attendance_early_leave_hours(wb)
    leave_by_key, _leave_period = load_invoice_leave_sheet(wb)
    workers_day_col = find_workers_day_column(ws)
    rows: list[dict[str, Any]] = []

    def raw(row_no: int, key: str) -> Any:
        col = layout.columns.get(key)
        if col is None:
            return None
        return ws.cell(row_no, col).value

    def num(row_no: int, key: str, default: float = 0.0) -> float:
        return safe_number(raw(row_no, key), default)

    def computed_subtotal(row_no: int, workers_day_pay: int) -> int:
        return int(
            num(row_no, "base_salary")
            + num(row_no, "base_deduction")
            + num(row_no, "ot_pay")
            + num(row_no, "shift_pay")
            + num(row_no, "night_pay")
            + num(row_no, "special_pay")
            + num(row_no, "special_ext_pay")
            + num(row_no, "position_pay")
            + workers_day_pay
        )

    def early_leave_value(row_no: int) -> float:
        value = num(row_no, "early_leave")
        col = layout.columns.get("early_leave")
        if col is not None and layout.source == "keyword":
            header = ws.cell(layout.header_row, col).value
            if _header_value_is_minutes(header):
                return value / 60.0
        return value

    for row in range(layout.data_start_row, ws.max_row + 1):
        name = _cell_str(raw(row, "name"))
        if not _is_valid_name(name):
            continue

        workers_day_pay = 0
        if workers_day_col:
            workers_day_pay = int(safe_number(ws.cell(row, workers_day_col).value))

        transport = num(row, "transport")
        gross_raw = num(row, "gross_pay")
        has_subtotal_col = "subtotal" in layout.columns
        subtotal = num(row, "subtotal")
        if subtotal <= 0 and gross_raw <= 0:
            subtotal = computed_subtotal(row, workers_day_pay)
        if gross_raw > 0:
            gross = gross_raw
            if not has_subtotal_col or subtotal <= 0:
                subtotal = max(0, gross - transport)
        else:
            # 업무추진비(AE 등)는 소계(AB)에 포함 — 별도 가산하지 않음
            gross = subtotal + transport
        if gross <= 0 and subtotal <= 0:
            continue

        health = int(num(row, "health"))
        ltc = int(num(row, "long_term_care"))
        pension = int(num(row, "pension"))
        employment = int(num(row, "employment"))
        insurance_total = health + ltc + pension + employment

        name_key = norm_name_key(name)
        early_leave_hours, _fuzzy_att_key = _resolve_attendance_hours(name_key, attendance_early)
        if early_leave_hours == 0.0 and name_key not in attendance_early:
            early_leave_hours = early_leave_value(row)

        rows.append({
            "row": row,
            "name": name,
            "dept": _cell_str(raw(row, "dept")),
            "hire_date": _cell_str(raw(row, "hire")),
            "base_hourly": num(row, "base_hourly"),
            "ordinary_hourly": num(row, "ordinary_hourly"),
            "base_days": num(row, "base_days"),
            "work_days": num(row, "work_days"),
            "unpaid_days": num(row, "unpaid"),
            "leave_days": num(row, "leave"),
            "ot_hours": num(row, "ot_hours"),
            "shift_hours": num(row, "shift_hours"),
            "night_hours": num(row, "night_hours"),
            "special_hours": num(row, "special_hours"),
            "special_ext_hours": num(row, "special_ext_hours"),
            "early_leave_hours": early_leave_hours,
            "base_salary": int(num(row, "base_salary")),
            "base_deduction": int(num(row, "base_deduction")),
            "ot_pay": int(num(row, "ot_pay")),
            "night_pay": int(num(row, "night_pay")),
            "special_pay": int(num(row, "special_pay")),
            "special_ext_pay": int(num(row, "special_ext_pay")),
            "position_pay": int(num(row, "position_pay")),
            "shift_pay": int(num(row, "shift_pay")),
            "workers_day_pay": workers_day_pay,
            "annual_pay": int(num(row, "annual_pay")),
            "transport": int(transport),
            "subtotal": int(subtotal),
            "gross_pay": int(gross),
            "health_insurance": health,
            "long_term_care": ltc,
            "national_pension": pension,
            "employment_insurance": employment,
            "insurance_total": insurance_total,
        })

    apply_leave_sheet_to_invoice_rows(rows, leave_by_key, prefer_sheet=True)
    wb.close()
    return rows
