"""
services/payroll_output_refresh.py - 저장 청구서 기준 급여 산출물 자동 갱신
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from core.payroll_calc_rules import PAYROLL_OUTPUT_ENGINE_VERSION
from excel_writer import OUTPUT_DIR
from insurance import EMPLOYMENT_INSURANCE_WORKER_RATE
from services.archive_storage import (
    INVOICE_STORED_NAME,
    load_scope_manifest,
    scope_output_dir,
    update_scope_manifest,
)
from services.payroll_scope import PayrollScope

_PERIOD_RE = re.compile(r"^\d{4}-\d{2}$")


@dataclass(frozen=True)
class InvoiceHit:
    invoice_path: Path
    scope: PayrollScope


def iter_saved_invoice_scopes() -> list[InvoiceHit]:
    hits: list[InvoiceHit] = []
    if not OUTPUT_DIR.exists():
        return hits
    for inv in OUTPUT_DIR.rglob(INVOICE_STORED_NAME):
        try:
            period_dir = inv.parent
            workplace_dir = period_dir.parent
            affiliate_dir = workplace_dir.parent
            period = period_dir.name
            if not _PERIOD_RE.match(period):
                continue
            if workplace_dir.resolve() == OUTPUT_DIR.resolve():
                continue
            scope = PayrollScope(affiliate_dir.name, workplace_dir.name, period)
            hits.append(InvoiceHit(inv, scope))
        except Exception:
            continue
    hits.sort(key=lambda h: (h.scope.period, h.scope.affiliate, h.scope.workplace))
    return hits


def manifest_engine_version(manifest: dict[str, Any]) -> str:
    return str(manifest.get("engine_version") or "1")


def ledger_has_current_formulas(ledger_path: Path) -> bool:
    """급여대장에 현행 고용보험·차인지급 수식이 있는지 확인."""
    if not ledger_path.is_file():
        return False
    try:
        from openpyxl import load_workbook

        wb = load_workbook(ledger_path, read_only=True, data_only=False)
        ws = wb.active
        rate_token = str(EMPLOYMENT_INSURANCE_WORKER_RATE)
        found_emp = False
        found_net = False
        for r in range(5, min(ws.max_row, 80), 2):
            emp = ws.cell(r, 22).value
            net = ws.cell(r + 1, 23).value
            if isinstance(emp, str) and "ROUND" in emp.upper() and rate_token in emp:
                found_emp = True
            if isinstance(net, str) and net.startswith("=") and "S" in net and "W" in net:
                found_net = True
            if found_emp and found_net:
                break
        wb.close()
        return found_emp and found_net
    except Exception:
        return False


def scope_outputs_stale(scope: PayrollScope) -> bool:
    manifest = load_scope_manifest(scope)
    if manifest_engine_version(manifest) != PAYROLL_OUTPUT_ENGINE_VERSION:
        return True
    ledger = scope_output_dir(scope) / "급여대장.xlsx"
    return not ledger_has_current_formulas(ledger)


def refresh_scope_from_saved_invoice(
    scope: PayrollScope,
    *,
    invoice_path: Path | None = None,
    interactive_parent=None,
) -> dict[str, Any]:
    """저장된 청구서로 해당 월 산출물을 재생성합니다."""
    from main import process_invoice

    inv = invoice_path or (scope_output_dir(scope) / INVOICE_STORED_NAME)
    if not inv.is_file():
        return {"ok": False, "scope": scope.key, "error": "청구서 없음"}
    process_invoice(inv, scope, interactive_parent=interactive_parent)
    update_scope_manifest(scope, engine_version=PAYROLL_OUTPUT_ENGINE_VERSION)
    return {"ok": True, "scope": scope.key, "invoice": str(inv)}


def refresh_stale_payroll_outputs(
    *,
    interactive_parent=None,
    only_scope: PayrollScope | None = None,
) -> dict[str, Any]:
    """엔진 버전·수식 기준이 낡은 월별 산출물을 청구서 기준으로 일괄 갱신."""
    hits = iter_saved_invoice_scopes()
    if only_scope is not None:
        hits = [h for h in hits if h.scope.key == only_scope.key]

    refreshed: list[str] = []
    skipped: list[str] = []
    failed: list[dict[str, str]] = []

    for hit in hits:
        if not scope_outputs_stale(hit.scope):
            skipped.append(hit.scope.key)
            continue
        try:
            refresh_scope_from_saved_invoice(
                hit.scope,
                invoice_path=hit.invoice_path,
                interactive_parent=interactive_parent,
            )
            refreshed.append(hit.scope.key)
        except Exception as exc:
            failed.append({"scope": hit.scope.key, "error": str(exc)})

    return {
        "engine_version": PAYROLL_OUTPUT_ENGINE_VERSION,
        "refreshed": refreshed,
        "skipped": skipped,
        "failed": failed,
    }


def stamp_manifest_engine_version(scope: PayrollScope) -> None:
    update_scope_manifest(scope, engine_version=PAYROLL_OUTPUT_ENGINE_VERSION)
