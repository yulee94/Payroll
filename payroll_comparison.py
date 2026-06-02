"""
payroll_comparison.py - 전월 대비 급여 차이 비교 (임원 보고용)

- 매월 급여 처리 시 output/{YYYY-MM}/payroll_snapshot.json 저장
- 급여차이내역/ 폴더에 전월 대비 비교 Excel 생성
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from roster_constants import norm_name_key
from utils import safe_number

from core.file_save import save_workbook
from core.paths import dev_root, payroll_diff_dir

BASE_DIR = dev_root()
PAYROLL_DIFF_DIR = payroll_diff_dir()
SNAPSHOT_FILENAME = "payroll_snapshot.json"
SNAPSHOT_TABLE_XLSX = "payroll_snapshot.xlsx"
SNAPSHOT_TABLE_CSV = "payroll_snapshot.csv"

# 비교·차이내역에 포함할 금액 항목 (records 키 → 한글명)
AMOUNT_FIELDS: tuple[tuple[str, str], ...] = (
    ("gross_pay", "총지급액"),
    ("net_pay", "실수령액"),
    ("total_deduction", "공제합계"),
    ("base_salary", "기본급"),
    ("base_deduction", "기본공제"),
    ("early_leave_deduction", "조퇴공제"),
    ("ot_pay", "연장수당"),
    ("shift_pay", "교대수당"),
    ("night_pay", "심야수당"),
    ("special_pay", "특근수당"),
    ("special_ext_pay", "특근연장"),
    ("position_pay", "직책수당"),
    ("transport", "교통비"),
    ("health_insurance", "건강보험"),
    ("long_term_care", "장기요양"),
    ("national_pension", "국민연금"),
    ("employment_insurance", "고용보험"),
    ("income_tax", "소득세"),
    ("local_income_tax", "지방소득세"),
)

SUMMARY_FIELDS = ("gross_pay", "net_pay", "total_deduction")


def ensure_payroll_diff_dir() -> Path:
    PAYROLL_DIFF_DIR.mkdir(parents=True, exist_ok=True)
    return PAYROLL_DIFF_DIR


def prev_period_label(period: str) -> str:
    """YYYY-MM → 전월 YYYY-MM."""
    year, month = period.split("-")
    y, m = int(year), int(month)
    if m == 1:
        return f"{y - 1:04d}-12"
    return f"{y:04d}-{m - 1:02d}"


def format_period_korean(period: str) -> str:
    """2026-05 → 2026년 05월."""
    year, month = period.split("-")
    return f"{year}년 {int(month):02d}월"


def _records_by_name(records: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for rec in records:
        key = norm_name_key(rec.get("name"))
        if key:
            out[key] = rec
    return out


def _snapshot_fields(rec: dict[str, Any]) -> dict[str, Any]:
    data: dict[str, Any] = {
        "name": rec.get("name"),
        "dept": rec.get("dept") or "",
        "workplace": rec.get("workplace") or "",
        "affiliate": rec.get("affiliate") or "",
    }
    for field, _ in AMOUNT_FIELDS:
        data[field] = int(safe_number(rec.get(field), 0))
    for field in ("work_days", "leave_days", "unpaid_days", "ot_hours"):
        data[field] = safe_number(rec.get(field), 0.0)
    data["leave_usage_display"] = rec.get("leave_usage_display") or ""
    return data


def save_payroll_snapshot(
    records: list[dict[str, Any]],
    out_dir: Path,
    period: str,
) -> Path:
    """당월 급여 스냅샷을 산출 폴더에 저장."""
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / SNAPSHOT_FILENAME
    payload = {
        "period": period,
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "count": len(records),
        "records": [_snapshot_fields(r) for r in records],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def _snapshot_table_columns() -> list[tuple[str, str]]:
    cols: list[tuple[str, str]] = [("name", "성명"), ("dept", "부서"), ("workplace", "사업장"), ("affiliate", "계열사")]
    cols.extend([(k, label) for k, label in AMOUNT_FIELDS])
    cols.extend(
        [
            ("work_days", "근무일"),
            ("leave_days", "연차"),
            ("unpaid_days", "무급/결근"),
            ("ot_hours", "연장시간"),
            ("leave_usage_display", "연차/결근표기"),
        ]
    )
    return cols


def write_payroll_snapshot_table_excel(
    records: list[dict[str, Any]],
    output_path: Path,
) -> Path:
    """스냅샷을 사람이 보기 쉬운 표 형태 Excel로 저장."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "snapshot"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    cols = _snapshot_table_columns()
    for c, (_key, label) in enumerate(cols, 1):
        ws.cell(1, c, label)
    _style_header_row(ws, 1, len(cols))

    snap = [_snapshot_fields(r) for r in records]
    for i, r in enumerate(snap, 1):
        row = i + 1
        for c, (key, _label) in enumerate(cols, 1):
            val = r.get(key, "")
            cell = ws.cell(row, c, val)
            if key in {"name", "dept", "leave_usage_display"}:
                _style_body_cell(cell, align="left")
            elif key.endswith("_days") or key.endswith("_hours"):
                _style_body_cell(cell, align="right")
            else:
                _style_body_cell(cell, align="right", number=True)

    # widths
    for idx, (key, _label) in enumerate(cols, 1):
        if key == "name":
            w = 10
        elif key == "dept":
            w = 10
        elif key in {"leave_usage_display"}:
            w = 16
        elif key.endswith("_days") or key.endswith("_hours"):
            w = 10
        else:
            w = 12
        ws.column_dimensions[get_column_letter(idx)].width = w

    save_workbook(wb, output_path)
    wb.close()
    return output_path


def write_payroll_snapshot_table_csv(
    records: list[dict[str, Any]],
    output_path: Path,
) -> Path:
    """스냅샷을 CSV(UTF-8 BOM)로 저장. (엑셀에서 바로 열기 용이)"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    cols = _snapshot_table_columns()
    snap = [_snapshot_fields(r) for r in records]

    def _csv_cell(v: Any) -> str:
        s = "" if v is None else str(v)
        s = s.replace("\r", " ").replace("\n", " ")
        if any(ch in s for ch in [",", "\"", "\n"]):
            s = "\"" + s.replace("\"", "\"\"") + "\""
        return s

    lines: list[str] = []
    lines.append(",".join(_csv_cell(label) for _k, label in cols))
    for r in snap:
        lines.append(",".join(_csv_cell(r.get(k, "")) for k, _label in cols))

    output_path.write_text("\n".join(lines), encoding="utf-8-sig")
    return output_path


def _resolve_writable_path(path: Path) -> Path:
    """
    동일 파일이 열려있어 잠긴 경우(Excel) 대비:
    같은 폴더에 _1, _2 ... 를 붙여 저장 경로를 찾습니다.
    """
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    for i in range(1, 50):
        candidate = parent / f"{stem}_{i}{suffix}"
        if not candidate.exists():
            return candidate
    # 마지막 수단: 타임스탬프
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return parent / f"{stem}_{ts}{suffix}"


def load_payroll_snapshot(out_dir: Path) -> list[dict[str, Any]] | None:
    path = out_dir / SNAPSHOT_FILENAME
    if not path.exists():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        records = data.get("records")
        if isinstance(records, list) and records:
            return records
    except (json.JSONDecodeError, OSError):
        return None
    return None


def _sum_field(records: list[dict[str, Any]], field: str) -> int:
    return sum(int(safe_number(r.get(field), 0)) for r in records)


def _pct_change(prev: int, curr: int) -> float | None:
    if prev == 0:
        return None if curr == 0 else 100.0
    return (curr - prev) / prev * 100.0


def _top_reasons(prev: dict[str, Any], curr: dict[str, Any], limit: int = 3) -> str:
    """항목별 차이 중 큰 순으로 요약 문자열."""
    parts: list[tuple[int, str]] = []
    for field, label in AMOUNT_FIELDS:
        if field in SUMMARY_FIELDS:
            continue
        diff = int(safe_number(curr.get(field), 0)) - int(safe_number(prev.get(field), 0))
        if diff != 0:
            parts.append((abs(diff), f"{label} {diff:+,}"))
    parts.sort(reverse=True)
    return ", ".join(p[1] for p in parts[:limit]) if parts else "-"


def build_payroll_comparison(
    current_records: list[dict[str, Any]],
    prior_records: list[dict[str, Any]] | None,
    current_period: str,
    prior_period: str | None,
) -> dict[str, Any]:
    """전월·금월 비교 데이터 생성."""
    curr_map = _records_by_name(current_records)
    prior_map = _records_by_name(prior_records or [])
    all_keys = sorted(set(curr_map) | set(prior_map))

    rows: list[dict[str, Any]] = []
    detail_rows: list[dict[str, Any]] = []

    for key in all_keys:
        curr = curr_map.get(key)
        prev = prior_map.get(key)
        name = (curr or prev or {}).get("name") or key
        dept = (curr or prev or {}).get("dept") or ""

        if curr and not prev:
            status = "신규"
        elif prev and not curr:
            status = "전월만"
        else:
            status = "동일"

        prev_gross = int(safe_number(prev.get("gross_pay"), 0)) if prev else 0
        curr_gross = int(safe_number(curr.get("gross_pay"), 0)) if curr else 0
        prev_net = int(safe_number(prev.get("net_pay"), 0)) if prev else 0
        curr_net = int(safe_number(curr.get("net_pay"), 0)) if curr else 0
        gross_diff = curr_gross - prev_gross
        net_diff = curr_net - prev_net

        reasons = _top_reasons(prev or {}, curr or {}) if status == "동일" else status

        row = {
            "name": name,
            "dept": dept,
            "status": status,
            "prev_gross": prev_gross,
            "curr_gross": curr_gross,
            "gross_diff": gross_diff,
            "gross_pct": _pct_change(prev_gross, curr_gross),
            "prev_net": prev_net,
            "curr_net": curr_net,
            "net_diff": net_diff,
            "net_pct": _pct_change(prev_net, curr_net),
            "reasons": reasons,
        }
        rows.append(row)

        if status != "동일" or gross_diff != 0 or net_diff != 0:
            for field, label in AMOUNT_FIELDS:
                pv = int(safe_number(prev.get(field), 0)) if prev else 0
                cv = int(safe_number(curr.get(field), 0)) if curr else 0
                diff = cv - pv
                if diff == 0:
                    continue
                detail_rows.append(
                    {
                        "name": name,
                        "dept": dept,
                        "status": status,
                        "field_label": label,
                        "prev": pv,
                        "curr": cv,
                        "diff": diff,
                    }
                )

    curr_total_gross = _sum_field(current_records, "gross_pay")
    curr_total_net = _sum_field(current_records, "net_pay")
    curr_total_ded = _sum_field(current_records, "total_deduction")
    prior_total_gross = _sum_field(prior_records or [], "gross_pay")
    prior_total_net = _sum_field(prior_records or [], "net_pay")
    prior_total_ded = _sum_field(prior_records or [], "total_deduction")

    changed = [r for r in rows if r["status"] == "동일" and r["gross_diff"] != 0]
    increased = sorted(changed, key=lambda r: r["gross_diff"], reverse=True)[:5]
    decreased = sorted(changed, key=lambda r: r["gross_diff"])[:5]

    new_hires = [r for r in rows if r["status"] == "신규"]
    departed = [r for r in rows if r["status"] == "전월만"]

    bullets: list[str] = []
    if prior_records:
        bullets.append(
            f"총지급액 {prior_total_gross:,}원 → {curr_total_gross:,}원 "
            f"({curr_total_gross - prior_total_gross:+,}원)"
        )
        bullets.append(
            f"실수령액 {prior_total_net:,}원 → {curr_total_net:,}원 "
            f"({curr_total_net - prior_total_net:+,}원)"
        )
        bullets.append(
            f"인원 {len(prior_map)}명 → {len(curr_map)}명 "
            f"({len(curr_map) - len(prior_map):+d}명)"
        )
        if new_hires:
            bullets.append(f"신규 {len(new_hires)}명: {', '.join(r['name'] for r in new_hires[:5])}")
        if departed:
            bullets.append(f"전월만 {len(departed)}명: {', '.join(r['name'] for r in departed[:5])}")
        bullets.append(f"총지급 변동 {len(changed)}명 (증가 {sum(1 for r in changed if r['gross_diff'] > 0)}명, "
                       f"감소 {sum(1 for r in changed if r['gross_diff'] < 0)}명)")
    else:
        bullets.append(f"전월({format_period_korean(prior_period or '')}) 스냅샷 없음 — 금월 합계만 표시")
        bullets.append(f"금월 총지급액 {curr_total_gross:,}원, 실수령액 {curr_total_net:,}원 ({len(curr_map)}명)")

    return {
        "current_period": current_period,
        "prior_period": prior_period,
        "has_prior": bool(prior_records),
        "summary": {
            "prior_count": len(prior_map),
            "current_count": len(curr_map),
            "prior_gross": prior_total_gross,
            "current_gross": curr_total_gross,
            "gross_diff": curr_total_gross - prior_total_gross,
            "prior_net": prior_total_net,
            "current_net": curr_total_net,
            "net_diff": curr_total_net - prior_total_net,
            "prior_deduction": prior_total_ded,
            "current_deduction": curr_total_ded,
            "deduction_diff": curr_total_ded - prior_total_ded,
        },
        "bullets": bullets,
        "increased_top5": increased,
        "decreased_top5": decreased,
        "rows": rows,
        "detail_rows": detail_rows,
    }


# ---------------------------------------------------------------------------
# Excel 스타일 (임원 보고용)
# ---------------------------------------------------------------------------
_FONT_TITLE = Font(name="맑은 고딕", size=16, bold=True, color="1F3864")
_FONT_SUB = Font(name="맑은 고딕", size=10, color="666666")
_FONT_HDR = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
_FONT_BODY = Font(name="맑은 고딕", size=10)
_FONT_BOLD = Font(name="맑은 고딕", size=10, bold=True)
_FILL_HDR = PatternFill("solid", fgColor="1F3864")
_FILL_SUBHDR = PatternFill("solid", fgColor="D6E4F0")
_FILL_TOTAL = PatternFill("solid", fgColor="EEF2F7")
_THIN = Side(style="thin", color="B4B4B4")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_NUM_FMT = "#,##0"
_PCT_FMT = "0.0%"


def _style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font = _FONT_HDR
        cell.fill = _FILL_HDR
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER


def _style_body_cell(cell, align: str = "center", bold: bool = False, number: bool = False) -> None:
    cell.font = _FONT_BOLD if bold else _FONT_BODY
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _BORDER
    if number:
        cell.number_format = _NUM_FMT


def _write_summary_sheet(wb: Workbook, comp: dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "보고요약"
    ws.sheet_view.showGridLines = False

    cur = comp["current_period"]
    pri = comp.get("prior_period") or prev_period_label(cur)
    s = comp["summary"]

    ws.merge_cells("A1:F1")
    ws["A1"] = "전월 대비 급여 변동 보고서"
    ws["A1"].font = _FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"대상: {format_period_korean(pri)} → {format_period_korean(cur)}  |  "
        f"작성일: {datetime.now():%Y-%m-%d}"
    )
    ws["A2"].font = _FONT_SUB
    ws["A2"].alignment = Alignment(horizontal="center")

    row = 4
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row, 1, "1. 핵심 지표").font = _FONT_BOLD
    row += 1

    headers = ["구분", "전월", "금월", "차이", "증감률", "비고"]
    for c, h in enumerate(headers, 1):
        ws.cell(row, c, h)
    _style_header_row(ws, row, len(headers))
    row += 1

    metrics = [
        ("인원(명)", s["prior_count"], s["current_count"], s["current_count"] - s["prior_count"], None, ""),
        ("총지급액(원)", s["prior_gross"], s["current_gross"], s["gross_diff"],
         _pct_change(s["prior_gross"], s["current_gross"]), ""),
        ("공제합계(원)", s["prior_deduction"], s["current_deduction"], s["deduction_diff"],
         _pct_change(s["prior_deduction"], s["current_deduction"]), ""),
        ("실수령액(원)", s["prior_net"], s["current_net"], s["net_diff"],
         _pct_change(s["prior_net"], s["current_net"]), ""),
    ]
    for label, pv, cv, diff, pct, note in metrics:
        ws.cell(row, 1, label)
        ws.cell(row, 2, pv)
        ws.cell(row, 3, cv)
        ws.cell(row, 4, diff)
        if pct is None:
            ws.cell(row, 5, "-")
        else:
            ws.cell(row, 5, pct / 100.0)
            ws.cell(row, 5).number_format = _PCT_FMT
        ws.cell(row, 6, note)
        for c in range(1, 7):
            _style_body_cell(ws.cell(row, c), align="right" if c > 1 else "left", number=c in (2, 3, 4))
        ws.cell(row, 1).alignment = Alignment(horizontal="left", vertical="center")
        diff_cell = ws.cell(row, 4)
        if diff > 0:
            diff_cell.font = Font(name="맑은 고딕", size=10, color="C00000")
        elif diff < 0:
            diff_cell.font = Font(name="맑은 고딕", size=10, color="0070C0")
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row, 1, "2. 주요 내용").font = _FONT_BOLD
    row += 1
    for bullet in comp["bullets"]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row, 1, f"• {bullet}")
        cell.font = _FONT_BODY
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row += 1

    def _write_top_section(title: str, items: list[dict[str, Any]]) -> int:
        nonlocal row
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row, 1, title).font = _FONT_BOLD
        row += 1
        top_hdr = ["순위", "성명", "부서", "전월 총지급", "금월 총지급", "차이"]
        for c, h in enumerate(top_hdr, 1):
            ws.cell(row, c, h)
        _style_header_row(ws, row, len(top_hdr))
        row += 1
        if not items:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row, 1, "(해당 없음)")
            row += 1
            return row
        for i, r in enumerate(items, 1):
            ws.cell(row, 1, i)
            ws.cell(row, 2, r["name"])
            ws.cell(row, 3, r["dept"])
            ws.cell(row, 4, r["prev_gross"])
            ws.cell(row, 5, r["curr_gross"])
            ws.cell(row, 6, r["gross_diff"])
            for c in range(1, 7):
                _style_body_cell(ws.cell(row, c), number=c >= 4)
            row += 1
        return row

    if comp["has_prior"]:
        _write_top_section("3. 총지급 증가 상위 5명", comp["increased_top5"])
        _write_top_section("4. 총지급 감소 상위 5명", comp["decreased_top5"])

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18


def _write_compare_sheet(wb: Workbook, comp: dict[str, Any]) -> None:
    ws = wb.create_sheet("급여비교표")
    cur = format_period_korean(comp["current_period"])
    pri = format_period_korean(comp["prior_period"] or prev_period_label(comp["current_period"]))

    headers = [
        "No", "성명", "부서", "구분",
        f"전월({pri}) 총지급", f"금월({cur}) 총지급", "총지급 차이", "증감률",
        f"전월 실수령", f"금월 실수령", "실수령 차이", "주요 차이 요인",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    _style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    for i, r in enumerate(comp["rows"], 1):
        row = i + 1
        pct = r["gross_pct"]
        values = [
            i, r["name"], r["dept"], r["status"],
            r["prev_gross"], r["curr_gross"], r["gross_diff"],
            (pct / 100.0) if pct is not None else "-",
            r["prev_net"], r["curr_net"], r["net_diff"], r["reasons"],
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row, c, val)
            if c in (5, 6, 7, 9, 10, 11):
                _style_body_cell(cell, number=True)
            elif c == 8 and isinstance(val, float):
                cell.number_format = _PCT_FMT
                _style_body_cell(cell)
            else:
                _style_body_cell(cell, align="left" if c in (2, 3, 12) else "center")
        diff_cell = ws.cell(row, 7)
        if r["gross_diff"] > 0:
            diff_cell.font = Font(name="맑은 고딕", size=10, color="C00000")
        elif r["gross_diff"] < 0:
            diff_cell.font = Font(name="맑은 고딕", size=10, color="0070C0")

    widths = [5, 10, 12, 8, 14, 14, 12, 8, 12, 12, 12, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_detail_sheet(wb: Workbook, comp: dict[str, Any]) -> None:
    ws = wb.create_sheet("차이내역")
    headers = ["No", "성명", "부서", "구분", "항목", "전월", "금월", "차이"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    _style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    for i, d in enumerate(comp["detail_rows"], 1):
        row = i + 1
        values = [i, d["name"], d["dept"], d["status"], d["field_label"], d["prev"], d["curr"], d["diff"]]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row, c, val)
            if c >= 6:
                _style_body_cell(cell, number=True)
            else:
                _style_body_cell(cell, align="left" if c in (2, 3, 5) else "center")

    if not comp["detail_rows"]:
        ws.cell(2, 1, "변동 내역 없음")

    for i, w in enumerate([5, 10, 12, 8, 14, 12, 12, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_payroll_comparison_report(comp: dict[str, Any], output_path: Path) -> Path:
    """임원 보고용 Excel 3시트 저장."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _write_summary_sheet(wb, comp)
    _write_compare_sheet(wb, comp)
    _write_detail_sheet(wb, comp)
    save_workbook(wb, output_path)
    wb.close()
    return output_path


def generate_payroll_comparison(
    records: list[dict[str, Any]],
    scope,
) -> dict[str, Any]:
    """
    스냅샷 저장 + 전월 대비 비교 보고서 생성 (사업장별 폴더 내).
    """
    from services.payroll_scope import PayrollScope, resolve_output_dir

    if not isinstance(scope, PayrollScope):
        raise TypeError("scope must be PayrollScope")

    period = scope.period
    out_dir = resolve_output_dir(scope)
    out_dir.mkdir(parents=True, exist_ok=True)

    save_payroll_snapshot(records, out_dir, period)

    snapshot_xlsx = _resolve_writable_path(out_dir / f"{period}_급여스냅샷.xlsx")
    snapshot_csv = _resolve_writable_path(out_dir / f"{period}_급여스냅샷.csv")
    write_payroll_snapshot_table_excel(records, snapshot_xlsx)
    write_payroll_snapshot_table_csv(records, snapshot_csv)

    prior_scope = scope.prior()
    prior_dir = resolve_output_dir(prior_scope)
    prior_records = load_payroll_snapshot(prior_dir)
    prior_period = prior_scope.period

    comp = build_payroll_comparison(records, prior_records, period, prior_period)

    report_name = f"{period}_전월대비_급여차이보고.xlsx"
    report_path = _resolve_writable_path(out_dir / report_name)
    write_payroll_comparison_report(comp, report_path)

    warning = None
    if not prior_records:
        warning = (
            f"전월({format_period_korean(prior_period)}) 급여 스냅샷이 없어 "
            "금월 합계만 포함된 보고서를 생성했습니다. "
            "다음 달부터 전월 대비 비교가 가능합니다."
        )

    return {
        "path": report_path,
        "prior_period": prior_period,
        "has_prior": bool(prior_records),
        "summary": comp["summary"],
        "warning": warning,
        "snapshot_xlsx": snapshot_xlsx,
        "snapshot_csv": snapshot_csv,
    }
