"""
leave_usage_ledger.py - 연차사용대장 (별도 폴더, 5년 보관)

- 저장 위치: 급여프로그램/연차사용대장/연차사용대장.xlsx
- 매월 급여 처리 시 당월분 반영(동일 월·성명은 갱신)
- 처리월 기준 5년 초과 분은 자동 삭제
- templates/근로자명부.xlsx 는 읽기만 (최초 1회 명부 연차대장 시트 → 이관 가능)
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from roster_constants import build_header_map, norm_name_key
from roster_workbook import (
    LEAVE_LEDGER_ALIASES,
    MONTHLY_LEAVE_ALIASES,
    LeaveLedgerEntry,
    MonthlyLeaveSummary,
    _ensure_leave_ledger_headers,
    _find_ledger_rows_for_period,
    find_leave_ledger_sheet,
)

from core.file_save import save_workbook
from core.paths import dev_root, leave_usage_ledger_dir

BASE_DIR = dev_root()
LEAVE_USAGE_LEDGER_DIR = leave_usage_ledger_dir()
LEAVE_USAGE_LEDGER_FILENAME = "연차사용대장.xlsx"
ARCHIVE_SUBDIR = "보관"
RETENTION_YEARS = 5
SHEET_NAME = "연차사용대장"
MONTHLY_SHEET_NAME = "월별현황"

_PERIOD_YM_RE = re.compile(r"(20\d{2})\D{0,3}(\d{1,2})")


def get_leave_usage_ledger_path() -> Path:
    """연차사용대장 통합 파일 경로."""
    return LEAVE_USAGE_LEDGER_DIR / LEAVE_USAGE_LEDGER_FILENAME


def ensure_leave_usage_ledger_dir() -> Path:
    """연차사용대장 폴더 생성."""
    LEAVE_USAGE_LEDGER_DIR.mkdir(parents=True, exist_ok=True)
    (LEAVE_USAGE_LEDGER_DIR / ARCHIVE_SUBDIR).mkdir(parents=True, exist_ok=True)
    return LEAVE_USAGE_LEDGER_DIR


def normalize_period_label(value: Any) -> str | None:
    """셀 값 → YYYY-MM (파싱 실패 시 None)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    text = str(value).strip()
    if not text:
        return None
    m = _PERIOD_YM_RE.search(text.replace("년", "-").replace("월", "-"))
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    if re.fullmatch(r"\d{4}-\d{2}", text):
        return text
    return None


def retention_cutoff_ym(reference_period: str, years: int = RETENTION_YEARS) -> str:
    """reference_period(YYYY-MM) 기준 years년 이전 월 (미만 삭제)."""
    ref = normalize_period_label(reference_period) or datetime.now().strftime("%Y-%m")
    y, m = map(int, ref.split("-"))
    y -= years
    return f"{y:04d}-{m:02d}"


def _ym_lt(a: str, b: str) -> bool:
    """a < b (YYYY-MM)."""
    ay, am = map(int, a.split("-"))
    by, bm = map(int, b.split("-"))
    return (ay, am) < (by, bm)


def _create_ledger_workbook(path: Path) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    headers = [
        "성명",
        "사번",
        "처리월",
        "사용월",
        "구분",
        "사용일수",
        "발생연차",
        "누적사용연차",
        "잔여연차",
        "무급횟수",
        "사용내역",
        "최종갱신",
    ]
    for col, label in enumerate(headers, start=1):
        ws.cell(1, col, label)

    ws_monthly = wb.create_sheet(MONTHLY_SHEET_NAME)
    monthly_headers = [
        "성명",
        "사번",
        "처리월",
        "발생연차",
        "당월연차",
        "누적사용연차",
        "잔여연차",
        "무급일수",
        "무급횟수",
        "연차내역",
        "무급내역",
        "최종갱신",
    ]
    for col, label in enumerate(monthly_headers, start=1):
        ws_monthly.cell(1, col, label)

    save_workbook(wb, path)
    return wb


def _open_or_create_ledger(path: Path) -> tuple[openpyxl.Workbook, Worksheet, Worksheet | None]:
    ensure_leave_usage_ledger_dir()
    if not path.exists():
        wb = _create_ledger_workbook(path)
        return wb, wb[SHEET_NAME], wb[MONTHLY_SHEET_NAME]
    wb = openpyxl.load_workbook(path)
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.active
        ws.title = SHEET_NAME
    if MONTHLY_SHEET_NAME in wb.sheetnames:
        ws_monthly = wb[MONTHLY_SHEET_NAME]
    else:
        ws_monthly = wb.create_sheet(MONTHLY_SHEET_NAME)
        monthly_headers = [
            "성명", "사번", "처리월", "발생연차", "당월연차", "누적사용연차",
            "잔여연차", "무급일수", "무급횟수", "연차내역", "무급내역", "최종갱신",
        ]
        for col, label in enumerate(monthly_headers, start=1):
            ws_monthly.cell(1, col, label)
    return wb, ws, ws_monthly


def _row_period_ym(
    ws: Worksheet,
    row: int,
    headers: dict[str, int],
) -> str | None:
    period_col = headers.get("처리월") or headers.get("사용월")
    if not period_col:
        return None
    return normalize_period_label(ws.cell(row, period_col).value)


def purge_leave_usage_older_than(
    ws: Worksheet,
    headers: dict[str, int],
    reference_period: str,
    years: int = RETENTION_YEARS,
) -> int:
    """처리월이 보관 기준보다 오래된 행 삭제. 삭제 건수 반환."""
    cutoff = retention_cutoff_ym(reference_period, years)
    to_delete: list[int] = []
    for r in range(2, ws.max_row + 1):
        name_col = headers.get("성명")
        if name_col and not ws.cell(r, name_col).value:
            continue
        period = _row_period_ym(ws, r, headers)
        if period is None:
            continue
        if _ym_lt(period, cutoff):
            to_delete.append(r)
    for r in sorted(to_delete, reverse=True):
        ws.delete_rows(r, 1)
    return len(to_delete)


def _copy_row_to_ledger(
    ws_src: Worksheet,
    row: int,
    headers_src: dict[str, int],
    ws_dest: Worksheet,
    headers_dest: dict[str, int],
) -> None:
    dest_row = ws_dest.max_row + 1
    mapping = (
        ("성명", "성명"),
        ("사번", "사번"),
        ("처리월", "처리월"),
        ("사용월", "사용월"),
        ("구분", "구분"),
        ("사용일수", "사용일수"),
        ("사용내역", "사용내역"),
    )
    for src_key, dst_key in mapping:
        sc = headers_src.get(src_key)
        dc = headers_dest.get(dst_key)
        if sc and dc:
            ws_dest.cell(dest_row, dc, ws_src.cell(row, sc).value)
    if "최종갱신" in headers_dest:
        ws_dest.cell(dest_row, headers_dest["최종갱신"], datetime.now().strftime("%Y-%m-%d %H:%M"))


def migrate_from_roster_if_needed(
    roster_path: Path | None,
    ledger_path: Path,
    ws: Worksheet,
    headers: dict[str, int],
) -> int:
    """
    통합 연차사용대장이 비어 있고 명부에 연차대장 시트가 있으면 1회 이관.
    """
    if roster_path is None or not roster_path.exists():
        return 0
    if ws.max_row > 1:
        return 0

    wb = openpyxl.load_workbook(roster_path, data_only=True)
    ws_src = find_leave_ledger_sheet(wb)
    if ws_src is None:
        wb.close()
        return 0

    headers_src = build_header_map(ws_src, LEAVE_LEDGER_ALIASES)
    if "성명" not in headers_src:
        wb.close()
        return 0

    copied = 0
    for r in range(2, ws_src.max_row + 1):
        if not ws_src.cell(r, headers_src["성명"]).value:
            continue
        _copy_row_to_ledger(ws_src, r, headers_src, ws, headers)
        copied += 1
    wb.close()
    return copied


def _ledger_period_tag(entry: LeaveLedgerEntry) -> str:
    days = entry.days
    days_s = int(days) if days == int(days) else days
    kind = entry.record_kind or "연차"
    if kind == "결근":
        return f"[{entry.period_label}:결근:{days_s}]"
    return f"[{entry.period_label}:연차:{days_s}]"


def _write_entry_row(
    ws: Worksheet,
    headers: dict[str, int],
    entry: LeaveLedgerEntry,
    updated_at: str,
) -> None:
    row = ws.max_row + 1
    kind = entry.record_kind or "연차"
    tag = _ledger_period_tag(entry)
    memo = f"{tag} {entry.usage_memo}".strip()

    if "성명" in headers:
        ws.cell(row, headers["성명"], entry.name)
    if "사번" in headers and entry.emp_no:
        ws.cell(row, headers["사번"], entry.emp_no)
    if "처리월" in headers:
        ws.cell(row, headers["처리월"], entry.period_label)
    if "사용월" in headers:
        ws.cell(row, headers["사용월"], entry.usage_month)
    if "구분" in headers:
        ws.cell(row, headers["구분"], kind)
    if "사용일수" in headers:
        ws.cell(row, headers["사용일수"], entry.days)
    if "사용내역" in headers:
        ws.cell(row, headers["사용내역"], memo)
    if entry.accrued is not None and "발생연차" in headers:
        ws.cell(row, headers["발생연차"], entry.accrued)
    if entry.used_total is not None and "누적사용연차" in headers:
        ws.cell(row, headers["누적사용연차"], entry.used_total)
    if entry.remaining is not None and "잔여연차" in headers:
        ws.cell(row, headers["잔여연차"], entry.remaining)
    if entry.occurrence_count is not None and "무급횟수" in headers:
        ws.cell(row, headers["무급횟수"], entry.occurrence_count)
    if "최종갱신" in headers:
        ws.cell(row, headers["최종갱신"], updated_at)


def _header_col_by_label(ws: Worksheet, label: str) -> int | None:
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None and str(v).strip() == label:
            return c
    return None


def _ensure_column(ws: Worksheet, headers: dict[str, int], canonical: str, label: str) -> None:
    if canonical in headers:
        return
    existing = _header_col_by_label(ws, label)
    if existing:
        headers[canonical] = existing
        return
    col = ws.max_column + 1
    ws.cell(1, col, label)
    headers[canonical] = col


def _ensure_detail_columns(ws: Worksheet, headers: dict[str, int]) -> dict[str, int]:
    headers = _ensure_leave_ledger_headers(ws, headers)
    for canonical, label in [
        ("구분", "구분"),
        ("발생연차", "발생연차"),
        ("누적사용연차", "누적사용연차"),
        ("잔여연차", "잔여연차"),
        ("무급횟수", "무급횟수"),
        ("최종갱신", "최종갱신"),
    ]:
        _ensure_column(ws, headers, canonical, label)
    return headers


def _ensure_monthly_columns(ws: Worksheet, headers: dict[str, int]) -> dict[str, int]:
    for canonical, label in [
        ("성명", "성명"),
        ("사번", "사번"),
        ("처리월", "처리월"),
        ("발생연차", "발생연차"),
        ("당월연차", "당월연차"),
        ("누적사용연차", "누적사용연차"),
        ("잔여연차", "잔여연차"),
        ("무급일수", "무급일수"),
        ("무급횟수", "무급횟수"),
        ("연차내역", "연차내역"),
        ("무급내역", "무급내역"),
        ("최종갱신", "최종갱신"),
    ]:
        _ensure_column(ws, headers, canonical, label)
    return headers


def _find_monthly_rows_for_period(
    ws: Worksheet,
    headers: dict[str, int],
    period_label: str,
) -> list[int]:
    """같은 처리월의 월별현황 행(재처리 시 삭제)."""
    period_col = headers.get("처리월")
    name_col = headers.get("성명", 0)
    rows: list[int] = []
    if not period_col or not name_col:
        return rows
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, name_col).value
        if nm is None or str(nm).strip() == "":
            continue
        pv = ws.cell(r, period_col).value
        if pv is not None and normalize_period_label(pv) == period_label:
            rows.append(r)
    return rows


def _write_monthly_summary_row(
    ws: Worksheet,
    headers: dict[str, int],
    summary: MonthlyLeaveSummary,
    updated_at: str,
) -> None:
    row = ws.max_row + 1
    mapping: list[tuple[str, Any]] = [
        ("성명", summary.name),
        ("사번", summary.emp_no),
        ("처리월", summary.period_label),
        ("발생연차", summary.accrued),
        ("당월연차", summary.month_leave_used),
        ("누적사용연차", summary.used_total),
        ("잔여연차", summary.remaining),
        ("무급일수", summary.absence_days),
        ("무급횟수", summary.absence_count),
        ("연차내역", summary.leave_memo),
        ("무급내역", summary.absence_memo),
        ("최종갱신", updated_at),
    ]
    for key, val in mapping:
        col = headers.get(key)
        if col and val is not None and val != "":
            ws.cell(row, col, val)
        elif col and key in ("발생연차", "당월연차", "누적사용연차", "잔여연차", "무급일수", "무급횟수"):
            ws.cell(row, col, val if val is not None else 0)


def save_leave_usage_ledger_entries(
    entries: list[LeaveLedgerEntry],
    payroll_period: str,
    *,
    roster_path_for_migration: Path | None = None,
    retention_years: int = RETENTION_YEARS,
    monthly_summaries: list[MonthlyLeaveSummary] | None = None,
) -> dict[str, Any]:
    """
    연차사용대장 폴더의 통합 파일에 당월 기록을 반영하고 5년 초과분을 삭제합니다.

    Returns:
        written, purged, migrated, monthly_written, path
    """
    result: dict[str, Any] = {
        "written": 0,
        "purged": 0,
        "migrated": 0,
        "monthly_written": 0,
        "path": None,
    }

    ledger_path = get_leave_usage_ledger_path()
    ensure_leave_usage_ledger_dir()
    wb, ws, ws_monthly = _open_or_create_ledger(ledger_path)
    headers = _ensure_detail_columns(ws, build_header_map(ws, LEAVE_LEDGER_ALIASES))

    migrated = migrate_from_roster_if_needed(
        roster_path_for_migration, ledger_path, ws, headers
    )
    result["migrated"] = migrated

    updated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    written = 0
    for entry in entries:
        kind = entry.record_kind or "연차"
        if entry.days <= 0:
            for r in _find_ledger_rows_for_period(
                ws, headers, entry.name, entry.period_label, record_kind=kind
            ):
                ws.delete_rows(r, 1)
            continue

        for r in sorted(
            _find_ledger_rows_for_period(
                ws, headers, entry.name, entry.period_label, record_kind=kind
            ),
            reverse=True,
        ):
            ws.delete_rows(r, 1)

        _write_entry_row(ws, headers, entry, updated_at)
        written += 1

    monthly_written = 0
    if ws_monthly is not None and monthly_summaries:
        m_headers = _ensure_monthly_columns(
            ws_monthly, build_header_map(ws_monthly, MONTHLY_LEAVE_ALIASES)
        )
        period_norm = normalize_period_label(payroll_period) or payroll_period
        for r in sorted(_find_monthly_rows_for_period(ws_monthly, m_headers, period_norm), reverse=True):
            ws_monthly.delete_rows(r, 1)
        for summary in monthly_summaries:
            _write_monthly_summary_row(ws_monthly, m_headers, summary, updated_at)
            monthly_written += 1

    purged = purge_leave_usage_older_than(ws, headers, payroll_period, retention_years)
    if ws_monthly is not None:
        m_headers = build_header_map(ws_monthly, MONTHLY_LEAVE_ALIASES)
        if m_headers.get("처리월"):
            cutoff = retention_cutoff_ym(payroll_period, retention_years)
            to_delete: list[int] = []
            for r in range(2, ws_monthly.max_row + 1):
                period = _row_period_ym(ws_monthly, r, m_headers)
                if period and _ym_lt(period, cutoff):
                    to_delete.append(r)
            for r in sorted(to_delete, reverse=True):
                ws_monthly.delete_rows(r, 1)

    result["written"] = written
    result["monthly_written"] = monthly_written
    result["purged"] = purged
    result["path"] = save_workbook(wb, ledger_path)
    wb.close()
    return result


def save_leave_ledger_entries(
    roster_path: Path,
    entries: list[LeaveLedgerEntry],
    payroll_period: str | None = None,
) -> int:
    """
    하위 호환 래퍼 — 실제 저장은 연차사용대장/ 폴더로 수행합니다.

    roster_path 는 최초 이관용으로만 사용합니다.
    """
    period = payroll_period or datetime.now().strftime("%Y-%m")
    info = save_leave_usage_ledger_entries(
        entries,
        period,
        roster_path_for_migration=roster_path,
    )
    return int(info.get("written") or 0)
