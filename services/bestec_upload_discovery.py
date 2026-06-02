"""
베스텍 밀양 — Downloads·카카오톡 등에서 근로자 명부·연차대장 업로드 파일 탐색·설치.

권장 파일명 (사용자 업로드 시):
  - templates/bestec/베스텍_재직증명서_2026.xlsx  (재직증명서 명부)
  - templates/bestec/연차관리_호민.xlsx (연차 통합 시트)
  - templates/bestec/근로자명부.xlsx  (씨앤엘+연차 시트, 대체)
"""

from __future__ import annotations

import re
import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

from roster_constants import norm_name_key
from roster_leave_sheet import leave_snapshot_for_period, parse_leave_annual_sheet
from roster_workbook import load_employee_roster_from_workbook
from services.bestec_leave_import import (
    apply_leave_days_to_invoices as apply_homin_leave_to_invoices,
    apply_leave_to_roster,
    load_homin_leave_for_period,
)
from services.bestec_roster_import import (
    load_bestec_certificate_roster,
    merge_bestec_roster,
    roster_stats,
)

WORKPLACE_FILTER = "밀양"
AFFILIATE_TAG = "베스텍"

_ROSTER_HINTS = ("명부", "근로자", "roster", "재직증명")
_LEAVE_HINTS = ("연차대장", "연차 대장", "연차사용대장", "연차 사용", "연차관리")
_SKIP_HINTS = ("급여대장", "근태", "p&l", "손익", "견적", "입찰")
_PERIOD_IN_NAME = re.compile(r"(20)?26[-_.]?\s*0?([1-4])\s*월?|(?:^|[_\s-])0?([1-4])월")
_MONTHLY_ROSTER = re.compile(
    r"근로자\s*명부[_\s-]*(20)?26[-_.]0?([1-4])",
    re.IGNORECASE,
)


@dataclass
class BestecUploadScan:
    downloads_dir: Path
    template_dir: Path
    roster_canonical: Path | None = None
    roster_certificate: Path | None = None
    roster_monthly: dict[str, Path] = field(default_factory=dict)
    leave_standalone: Path | None = None
    installed: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)


def bestec_template_dir(root: Path) -> Path:
    d = root / "templates" / "bestec"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _period_from_filename(name: str) -> str | None:
    m = _MONTHLY_ROSTER.search(name.replace(" ", ""))
    if m:
        mo = int(m.group(2))
        return f"2026-{mo:02d}"
    m2 = _PERIOD_IN_NAME.search(name)
    if m2:
        mo = int(m2.group(2) or m2.group(3) or 0)
        if 1 <= mo <= 4:
            return f"2026-{mo:02d}"
    return None


def _is_candidate(path: Path) -> bool:
    ext = path.suffix.lower()
    if ext not in (".xlsx", ".xlsm"):
        return False
    low = path.name.lower()
    if any(s in low for s in ("급여대장", "근태")):
        return False
    if any(s in path.name for s in _SKIP_HINTS):
        return False
    return True


def _score_roster(path: Path) -> int:
    n = path.name
    score = 0
    if AFFILIATE_TAG in n or "bestec" in n.lower():
        score += 3
    if "재직증명" in n:
        score += 8
    if "명부" in n:
        score += 5
    if "근로자" in n:
        score += 4
    if "연차" in n and "명부" not in n:
        score -= 10
    if _period_from_filename(n):
        score += 2
    return score


def _score_leave(path: Path) -> int:
    n = path.name
    score = 0
    if any(h in n for h in _LEAVE_HINTS):
        score += 8
    if "연차" in n and "대장" in n:
        score += 6
    if "명부" in n:
        score -= 5
    return score


def scan_uploads(
    downloads_dir: Path | None = None,
    template_dir: Path | None = None,
    *,
    root: Path | None = None,
    extra_dirs: list[Path] | None = None,
) -> BestecUploadScan:
    from excel_writer import TEMPLATES_DIR

    root = root or Path(__file__).resolve().parents[1]
    downloads = downloads_dir or Path(r"c:\Users\MY\Downloads")
    tdir = template_dir or bestec_template_dir(root)

    scan = BestecUploadScan(downloads_dir=downloads, template_dir=tdir)

    roster_candidates: list[Path] = []
    leave_candidates: list[Path] = []

    search_roots = [downloads, tdir, TEMPLATES_DIR]
    if extra_dirs:
        search_roots = list(extra_dirs) + search_roots
    for base in search_roots:
        if not base.is_dir():
            continue
        for path in base.rglob("*.xlsx"):
            if not _is_candidate(path):
                continue
            name = path.name
            if any(h in name for h in _LEAVE_HINTS) or (
                "연차" in name and "대장" in name and "명부" not in name
            ):
                leave_candidates.append(path)
            elif any(h in name for h in _ROSTER_HINTS) or "명부" in name:
                roster_candidates.append(path)
            elif "재직증명" in name and AFFILIATE_TAG in name:
                roster_candidates.append(path)

    # 월별 명부
    monthly: dict[str, Path] = {}
    for p in sorted(roster_candidates, key=lambda x: x.stat().st_mtime, reverse=True):
        per = _period_from_filename(p.name)
        if per and per not in monthly:
            monthly[per] = p
    scan.roster_monthly = monthly

    canonical = sorted(
        [p for p in roster_candidates if _period_from_filename(p.name) is None],
        key=lambda x: (_score_roster(x), x.stat().st_mtime),
        reverse=True,
    )
    cert_sorted = sorted(
        [p for p in roster_candidates if "재직증명" in p.name],
        key=lambda x: (_score_roster(x), x.stat().st_mtime),
        reverse=True,
    )
    if cert_sorted:
        scan.roster_certificate = cert_sorted[0]
    if canonical:
        scan.roster_canonical = canonical[0]
    elif monthly:
        scan.roster_canonical = monthly.get("2026-04") or next(iter(monthly.values()))

    leave_sorted = sorted(
        leave_candidates,
        key=lambda x: (_score_leave(x), x.stat().st_mtime),
        reverse=True,
    )
    if leave_sorted:
        scan.leave_standalone = leave_sorted[0]

    if not scan.roster_canonical and not scan.roster_monthly:
        scan.notes.append(
            "근로자 명부 파일 없음 — Downloads에 *명부*.xlsx 또는 "
            f"{tdir / '근로자명부.xlsx'} 업로드 필요"
        )
    return scan


def install_uploads(
    scan: BestecUploadScan,
    *,
    explicit_roster: Path | None = None,
    explicit_leave: Path | None = None,
) -> BestecUploadScan:
    """탐색·지정 파일을 templates/bestec/ 에 복사."""
    tdir = scan.template_dir
    tdir.mkdir(parents=True, exist_ok=True)

    if explicit_roster and explicit_roster.is_file():
        scan.roster_certificate = explicit_roster
    if explicit_leave and explicit_leave.is_file():
        scan.leave_standalone = explicit_leave

    if scan.roster_certificate:
        dest = tdir / "베스텍_재직증명서_2026.xlsx"
        src = scan.roster_certificate
        if not dest.exists() or src.stat().st_mtime > dest.stat().st_mtime:
            shutil.copy2(src, dest)
            scan.installed.append(str(dest))
        scan.roster_certificate = dest

    if scan.roster_canonical and scan.roster_canonical != scan.roster_certificate:
        dest = tdir / "근로자명부.xlsx"
        if not dest.exists() or scan.roster_canonical.stat().st_mtime > dest.stat().st_mtime:
            shutil.copy2(scan.roster_canonical, dest)
            scan.installed.append(str(dest))
        scan.roster_canonical = dest

    for period, src in scan.roster_monthly.items():
        dest = tdir / f"근로자명부_{period}.xlsx"
        if src.resolve() == dest.resolve():
            scan.roster_monthly[period] = dest
            continue
        if not dest.exists() or src.stat().st_mtime > dest.stat().st_mtime:
            shutil.copy2(src, dest)
            scan.installed.append(str(dest))
        scan.roster_monthly[period] = dest

    if scan.leave_standalone:
        dest = tdir / "연차관리_호민.xlsx"
        src = scan.leave_standalone
        if not dest.exists() or src.stat().st_mtime > dest.stat().st_mtime:
            shutil.copy2(src, dest)
            scan.installed.append(str(dest))
        scan.leave_standalone = dest

    return scan


def roster_path_for_period(scan: BestecUploadScan, period: str) -> Path | None:
    if period in scan.roster_monthly and scan.roster_monthly[period].is_file():
        return scan.roster_monthly[period]
    if scan.roster_canonical and scan.roster_canonical.is_file():
        return scan.roster_canonical
    return None


def _filter_milyang(roster: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for key, rec in roster.items():
        wp = str(rec.get("근무지") or "").strip()
        aff = str(rec.get("계열사") or "").strip()
        if wp and WORKPLACE_FILTER not in wp and AFFILIATE_TAG not in wp:
            if aff and AFFILIATE_TAG not in aff:
                continue
        if wp and WORKPLACE_FILTER not in wp and not aff:
            continue
        out[key] = rec
    return out or roster


def load_bestec_roster(
    scan: BestecUploadScan,
    period: str,
    *,
    reference_fallback: dict[str, dict[str, Any]] | None = None,
    reference: dict | None = None,
) -> tuple[dict[str, dict[str, Any]], str]:
    """재직증명서 명부 + 참조 급여대장 명부값 병합."""
    source = "참조 급여대장 추출"
    merged: dict[str, dict[str, Any]] = {}

    cert_path = scan.roster_certificate
    if cert_path and cert_path.is_file() and reference:
        cert = load_bestec_certificate_roster(cert_path)
        if cert:
            merged = merge_bestec_roster(cert, reference)
            stats = roster_stats(merged)
            source = (
                f"재직증명서+참조대장 ({cert_path.name}, "
                f"{stats['cert_matched']}명 매칭, 시급 {stats['with_hourly_rate']}명)"
            )
        elif cert:
            merged = _filter_milyang(cert)
            source = f"재직증명서 ({cert_path.name})"

    path = roster_path_for_period(scan, period)
    if path and path.is_file() and path != cert_path:
        loaded = load_employee_roster_from_workbook(path, period_hint=period)
        loaded = _filter_milyang(loaded)
        if loaded:
            for key, rec in loaded.items():
                if key not in merged:
                    merged[key] = rec
            if "재직증명" not in source:
                source = f"업로드 명부 ({path.name})"

    if reference_fallback:
        for key, rec in reference_fallback.items():
            if key not in merged:
                merged[key] = rec

    if not merged and reference_fallback:
        merged = dict(reference_fallback)
        source = "참조 급여대장 추출 (명부 미업로드)"

    return merged, source


def load_bestec_leave_for_period(
    scan: BestecUploadScan,
    period: str,
    roster: dict[str, dict[str, Any]],
) -> tuple[dict[str, dict[str, Any]], int]:
    """연차관리(호민) 통합 시트 → 당월 사용일 (청구서용)."""
    leave_path = scan.leave_standalone
    if not leave_path or not leave_path.is_file():
        return {}, 0
    leave_by_key = load_homin_leave_for_period(leave_path, period)
    applied = sum(1 for v in leave_by_key.values() if v.get("leave_days", 0) > 0)
    return leave_by_key, applied


def _period_month_labels(period: str) -> tuple[str, str, str]:
    yy = f"{int(period.split('-')[0]) % 100:02d}"
    mm = f"{int(period.split('-')[1]):02d}"
    return yy, mm, f"{yy}.{mm}"


def monthly_leave_days_from_roster(
    roster: dict[str, dict[str, Any]],
    period: str,
) -> dict[str, float]:
    """명부 로드 시 병합된 _monthly_leave_usage → {이름키: 당월일수}."""
    _, mm, label = _period_month_labels(period)
    alt = f"{label.split('.')[0]}.{int(mm)}"
    out: dict[str, float] = {}
    seen: set[str] = set()
    for rec in roster.values():
        if not isinstance(rec, dict):
            continue
        name = str(rec.get("성명") or "").strip()
        if not name:
            continue
        key = norm_name_key(name)
        if key in seen:
            continue
        seen.add(key)
        monthly = rec.get("_monthly_leave_usage") or {}
        days = float(monthly.get(label) or monthly.get(alt) or 0)
        if days <= 0:
            for lab, val in monthly.items():
                if lab.endswith(f".{mm}") or lab.endswith(f".{int(mm)}"):
                    days = max(days, float(val))
        if days > 0:
            out[key] = days
    return out


def monthly_leave_days_from_workbook(path: Path, period: str) -> dict[str, float]:
    """연차 시트 월별 사용일수 → {이름키: 당월일수} (명부 파일 직접 파싱, 느릴 수 있음)."""
    if not path.is_file():
        return {}
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        snaps = parse_leave_annual_sheet(wb)
        wb.close()
    except Exception:
        return {}

    _, _, label = _period_month_labels(period)
    yy, mm, _ = _period_month_labels(period)
    alt = f"{yy}.{int(mm)}"

    out: dict[str, float] = {}
    for key, snap in snaps.items():
        view = leave_snapshot_for_period(snap, period)
        monthly = view.get("_monthly_leave_usage") or {}
        days = float(monthly.get(label) or monthly.get(alt) or 0)
        if days > 0:
            out[key] = days
    return out


def apply_leave_days_to_invoices(
    invoice_rows: list[dict[str, Any]],
    leave_by_key: dict[str, float],
) -> int:
    n = 0
    for inv in invoice_rows:
        key = norm_name_key(inv.get("name"))
        days = leave_by_key.get(key)
        if days is not None and days > 0:
            inv["leave_days"] = days
            n += 1
    return n


def merge_leave_workbook_into_roster_path(
    roster_path: Path,
    leave_path: Path,
) -> bool:
    """독립 연차대장의 연차 시트를 명부 파일에 없으면 복사하지 않음 — 동일 파일이면 스킵."""
    if not leave_path.is_file() or leave_path.resolve() == roster_path.resolve():
        return False
    try:
        lw = openpyxl.load_workbook(leave_path, read_only=True)
        if "연차" not in lw.sheetnames:
            lw.close()
            return False
        lw.close()
    except Exception:
        return False
    return True


def expected_upload_doc() -> str:
    return (
        "권장 업로드 위치: `급여프로그램/templates/bestec/`\n"
        "- `베스텍_재직증명서_2026.xlsx` (재직증명서 근로자명부)\n"
        "- `연차관리_호민.xlsx` (연차 통합 시트)\n"
        "- `근로자명부.xlsx` (씨엔엘/연차 시트, 대체)\n"
        "또는 `Downloads`·`카카오톡 받은 파일`에 **재직증명**, **연차관리**, **베스텍** 포함"
    )
