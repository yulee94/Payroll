from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from invoice_parser import extract_invoice_data, find_billing_column_layout


class InvoiceParserKeywordTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmpdir = tempfile.TemporaryDirectory()
        self._root = Path(self._tmpdir.name)

    def tearDown(self) -> None:
        self._tmpdir.cleanup()

    def test_extracts_reordered_invoice_columns_by_header_keywords(self) -> None:
        path = self._root / "keyword_invoice.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "급여자료"
        headers = [
            "직원명",
            "현장부서",
            "교통수당",
            "총지급액",
            "기본급",
            "연장시간",
            "연장수당",
            "건강보험료",
            "장기요양보험",
            "국민연금",
            "고용보험",
            "조퇴분",
            "무급일",
            "연차사용",
            "특근연장수당",
            "특근연장시간",
            "교대수당",
            "교대시간",
            "심야수당",
            "심야시간",
            "직책수당",
        ]
        for col, title in enumerate(headers, 1):
            ws.cell(3, col).value = title
        values = [
            "홍길동",
            "생산1",
            50_000,
            3_000_000,
            2_500_000,
            2,
            36_000,
            100_000,
            12_000,
            90_000,
            27_000,
            30,
            1,
            0.5,
            8_000,
            1,
            2_000,
            0.5,
            6_000,
            2,
            100_000,
        ]
        for col, value in enumerate(values, 1):
            ws.cell(4, col).value = value
        wb.save(path)
        wb.close()

        rows = extract_invoice_data(path)

        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row["name"], "홍길동")
        self.assertEqual(row["dept"], "생산1")
        self.assertEqual(row["gross_pay"], 3_000_000)
        self.assertEqual(row["subtotal"], 2_950_000)
        self.assertEqual(row["transport"], 50_000)
        self.assertEqual(row["base_salary"], 2_500_000)
        self.assertAlmostEqual(row["ot_hours"], 2)
        self.assertEqual(row["ot_pay"], 36_000)
        self.assertAlmostEqual(row["early_leave_hours"], 0.5)
        self.assertEqual(row["health_insurance"], 100_000)
        self.assertEqual(row["long_term_care"], 12_000)
        self.assertEqual(row["national_pension"], 90_000)
        self.assertEqual(row["employment_insurance"], 27_000)

    def test_embedded_attendance_sheet_can_override_invoice_early_leave_by_keywords(self) -> None:
        path = self._root / "keyword_invoice_with_attendance.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "청구"
        for col, title in enumerate(("성명", "소계", "조퇴시간"), 1):
            ws.cell(4, col).value = title
        ws.cell(5, 1).value = "김철수"
        ws.cell(5, 2).value = 1_000_000
        ws.cell(5, 3).value = 0

        att = wb.create_sheet("월별 근태")
        att.cell(2, 1).value = "직원명"
        att.cell(2, 2).value = "조퇴분"
        att.cell(3, 1).value = "김철수"
        att.cell(3, 2).value = 30
        wb.save(path)
        wb.close()

        rows = extract_invoice_data(path)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["name"], "김철수")
        self.assertAlmostEqual(rows[0]["early_leave_hours"], 0.5)

    def test_layout_falls_back_to_legacy_columns_without_detectable_headers(self) -> None:
        wb = Workbook()
        ws = wb.active
        layout = find_billing_column_layout(ws)
        wb.close()

        self.assertEqual(layout.source, "legacy-fixed")
        self.assertEqual(layout.data_start_row, 5)


if __name__ == "__main__":
    unittest.main()
