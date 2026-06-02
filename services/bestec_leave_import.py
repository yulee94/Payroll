"""
베스텍 ○ 연차관리(호민) 통합 시트 — 월별 연차 사용일 파싱.

'통합' 시트: 성명 + 월별(발생/사용) 열 + 연차발생/사용/잔여
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl

from roster_constants import norm_name_key
from utils import safe_number

TONGHAP_SHEET = "통합"
_MONTH_RE = re.compile(r"^(\d{1,2})\s*월$")


def _parse_usage(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("일", "").replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _month_usage_columns(ws, *, payroll_month: int) -> tuple[int | None, int | None]:
    """
    통합 시트에서 해당 급여월 '사용' 열 찾기.

    헤더 1행: … 10월, (빈), 11월, … 12월, … 1월, …
    헤더 2행: … 발생, 사용, …
    """
    month_cols: list[tuple[int, str]] = []
    for col in range(1, ws.max_column + 1):
        h1 = ws.cell(1, col).value
        if h1 is None:
            continue
        text = str(h1).strip().replace("\n", "")
        m = _MONTH_RE.match(text)
        if m:
            month_cols.append((col, text))

    if not month_cols:
        return None, None

    # 동일 월 표기가 2번(전년 10~12 + 당해 1~12) 있을 수 있음 → 두 번째 블록(1~9월) 우선
    candidates = [c for c, label in month_cols if int(_MONTH_RE.match(label).group(1)) == payroll_month]
    if not candidates:
        return None, None
    use_col = candidates[-1] + 1
    if use_col > ws.max_column:
        return None, None
    sub = str(ws.cell(2, use_col).value or "").strip()
    if sub != "사용":
        # 발생/사용 순서가 바뀐 경우 보정
        if str(ws.cell(2, candidates[-1]).value or "").strip() == "사용":
            use_col = candidates[-1]
        else:
            return None, None
    accrued_col = use_col - 1 if use_col > 1 else None
    return accrued_col, use_col


def _summary_columns(ws) -> dict[str, int]:
    out: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        h = ws.cell(1, col).value
        if h is None:
            continue
        text = str(h).strip().replace("\n", "")
        if text in ("연차발생", "연차 발생"):
            out["발생연차"] = col
        elif text in ("연차사용", "연차 사용"):
            out["사용연차"] = col
        elif text in ("잔여연차", "잔여 연차"):
            out["잔여연차"] = col
        elif text == "통상시급":
            out["통상시급"] = col
    return out


def load_homin_leave_for_period(path: Path, period: str) -> dict[str, dict[str, Any]]:
    """
    {이름키: {leave_days, 발생연차, 사용연차, 잔여연차, 통상시급}}.
    """
    if not path.is_file():
        return {}
    try:
        month = int(period.split("-")[1])
    except (IndexError, ValueError):
        return {}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if TONGHAP_SHEET not in wb.sheetnames:
            return {}
        ws = wb[TONGHAP_SHEET]
        accrued_col, use_col = _month_usage_columns(ws, payroll_month=month)
        summary = _summary_columns(ws)
        if use_col is None:
            return {}

        out: dict[str, dict[str, Any]] = {}
        empty_streak = 0
        for row_idx in range(3, min(ws.max_row or 0, 800) + 1):
            name = ws.cell(row_idx, 1).value
            if name is None or not str(name).strip():
                empty_streak += 1
                if empty_streak >= 30:
                    break
                continue
            empty_streak = 0
            name_s = str(name).strip()
            key = norm_name_key(name_s)
            if not key:
                continue

            leave_days = _parse_usage(ws.cell(row_idx, use_col).value)
            rec: dict[str, Any] = {"leave_days": leave_days, "성명": name_s}
            if accrued_col:
                rec["당월발생"] = _parse_usage(ws.cell(row_idx, accrued_col).value)
            for field, col in summary.items():
                val = ws.cell(row_idx, col).value
                if val not in (None, ""):
                    rec[field] = safe_number(val, 0.0)
            out[key] = rec
        return out
    finally:
        wb.close()


def apply_leave_to_roster(
    roster: dict[str, dict[str, Any]],
    leave_by_key: dict[str, dict[str, Any]],
    period: str,
) -> int:
    """명부에 연차 잔여·발생·사용 반영. 반영 인원 수."""
    n = 0
    for key, leave_rec in leave_by_key.items():
        emp = roster.get(key)
        if not emp:
            continue
        if leave_rec.get("leave_days", 0) > 0:
            emp["연차사용메모"] = f"{period}:{leave_rec['leave_days']:g}일"
        for field in ("발생연차", "사용연차", "잔여연차", "통상시급"):
            if field in leave_rec and leave_rec[field]:
                emp[field] = leave_rec[field]
        n += 1
    return n


def apply_leave_days_to_invoices(
    invoice_rows: list[dict[str, Any]],
    leave_by_key: dict[str, dict[str, Any]],
    *,
    reference_hours: dict[str, float] | None = None,
    standard_hours: float = 209.0,
    derive_from_reference_hours: bool = False,
) -> int:
    """청구서 leave_days 반영 (연차대장 통합 시트 기준)."""
    n = 0
    reference_hours = reference_hours or {}
    for inv in invoice_rows:
        key = norm_name_key(inv.get("name"))
        leave_rec = leave_by_key.get(key, {})
        days = safe_number(leave_rec.get("leave_days"), 0.0)

        if days <= 0 and derive_from_reference_hours:
            ref_h = safe_number(reference_hours.get(key), 0.0)
            if ref_h > 0 and ref_h < standard_hours - 0.5:
                days = round((standard_hours - ref_h) / 8.0, 4)

        if days > 0:
            inv["leave_days"] = days
            inv["leave_sheet_leave_days"] = days
            n += 1
    return n
