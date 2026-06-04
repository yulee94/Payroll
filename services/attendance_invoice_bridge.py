"""
attendance_invoice_bridge.py - build invoice-compatible workbooks from attendance.

The payroll engine still writes the same ledger/payslip/payment outputs. This
bridge simply gives attendance-only or external attendance uploads the workbook
shape that the existing invoice parser already understands.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from invoice_parser import COL, DATA_START_ROW
from services.attendance_import import AttendanceImportResult


BILLING_SHEET_NAME = "청구내역_근태기반"
ATTENDANCE_SHEET_NAME = "근태_외부업로드"


def _safe_sheet_title(title: str) -> str:
    cleaned = "".join("_" if ch in "[]:*?/\\" else ch for ch in title).strip()
    return (cleaned or "sheet")[:31]


def _write_billing_headers(ws) -> None:
    headers = {
        COL["dept"]: "소속",
        COL["name"]: "성명",
        COL["hire"]: "입사일",
        COL["base_hourly"]: "기본시급",
        COL["ordinary_hourly"]: "통상시급",
        COL["base_days"]: "기준시간",
        COL["work_days"]: "근무시간",
        COL["unpaid"]: "무급/결근",
        COL["leave"]: "휴가/연차",
        COL["ot_hours"]: "O/T(150%)",
        COL["night_hours"]: "심야(50%)",
        COL["special_hours"]: "특근(150%)",
        COL["special_ext_hours"]: "특근연장(50%)",
        COL["early_leave"]: "지조외",
        COL["base_salary"]: "기본급",
        COL["base_deduction"]: "기본공제",
        COL["ot_pay"]: "O/T수당",
        COL["night_pay"]: "심야수당",
        COL["special_pay"]: "특근수당",
        COL["special_ext_pay"]: "특근연장",
        COL["position_pay"]: "직책수당",
        COL["shift_pay"]: "교대수당",
        COL["subtotal"]: "소계",
        COL["transport"]: "교통비",
        COL["health"]: "건강보험",
        COL["long_term_care"]: "장기요양",
        COL["pension"]: "국민연금",
        COL["employment"]: "고용보험",
    }
    for col, title in headers.items():
        ws.cell(4, col).value = title


def _write_invoice_row(ws, row_no: int, row: dict[str, Any]) -> None:
    ws.cell(row_no, COL["no"]).value = row_no - DATA_START_ROW + 1
    ws.cell(row_no, COL["dept"]).value = row.get("dept", "")
    ws.cell(row_no, COL["name"]).value = row.get("name", "")
    ws.cell(row_no, COL["hire"]).value = row.get("hire_date", "")
    for key in (
        "base_hourly",
        "ordinary_hourly",
        "base_days",
        "work_days",
        "unpaid",
        "leave",
        "ot_hours",
        "night_hours",
        "special_hours",
        "special_ext_hours",
        "early_leave",
        "base_salary",
        "base_deduction",
        "ot_pay",
        "night_pay",
        "special_pay",
        "special_ext_pay",
        "position_pay",
        "shift_pay",
        "transport",
        "health",
        "long_term_care",
        "pension",
        "employment",
    ):
        source_key = {
            "unpaid": "unpaid_days",
            "leave": "leave_days",
            "early_leave": "early_leave_hours",
            "health": "health_insurance",
            "long_term_care": "long_term_care",
            "pension": "national_pension",
            "employment": "employment_insurance",
        }.get(key, key)
        ws.cell(row_no, COL[key]).value = row.get(source_key, 0)

    subtotal = int(row.get("subtotal") or 0)
    # The invoice parser skips rows whose subtotal is 0. Attendance-based rows
    # are recalculated from the roster later, so keep a minimal include marker.
    ws.cell(row_no, COL["subtotal"]).value = subtotal if subtotal > 0 else 1


def build_attendance_invoice_workbook(
    attendance_result: AttendanceImportResult,
    target_path: Path,
    *,
    period: str,
    workplace: str = "",
) -> Path:
    """Create an invoice-shaped workbook from aggregated attendance rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_title(BILLING_SHEET_NAME)
    ws.cell(1, 1).value = f"{period} 근태기반 급여입력"
    ws.cell(2, 1).value = workplace
    _write_billing_headers(ws)
    for offset, row in enumerate(attendance_result.invoice_rows):
        _write_invoice_row(ws, DATA_START_ROW + offset, row)
    _write_attendance_sheet(wb, attendance_result.invoice_rows)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target_path)
    wb.close()
    return target_path


def _write_attendance_sheet(wb, rows: list[dict[str, Any]]) -> None:
    title = _safe_sheet_title(ATTENDANCE_SHEET_NAME)
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws.cell(1, 1).value = "성명"
    ws.cell(1, 2).value = "지조외"
    ws.cell(1, 3).value = "근무시간"
    for idx, row in enumerate(rows, 2):
        ws.cell(idx, 1).value = row.get("name", "")
        ws.cell(idx, 2).value = row.get("early_leave_hours", 0)
        ws.cell(idx, 3).value = row.get("work_days", 0)


def attach_attendance_sheet(
    invoice_path: Path,
    attendance_result: AttendanceImportResult,
    target_path: Path,
) -> Path:
    """Copy an invoice workbook and add an external attendance sheet to it."""
    wb = load_workbook(invoice_path)
    _write_attendance_sheet(wb, attendance_result.invoice_rows)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target_path)
    wb.close()
    return target_path
