"""
5월 기준 templates/근로자명부.xlsx 에서 2026-01 ~ 2026-05 월별 명부 생성.

실행 (급여프로그램 폴더에서):
  python scripts/build_monthly_roster_files.py
"""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from excel_writer import TEMPLATES_DIR
from roster_constants import ROSTER_HEADER_ALIASES, build_header_map, norm_name_key
from roster_leave_sheet import leave_snapshot_for_period, parse_leave_annual_sheet
from roster_workbook import find_main_roster_sheet
from services.employee_roster_store import MONTHLY_ROSTER_PREFIX, canonical_roster_path

PERIODS = [f"2026-{m:02d}" for m in range(1, 6)]


def _write_leave_values(ws, headers: dict[str, int], row: int, leave: dict) -> None:
    for col_name in ("발생연차", "사용연차", "잔여연차"):
        if col_name not in headers:
            continue
        val = leave.get(col_name)
        if val is None:
            continue
        ws.cell(row, headers[col_name], val)


def build_monthly_files(source: Path | None = None) -> list[Path]:
    src = source or canonical_roster_path()
    if not src.is_file():
        raise FileNotFoundError(f"명부 없음: {src}")

    wb = openpyxl.load_workbook(src, data_only=True)
    leave_snaps = parse_leave_annual_sheet(wb)
    wb.close()

    created: list[Path] = []
    TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)

    for period in PERIODS:
        dest = TEMPLATES_DIR / f"{MONTHLY_ROSTER_PREFIX}{period}.xlsx"
        shutil.copy2(src, dest)

        out = openpyxl.load_workbook(dest)
        ws = find_main_roster_sheet(out)
        headers = build_header_map(ws, ROSTER_HEADER_ALIASES)
        if "성명" not in headers:
            out.close()
            continue

        for r in range(2, min(ws.max_row or 0, 5000) + 1):
            name = ws.cell(r, headers["성명"]).value
            if name is None or not str(name).strip():
                continue
            snap = leave_snaps.get(norm_name_key(str(name).strip()))
            if snap is None:
                continue
            leave = leave_snapshot_for_period(snap, period)
            _write_leave_values(ws, headers, r, leave)

        out.save(dest)
        out.close()
        created.append(dest)
        print(f"생성: {dest.name}")

    return created


if __name__ == "__main__":
    build_monthly_files()
