"""

services/monthly_report.py - 월별 임원 요약 보고 (UI + Excel)

"""



from __future__ import annotations



from datetime import datetime

from pathlib import Path

from typing import Any



from openpyxl import Workbook

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1



from core.brand_display import company_name_line
from core.config import APP_CONFIG

from payroll_archive import (
    MonthSummary,
    format_executive_report_title,
    format_period_display,
    load_snapshot_records,
)
from services.executive_analytics import load_year_to_date_series

from payroll_comparison import prev_period_label

from payroll_archive import build_month_summary

from services.org_registry import enrich_records, group_by_workplace, summarize_records





def _fmt_won(won: int) -> str:

    if won >= 100_000_000:

        return f"{won / 100_000_000:.1f}억원"

    if won >= 10_000:

        return f"{won / 10_000:,.0f}만원"

    return f"{won:,}원"





def _fmt_delta(value: int, *, is_money: bool = False) -> str:

    if value == 0:

        return "전월과 동일"

    if is_money:

        sign = "+" if value > 0 else "-"

        return f"전월 대비 {sign}{_fmt_won(abs(value))}"

    sign = "+" if value > 0 else ""

    return f"전월 대비 {sign}{value}명"


def _section(title: str) -> list[str]:
    return ["", title, ""]


def _site_table_lines(records: list[dict[str, Any]], total_gross: int) -> list[str]:
    sites = group_by_workplace(enrich_records(records))
    if not sites:
        return []
    lines = ["  사업장·소속          인원    총 인건비        비중", "  " + "─" * 52]
    for wp, site_rows in sites[:12]:
        sm = summarize_records(site_rows)
        share = f"{(sm.total_gross / total_gross * 100):.0f}%" if total_gross else "-"
        name = (wp or "-")[:18].ljust(18)
        lines.append(
            f"  {name}  {sm.employee_count:>4}명  {_fmt_won(sm.total_gross):>12}  {share:>5}"
        )
    if len(sites) > 12:
        lines.append(f"  · 외 {len(sites) - 12}개 사업장·소속")
    return lines


def build_executive_report_text(

    period: str,

    summary: MonthSummary,

    records: list[dict[str, Any]],

) -> str:

    brand = APP_CONFIG.brand

    prior = build_month_summary(prev_period_label(period))

    hd = summary.employee_count - (prior.employee_count if prior.has_output else 0)

    gd = summary.total_gross - (prior.total_gross if prior.has_output else 0)



    title = format_executive_report_title(period)
    lines: list[str] = [
        "═" * 54,
        f"  {company_name_line()}",
        f"  {title}",
        f"  작성 {datetime.now():%Y-%m-%d %H:%M}",
        "═" * 54,
        *_section("【 경영 요약 】"),
        f"  ① 인원 {summary.employee_count}명 — {_fmt_delta(hd)}",
        f"  ② 총 인건비 {_fmt_won(summary.total_gross)} — {_fmt_delta(gd, is_money=True)}",
    ]

    sites = group_by_workplace(enrich_records(records))

    if sites:

        wp, rows = sites[0]

        sm = summarize_records(rows)

        share = (sm.total_gross / summary.total_gross * 100) if summary.total_gross else 0

        lines.append(

            f"  ③ 최대 부담 「{wp}」 {sm.employee_count}명 · {_fmt_won(sm.total_gross)} (약 {share:.0f}%)"

        )

    lines.extend(
        _section("【 핵심 지표 】")
        + [
            f"  실수령액     {_fmt_won(summary.total_net)}",
            f"  공제합계     {_fmt_won(summary.total_deduction)}",
            f"  연차 사용    {summary.leave_users}명",
            f"  무급/결근    {summary.absence_users}명",
        ]
    )

    site_lines = _site_table_lines(records, summary.total_gross)
    if site_lines:
        lines.extend(_section("【 사업장·소속별 】") + site_lines)

    if summary.has_comparison:

        lines.extend(_section("【 전월 대비 】") + ["  급여차이 보고서가 생성되었습니다."])

    ytd_label, ytd_months, ytd_deltas, ytd_total = load_year_to_date_series(period)
    if ytd_months:
        lines.append("")
        lines.append(f"【 연간 보고 · {ytd_label} 】")
        lines.append(f"  누적 총급여  {_fmt_won(ytd_total)}")
        for pt in ytd_months:
            lines.append(f"  · {pt.label}  {_fmt_won(pt.gross)}  ({pt.headcount}명)")
        if len(ytd_deltas) > 1:
            lines.append("  전월 대비 총급여")
            for d in ytd_deltas[1:]:
                sign = "+" if d.delta >= 0 else "-"
                lines.append(f"    {d.label}  {sign}{_fmt_won(abs(d.delta))}")
        lines.append("")

    top_gross = sorted(records, key=lambda r: int(r.get("gross_pay") or 0), reverse=True)[:5]

    if top_gross:

        lines.extend(_section("【 총지급 상위 】"))

        for i, r in enumerate(top_gross, 1):

            lines.append(

                f"  {i}. {r.get('name','')} ({r.get('workplace','')}) — "

                f"{_fmt_won(int(r.get('gross_pay') or 0))}"

            )

        lines.append("")



    leave_rows = [r for r in records if float(r.get("leave_days") or 0) > 0]

    if leave_rows:

        lines.extend(_section("【 당월 연차 사용 】"))

        for r in leave_rows[:10]:

            lines.append(

                f"  · {r.get('name','')} — {r.get('leave_days')}일 "

                f"({r.get('leave_usage_display') or '-'})"

            )

        if len(leave_rows) > 10:

            lines.append(f"  · 외 {len(leave_rows) - 10}명")

        lines.append("")



    abs_rows = [r for r in records if float(r.get("unpaid_days") or 0) > 0]

    if abs_rows:

        lines.extend(_section("【 당월 무급/결근 】"))

        for r in abs_rows[:10]:

            lines.append(f"  · {r.get('name','')} — {r.get('unpaid_days')}일")

        if len(abs_rows) > 10:

            lines.append(f"  · 외 {len(abs_rows) - 10}명")



    return "\n".join(lines).rstrip()





def export_monthly_report_excel(

    period: str,

    summary: MonthSummary,

    records: list[dict[str, Any]],

    output_path: Path,

) -> Path:

    """임원 보고용 월별 요약 Excel (1페이지 요약 + 상세 시트)."""

    brand = APP_CONFIG.brand

    output_path.parent.mkdir(parents=True, exist_ok=True)



    wb = Workbook()

    navy = brand.primary_navy.lstrip("#")

    fill_hdr = PatternFill("solid", fgColor=navy)

    fill_title = PatternFill("solid", fgColor=navy)

    font_hdr = Font(name="맑은 고딕", bold=True, color="FFFFFF")

    font_title = Font(name="맑은 고딕", size=16, bold=True, color="FFFFFF")

    font_sub = Font(name="맑은 고딕", size=11, bold=True)

    font_body = Font(name="맑은 고딕", size=10)



    prior = build_month_summary(prev_period_label(period))

    hd = summary.employee_count - (prior.employee_count if prior.has_output else 0)

    gd = summary.total_gross - (prior.total_gross if prior.has_output else 0)



    # --- 경영 요약 (1페이지) ---

    ws = wb.active

    ws.title = "경영요약"

    ws.merge_cells("A1:F1")

    c = ws["A1"]

    c.value = format_executive_report_title(period)

    c.font = font_title

    c.fill = fill_title

    c.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 36



    ws.merge_cells("A2:F2")

    ws["A2"] = company_name_line()

    ws["A2"].font = Font(name="맑은 고딕", size=10, color="64748B")

    ws["A2"].alignment = Alignment(horizontal="left")



    bullets = [

        f"① 인원 {summary.employee_count}명 — {_fmt_delta(hd)}",

        f"② 총 인건비 {_fmt_won(summary.total_gross)} — {_fmt_delta(gd, is_money=True)}",

    ]

    sites = group_by_workplace(enrich_records(records))

    if sites:

        wp, rows = sites[0]

        sm = summarize_records(rows)

        share = (sm.total_gross / summary.total_gross * 100) if summary.total_gross else 0

        bullets.append(

            f"③ 최대 부담 「{wp}」 {sm.employee_count}명 · {_fmt_won(sm.total_gross)} (약 {share:.0f}%)"

        )

    row = 4

    for b in bullets:

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

        cell = ws.cell(row, 1, b)

        cell.font = font_body

        cell.alignment = Alignment(wrap_text=True)

        row += 1



    row += 1

    kpi_headers = ["구분", "당월", "전월 대비"]

    for col, h in enumerate(kpi_headers, 1):

        cell = ws.cell(row, col, h)

        cell.fill = fill_hdr

        cell.font = font_hdr

        cell.alignment = Alignment(horizontal="center")

    row += 1

    for label, val, delta in [

        ("인원", f"{summary.employee_count}명", _fmt_delta(hd)),

        ("총 인건비", _fmt_won(summary.total_gross), _fmt_delta(gd, is_money=True)),

        ("실수령", _fmt_won(summary.total_net), ""),

        ("연차 사용자", f"{summary.leave_users}명", ""),

        ("무급/결근", f"{summary.absence_users}명", ""),

    ]:

        ws.cell(row, 1, label).font = font_sub

        ws.cell(row, 2, val)

        ws.cell(row, 3, delta)

        row += 1



    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    ws.cell(row, 1, "사업장·소속별 집계").font = font_sub

    row += 1

    site_headers = ["사업장·소속", "인원", "총 인건비", "비중"]

    for col, h in enumerate(site_headers, 1):

        cell = ws.cell(row, col, h)

        cell.fill = fill_hdr

        cell.font = font_hdr

        cell.alignment = Alignment(horizontal="center")

    row += 1

    for wp, site_rows in sites[:12]:

        sm = summarize_records(site_rows)

        share = f"{(sm.total_gross / summary.total_gross * 100):.0f}%" if summary.total_gross else "-"

        ws.cell(row, 1, wp)

        ws.cell(row, 2, sm.employee_count)

        gross_cell = ws.cell(row, 3, sm.total_gross)
        gross_cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1

        ws.cell(row, 4, share)

        row += 1



    for col, w in zip("ABCDEF", [14, 14, 16, 14, 14, 14]):

        ws.column_dimensions[col].width = w

    ytd_label, ytd_months, ytd_deltas, ytd_total = load_year_to_date_series(period)
    if ytd_months:
        row += 2
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row, 1, f"연간 보고 ({ytd_label})").font = font_sub
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row, 1, f"누적 총급여 {_fmt_won(ytd_total)}").font = font_body
        row += 1
        ytd_headers = ["월", "인원", "총급여", "전월 대비"]
        for col, h in enumerate(ytd_headers, 1):
            cell = ws.cell(row, col, h)
            cell.fill = fill_hdr
            cell.font = font_hdr
            cell.alignment = Alignment(horizontal="center")
        row += 1
        for i, pt in enumerate(ytd_months):
            delta_txt = "-"
            if i < len(ytd_deltas):
                d = ytd_deltas[i]
                if d.delta == 0 and i == 0:
                    delta_txt = "(기준)"
                elif d.delta != 0:
                    sign = "+" if d.delta > 0 else "-"
                    delta_txt = f"{sign}{_fmt_won(abs(d.delta))}"
            ws.cell(row, 1, pt.label)
            ws.cell(row, 2, pt.headcount)
            gcell = ws.cell(row, 3, pt.gross)
            gcell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
            ws.cell(row, 4, delta_txt)
            row += 1

    # --- 인원별 ---

    ws2 = wb.create_sheet("인원별")

    headers = [

        "성명", "계열사", "사업장", "부서", "총지급", "실수령", "공제합계",

        "연차", "무급/결근", "연차표기",

    ]

    for c, h in enumerate(headers, 1):

        cell = ws2.cell(1, c, h)

        cell.fill = fill_hdr

        cell.font = font_hdr

        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(sorted(records, key=lambda x: str(x.get("name") or "")), 2):

        ws2.cell(i, 1, r.get("name", ""))

        ws2.cell(i, 2, r.get("affiliate", ""))

        ws2.cell(i, 3, r.get("workplace", ""))

        ws2.cell(i, 4, r.get("dept", ""))

        gc = ws2.cell(i, 5, int(r.get("gross_pay") or 0))
        gc.number_format = FORMAT_NUMBER_COMMA_SEPARATED1

        ws2.cell(i, 6, int(r.get("net_pay") or 0))

        ws2.cell(i, 7, int(r.get("total_deduction") or 0))

        ws2.cell(i, 8, r.get("leave_days", 0))

        ws2.cell(i, 9, r.get("unpaid_days", 0))

        ws2.cell(i, 10, r.get("leave_usage_display", ""))



    for col, w in zip("ABCDEFGHIJ", [10, 12, 12, 10, 12, 12, 12, 8, 10, 22]):

        ws2.column_dimensions[col].width = w



    # --- 사업장·소속별 ---

    ws3 = wb.create_sheet("사업장별")

    site_headers = ["계열사", "사업장·소속", "인원", "총지급", "실수령", "공제합계", "연차", "무급/결근"]

    for c, h in enumerate(site_headers, 1):

        cell = ws3.cell(1, c, h)

        cell.fill = fill_hdr

        cell.font = font_hdr

        cell.alignment = Alignment(horizontal="center")

    for i, (wp, site_rows) in enumerate(group_by_workplace(enrich_records(records)), 2):

        sm = summarize_records(site_rows)

        aff = site_rows[0].get("affiliate", "") if site_rows else ""

        ws3.cell(i, 1, aff)

        ws3.cell(i, 2, wp)

        ws3.cell(i, 3, sm.employee_count)

        ws3.cell(i, 4, sm.total_gross)

        ws3.cell(i, 5, sm.total_net)

        ws3.cell(i, 6, sm.total_deduction)

        ws3.cell(i, 7, sm.leave_users)

        ws3.cell(i, 8, sm.absence_users)

    for col, w in zip("ABCDEFGH", [12, 16, 8, 12, 12, 12, 8, 10]):

        ws3.column_dimensions[col].width = w



    wb.save(output_path)

    wb.close()

    return output_path





def get_or_create_report_path(period: str, reports_dir: Path) -> Path:

    return reports_dir / f"{period}_월별요약보고.xlsx"





def build_report_bundle(period: str, summary: MonthSummary) -> tuple[str, list[dict[str, Any]]]:

    records = enrich_records(load_snapshot_records(period))

    text = build_executive_report_text(period, summary, records)

    return text, records


