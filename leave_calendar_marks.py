"""
leave_calendar_marks.py - 연차사용대장 일별 표시(● ◐ ✕ 등) 헬퍼
"""

from __future__ import annotations

import calendar
import re
from enum import Enum
from typing import Any

from leave_usage_ledger import MONTHLY_SHEET_NAME, SHEET_NAME, get_leave_usage_ledger_path, normalize_period_label
from roster_workbook import LEAVE_LEDGER_ALIASES, MONTHLY_LEAVE_ALIASES, build_header_map

_USAGE_DAY_RE = re.compile(
    r"(\d{1,2})\s*월\s*(\d{1,2})\s*일\s*(?:사용|결근|무급|결근/무급)?",
    re.IGNORECASE,
)
_HALF_DAY_RE = re.compile(
    r"(\d{1,2})\s*월\s*반차\s*\(([\d.]+)\s*일\)\s*(?:사용|결근|무급|결근/무급)?",
    re.IGNORECASE,
)
_PERIOD_TAG_RE = re.compile(r"\[(\d{4}-\d{2}):[^\]]+\]")


class LeaveMarkKind(str, Enum):
    FULL = "full"
    HALF = "half"
    QUARTER = "quarter"
    ABSENCE = "absence"
    ABSENCE_HALF = "absence_half"


LEAVE_MARK_SYMBOLS: dict[LeaveMarkKind, str] = {
    LeaveMarkKind.FULL: "●",
    LeaveMarkKind.HALF: "◐",
    LeaveMarkKind.QUARTER: "◑",
    LeaveMarkKind.ABSENCE: "✕",
    LeaveMarkKind.ABSENCE_HALF: "◔",
}

EMPTY_DAY_SYMBOL = "·"

LEAVE_MARK_LEGEND: tuple[tuple[str, str], ...] = (
    (LEAVE_MARK_SYMBOLS[LeaveMarkKind.FULL], "연차(종일)"),
    (LEAVE_MARK_SYMBOLS[LeaveMarkKind.HALF], "반차"),
    (LEAVE_MARK_SYMBOLS[LeaveMarkKind.QUARTER], "1/4차"),
    (LEAVE_MARK_SYMBOLS[LeaveMarkKind.ABSENCE], "결근/무급(종일)"),
    (LEAVE_MARK_SYMBOLS[LeaveMarkKind.ABSENCE_HALF], "결근/무급(반차)"),
    (EMPTY_DAY_SYMBOL, "미사용"),
)


def _period_year_month(period: str) -> tuple[int, int]:
    norm = normalize_period_label(period) or period
    try:
        y, m = norm.split("-")
        return int(y), int(m)
    except (ValueError, AttributeError):
        return 2000, 1


def days_in_month(period: str) -> int:
    year, month = _period_year_month(period)
    return calendar.monthrange(year, month)[1]


def _kind_from_fraction(days: float, *, absence: bool) -> LeaveMarkKind:
    if days <= 0:
        raise ValueError("days must be positive")
    if abs(days - 0.25) < 1e-9:
        return LeaveMarkKind.ABSENCE_HALF if absence else LeaveMarkKind.QUARTER
    if abs(days - 0.5) < 1e-9:
        return LeaveMarkKind.ABSENCE_HALF if absence else LeaveMarkKind.HALF
    return LeaveMarkKind.ABSENCE if absence else LeaveMarkKind.FULL


def _merge_mark(existing: LeaveMarkKind | None, new: LeaveMarkKind) -> LeaveMarkKind:
    if existing is None:
        return new
    priority = {
        LeaveMarkKind.FULL: 4,
        LeaveMarkKind.ABSENCE: 4,
        LeaveMarkKind.HALF: 3,
        LeaveMarkKind.ABSENCE_HALF: 3,
        LeaveMarkKind.QUARTER: 2,
    }
    if priority.get(new, 0) >= priority.get(existing, 0):
        return new
    return existing


def _strip_period_tags(text: str) -> str:
    cleaned = _PERIOD_TAG_RE.sub("", text or "")
    return cleaned.replace("|", ",").strip()


def parse_usage_memo(
    memo: str,
    *,
    default_month: int | None = None,
    absence: bool = False,
) -> dict[int, LeaveMarkKind]:
    """'MM월 DD일 사용' / 'MM월 반차(0.5일) 사용' 등에서 일별 표식 추출."""
    marks: dict[int, LeaveMarkKind] = {}
    text = _strip_period_tags(memo)
    if not text:
        return marks

    for m in _USAGE_DAY_RE.finditer(text):
        month = int(m.group(1))
        day = int(m.group(2))
        if default_month is not None and month != default_month:
            continue
        if 1 <= day <= 31:
            kind = LeaveMarkKind.ABSENCE if absence or "결근" in m.group(0) or "무급" in m.group(0) else LeaveMarkKind.FULL
            marks[day] = _merge_mark(marks.get(day), kind)

    for m in _HALF_DAY_RE.finditer(text):
        month = int(m.group(1))
        if default_month is not None and month != default_month:
            continue
        days_val = float(m.group(2))
        kind = _kind_from_fraction(days_val, absence=absence or "결근" in m.group(0) or "무급" in m.group(0))
        # 반차에 일자가 없으면 첫 빈 칸(1일)에 표시 — synthetic 메모 대비
        day = 1
        while day in marks and day <= 31:
            day += 1
        if day <= 31:
            marks[day] = _merge_mark(marks.get(day), kind)

    return marks


def build_day_marks(
    period: str,
    *,
    leave_memo: str = "",
    absence_memo: str = "",
    leave_days: float = 0.0,
    absence_days: float = 0.0,
    leave_dates: list[int] | None = None,
    absence_dates: list[int] | None = None,
) -> dict[int, LeaveMarkKind]:
    """한 달(period=YYYY-MM)의 일별 연차·결근 표식."""
    _, month = _period_year_month(period)
    marks: dict[int, LeaveMarkKind] = {}

    if leave_dates:
        for day in leave_dates:
            if 1 <= int(day) <= 31:
                d = int(day)
                marks[d] = _merge_mark(marks.get(d), LeaveMarkKind.FULL)

    if absence_dates:
        for day in absence_dates:
            if 1 <= int(day) <= 31:
                d = int(day)
                marks[d] = _merge_mark(marks.get(d), LeaveMarkKind.ABSENCE)

    for day, kind in parse_usage_memo(leave_memo, default_month=month, absence=False).items():
        marks[day] = _merge_mark(marks.get(day), kind)

    for day, kind in parse_usage_memo(absence_memo, default_month=month, absence=True).items():
        marks[day] = _merge_mark(marks.get(day), kind)

    # 메모에 일자가 없고 일수만 있는 경우 — 1일부터 순서대로 채움(기존 synthetic 메모 호환)
    if leave_days > 0 and not _leave_mark_days(marks):
        _fill_sequential(marks, leave_days, absence=False)

    if absence_days > 0 and not any(
        marks.get(d) in (LeaveMarkKind.ABSENCE, LeaveMarkKind.ABSENCE_HALF) for d in marks
    ):
        _fill_sequential(marks, absence_days, absence=True)

    return marks


def _leave_mark_days(marks: dict[int, LeaveMarkKind]) -> list[int]:
    return [
        d
        for d, k in marks.items()
        if k in (LeaveMarkKind.FULL, LeaveMarkKind.HALF, LeaveMarkKind.QUARTER)
    ]


def _fill_sequential(marks: dict[int, LeaveMarkKind], total_days: float, *, absence: bool) -> None:
    remaining = float(total_days)
    day = 1
    max_day = 31
    while remaining > 1e-9 and day <= max_day:
        if day in marks:
            day += 1
            continue
        if remaining >= 1.0 - 1e-9:
            kind = LeaveMarkKind.ABSENCE if absence else LeaveMarkKind.FULL
            marks[day] = kind
            remaining -= 1.0
        else:
            kind = _kind_from_fraction(remaining, absence=absence)
            marks[day] = kind
            remaining = 0.0
        day += 1


def symbol_for_day(marks: dict[int, LeaveMarkKind], day: int) -> str:
    kind = marks.get(day)
    if kind is None:
        return EMPTY_DAY_SYMBOL
    return LEAVE_MARK_SYMBOLS.get(kind, EMPTY_DAY_SYMBOL)


def render_month_mark_row(period: str, marks: dict[int, LeaveMarkKind]) -> str:
    """일별 표식 한 줄 (예: ··●●◐···)."""
    dim = days_in_month(period)
    return "".join(symbol_for_day(marks, d) for d in range(1, dim + 1))


def render_month_day_header(period: str) -> str:
    """일 번호 헤더 (예: 01 02 03 ...)."""
    dim = days_in_month(period)
    return " ".join(f"{d:02d}" for d in range(1, dim + 1))


def render_month_calendar_block(period: str, marks: dict[int, LeaveMarkKind]) -> str:
    """Treeview·라벨용 2줄 블록."""
    year, month = _period_year_month(period)
    header = render_month_day_header(period)
    body = render_month_mark_row(period, marks)
    return f"{month}월 ({year})\n{header}\n{body}"


def render_compact_calendar(period: str, marks: dict[int, LeaveMarkKind]) -> str:
    """한 줄 요약 — Treeview 열용."""
    _, month = _period_year_month(period)
    return f"{month:02d}월 " + render_month_mark_row(period, marks)


def format_legend_text() -> str:
    return "  ".join(f"{sym} {label}" for sym, label in LEAVE_MARK_LEGEND)


def format_usage_dates_summary(period: str, marks: dict[int, LeaveMarkKind]) -> str:
    """사용 일자 요약 — 예: '5/1(●), 5/12(◐)'."""
    _, month = _period_year_month(period)
    parts: list[str] = []
    for day in sorted(marks.keys()):
        kind = marks[day]
        sym = LEAVE_MARK_SYMBOLS.get(kind, "")
        if not sym:
            continue
        parts.append(f"{month}/{day}({sym})")
    return ", ".join(parts)


def _merge_memo_text(existing: str, new: str) -> str:
    existing = (existing or "").strip()
    new = (new or "").strip()
    if not new:
        return existing
    if not existing:
        return new
    if new in existing:
        return existing
    return f"{existing}, {new}"


def _merge_date_lists(existing: list[Any] | None, new: list[Any] | None) -> list[int]:
    merged = {int(d) for d in (existing or []) if d is not None}
    merged.update(int(d) for d in (new or []) if d is not None)
    return sorted(merged)


def aggregate_leave_usage_rows(
    rows: list[dict[str, Any]],
    period: str = "",
) -> list[dict[str, Any]]:
    """동일 직원의 연차·결근 행을 1인 1행으로 합칩니다."""
    from roster_constants import norm_name_key

    by_key: dict[str, dict[str, Any]] = {}
    for r in rows:
        name = str(r.get("name") or r.get("성명") or "").strip()
        if not name:
            continue
        key = norm_name_key(name)
        if not key:
            continue

        if key not in by_key:
            entry = dict(r)
            entry["name"] = name
            by_key[key] = entry
            continue

        existing = by_key[key]
        for field in ("month_leave", "absence_days"):
            existing[field] = float(existing.get(field) or 0) + float(r.get(field) or 0)
        existing["absence_count"] = int(existing.get("absence_count") or 0) + int(
            r.get("absence_count") or 0
        )

        for field in ("accrued", "used_total"):
            nv = r.get(field)
            if nv not in (None, "", 0, 0.0):
                existing[field] = nv

        remaining = r.get("remaining")
        if remaining not in (None, "", 0, 0.0):
            existing["remaining"] = remaining

        for memo_field in ("leave_memo", "absence_memo"):
            existing[memo_field] = _merge_memo_text(
                str(existing.get(memo_field) or ""),
                str(r.get(memo_field) or r.get("연차내역" if memo_field == "leave_memo" else "무급내역") or ""),
            )

        for date_field in (
            "leave_sheet_leave_dates",
            "leave_sheet_absence_dates",
            "leave_dates",
            "absence_dates",
        ):
            if r.get(date_field):
                existing[date_field] = _merge_date_lists(existing.get(date_field), r.get(date_field))

        if not existing.get("workplace") and r.get("workplace"):
            existing["workplace"] = r.get("workplace")
        if not existing.get("affiliate") and r.get("affiliate"):
            existing["affiliate"] = r.get("affiliate")
        existing["emp_no"] = existing.get("emp_no") or r.get("emp_no") or r.get("사번")

        if period:
            marks = build_marks_from_row(existing, period)
            existing["dates_summary"] = format_usage_dates_summary(period, marks)

    out = list(by_key.values())
    if period:
        for entry in out:
            if not entry.get("dates_summary"):
                marks = build_marks_from_row(entry, period)
                entry["dates_summary"] = format_usage_dates_summary(period, marks)
    return sorted(out, key=lambda x: (str(x.get("workplace") or ""), str(x.get("name") or "")))


def build_marks_from_row(row: dict[str, Any], period: str) -> dict[int, LeaveMarkKind]:
    """급여/대장 병합 행 → 일별 표식."""
    return build_day_marks(
        period,
        leave_memo=str(row.get("leave_memo") or row.get("연차내역") or ""),
        absence_memo=str(row.get("absence_memo") or row.get("무급내역") or ""),
        leave_days=float(row.get("month_leave") or row.get("leave_days") or row.get("당월연차") or 0),
        absence_days=float(row.get("absence_days") or row.get("unpaid_days") or row.get("무급일수") or 0),
        leave_dates=row.get("leave_sheet_leave_dates") or row.get("leave_dates"),
        absence_dates=row.get("leave_sheet_absence_dates") or row.get("absence_dates"),
    )


def load_ledger_monthly_rows(period: str) -> list[dict[str, Any]]:
    """연차사용대장 「월별현황」 시트 로드."""
    from services.monthly_leave_manager import load_monthly_leave_rows

    return load_monthly_leave_rows(period)


def load_ledger_years() -> list[str]:
    """대장에 기록된 연도 목록 (최신순)."""
    path = get_leave_usage_ledger_path()
    if not path.is_file():
        return []
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    years: set[str] = set()
    for sheet_name in (MONTHLY_SHEET_NAME, SHEET_NAME):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        aliases = MONTHLY_LEAVE_ALIASES if sheet_name == MONTHLY_SHEET_NAME else LEAVE_LEDGER_ALIASES
        headers = build_header_map(ws, aliases)
        period_col = headers.get("처리월") or headers.get("사용월")
        if not period_col:
            continue
        for r in range(2, (ws.max_row or 1) + 1):
            pv = normalize_period_label(ws.cell(r, period_col).value)
            if pv:
                years.add(pv[:4])
    wb.close()
    return sorted(years, reverse=True)
