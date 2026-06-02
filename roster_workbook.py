"""
roster_workbook.py - templates/근로자명부.xlsx (씨엔엘 + 연차대장) 읽기/쓰기

- 1번 시트(씨엔엘): 직원·시급·보험. 발생/사용/잔여연차는 수식(#N/A 가능) → 값만 읽고 0 처리.
- 2번 시트(연차대장): 읽기·최초 이관용 (신규 기록은 연차사용대장/ 폴더)
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from roster_constants import ROSTER_HEADER_ALIASES, build_header_map, norm_name_key
from roster_leave_sheet import (
    latest_period_hint_from_snapshot,
    merge_leave_into_roster_record,
    parse_leave_annual_sheet,
)
from utils import safe_number

_PERIOD_TAG_RE = re.compile(r"\[(\d{4}-\d{2})")

# 연차대장 시트 헤더 별칭
LEAVE_LEDGER_ALIASES: dict[str, tuple[str, ...]] = {
    "성명": ("성명", "이름", "사원명"),
    "사번": ("사번", "사원번호"),
    "사용월": ("사용월", "급여월", "해당월", "월"),
    "사용일수": ("사용일수", "일수", "연차일수", "사용", "연차"),
    "사용내역": ("사용내역", "내역", "비고", "메모", "연차사용메모"),
    "사용일": ("사용일", "일자", "날짜", "연차사용일"),
    "처리월": ("처리월", "급여처리월", "정산월"),
    "구분": ("구분", "유형", "기록구분", "근태구분"),
    "발생연차": ("발생연차", "발생 연차"),
    "누적사용연차": ("누적사용연차", "사용연차", "누적 사용"),
    "잔여연차": ("잔여연차", "잔여 연차"),
    "무급횟수": ("무급횟수", "결근횟수", "횟수"),
}

MONTHLY_LEAVE_ALIASES: dict[str, tuple[str, ...]] = {
    "성명": ("성명", "이름", "사원명"),
    "사번": ("사번", "사원번호"),
    "처리월": ("처리월", "급여처리월", "정산월"),
    "발생연차": ("발생연차", "발생 연차"),
    "당월연차": ("당월연차", "당월 사용", "당월연차사용"),
    "누적사용연차": ("누적사용연차", "사용연차", "누적 사용"),
    "잔여연차": ("잔여연차", "잔여 연차"),
    "무급일수": ("무급일수", "결근일수", "무급/결근일수"),
    "무급횟수": ("무급횟수", "결근횟수", "횟수"),
    "연차내역": ("연차내역", "연차 사용내역"),
    "무급내역": ("무급내역", "결근내역", "무급/결근내역"),
    "최종갱신": ("최종갱신", "갱신일시"),
}


def is_leave_deficit_display(value: Any) -> bool:
    """잔여연차가 '-' 또는 음수 → 이미 발생 연차를 초과 사용한 상태."""
    if value is None:
        return False
    if isinstance(value, str) and value.strip() in ("-", "－", "—"):
        return True
    if isinstance(value, (int, float)):
        try:
            return float(value) < 0
        except (TypeError, ValueError):
            return False
    return False


def read_leave_balance(value: Any) -> float | None:
    """
    연차 잔여·발생 등 숫자 읽기.

    '-' 표기는 None(초과 사용), #N/A 는 0, 음수는 그대로 반환.
    """
    if value is None:
        return None
    if is_leave_deficit_display(value) and isinstance(value, str):
        return None
    if isinstance(value, str):
        t = value.strip().upper()
        if t.startswith("#") or t in ("#N/A", "#NA", "#REF!", "#VALUE!", "#NAME?", "#DIV/0!"):
            return 0.0
    if isinstance(value, (int, float)):
        import math

        if math.isnan(value) or math.isinf(value):
            return 0.0
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def sanitize_roster_number(value: Any) -> float:
    """#N/A·#REF! 등 엑셀 오류를 0으로. '-' 단독은 0 (연차 잔여는 read_leave_balance 사용)."""
    if value is None:
        return 0.0
    if isinstance(value, str):
        t = value.strip().upper()
        if not t or t in ("—", "NONE", "NAN"):
            return 0.0
        if t in ("-", "－") and len(t.strip()) <= 1:
            return 0.0
        if t.startswith("#") or t in ("#N/A", "#NA", "#REF!", "#VALUE!", "#NAME?", "#DIV/0!"):
            return 0.0
    n = safe_number(value, 0.0)
    return max(0.0, n)


def find_main_roster_sheet(wb: openpyxl.Workbook) -> Worksheet:
    for name in wb.sheetnames:
        if "씨엔엘" in name.replace(" ", ""):
            return wb[name]
    for name in wb.sheetnames:
        if "연차" not in name:
            ws = wb[name]
            h = build_header_map(ws, ROSTER_HEADER_ALIASES)
            if "성명" in h:
                return ws
    return wb.active


def find_leave_ledger_sheet(wb: openpyxl.Workbook) -> Worksheet | None:
    for name in wb.sheetnames:
        if "연차" in name and ("대장" in name or "관리" in name):
            return wb[name]
    if len(wb.sheetnames) >= 2:
        second = wb[wb.sheetnames[1]]
        if "연차" in wb.sheetnames[1] or build_header_map(second, LEAVE_LEDGER_ALIASES).get("성명"):
            return second
    return None


_NUMERIC_ROSTER_FIELDS = frozenset(
    {
        "통상시급",
        "국민연금",
        "건강보험",
        "소득세",
        "기본시급",
        "수당",
        "발생연차",
        "사용연차",
        "잔여연차",
        "예상발생연차",
    }
)


def _cell_has_formula(ws: Worksheet, row: int, col: int) -> bool:
    val = ws.cell(row, col).value
    return isinstance(val, str) and val.startswith("=")


def load_employee_roster_from_workbook(
    roster_path: Path,
    *,
    detect_formulas: bool = False,
    period_hint: str = "",
) -> dict[str, dict[str, Any]]:
    """
    씨엔엘 시트에서 직원 명부를 읽습니다.

    detect_formulas=False(기본): 엑셀 1회 read_only 로드 — UI·캐시용.
    detect_formulas=True: 값·수식 2회 로드 — 구버전 호환(느림).
    """
    if not roster_path.exists():
        return {}

    if detect_formulas:
        return _load_roster_dual_workbook(roster_path, period_hint=period_hint)

    return _load_roster_fast(roster_path, period_hint=period_hint)


def _load_leave_snapshots(roster_path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(roster_path, data_only=True)
    try:
        return parse_leave_annual_sheet(wb)
    finally:
        wb.close()


def _load_roster_fast(roster_path: Path, *, period_hint: str = "") -> dict[str, dict[str, Any]]:
    leave_snaps = _load_leave_snapshots(roster_path)

    wb = openpyxl.load_workbook(roster_path, read_only=True, data_only=True)
    try:
        ws = find_main_roster_sheet(wb)
        headers = build_header_map(ws, ROSTER_HEADER_ALIASES)
        if "성명" not in headers:
            return {}

        wanted = [k for k in ROSTER_HEADER_ALIASES if k != "성명" and k in headers]
        max_col = max(headers.values())
        name_col = headers["성명"] - 1
        col_by_field = {k: headers[k] - 1 for k in wanted}

        roster: dict[str, dict[str, Any]] = {}
        empty_streak = 0
        sheet_title = ws.title

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, max_col=max_col, values_only=True),
            start=2,
        ):
            name = row[name_col] if name_col < len(row) else None
            if name is None or str(name).strip() == "":
                empty_streak += 1
                if empty_streak >= 25:
                    break
                continue
            empty_streak = 0

            rec: dict[str, Any] = {
                "성명": str(name).strip(),
                "_row": row_idx,
                "_main_sheet": sheet_title,
            }

            for k in wanted:
                idx = col_by_field[k]
                v = row[idx] if idx < len(row) else None
                if k == "잔여연차":
                    rec["잔여연차_raw"] = v
                    rec["_잔여연차_초과"] = is_leave_deficit_display(v)
                    bal = read_leave_balance(v)
                    rec[k] = bal if bal is not None else 0.0
                elif k in _NUMERIC_ROSTER_FIELDS:
                    rec[k] = sanitize_roster_number(v)
                else:
                    rec[k] = v

            emp_no = row[headers["사번"] - 1] if "사번" in headers and headers["사번"] - 1 < len(row) else None

            snap = leave_snaps.get(norm_name_key(str(name).strip()))
            hint = period_hint or (
                latest_period_hint_from_snapshot(snap) if snap is not None else ""
            )
            if hint or snap is not None:
                merge_leave_into_roster_record(
                    rec,
                    snap,
                    hint or period_hint or "2099-12",
                )
            elif period_hint:
                merge_leave_into_roster_record(rec, None, period_hint)

            n = norm_name_key(str(name).strip())
            e = norm_name_key(emp_no)
            if n:
                roster[n] = rec
            if e:
                roster[e] = rec

        return roster
    finally:
        wb.close()


def _load_roster_dual_workbook(
    roster_path: Path,
    *,
    period_hint: str = "",
) -> dict[str, dict[str, Any]]:
    leave_snaps = _load_leave_snapshots(roster_path)
    wb_values = openpyxl.load_workbook(roster_path, data_only=True)
    wb_formula = openpyxl.load_workbook(roster_path, data_only=False)
    try:
        ws_val = find_main_roster_sheet(wb_values)
        ws_formula = find_main_roster_sheet(wb_formula)
        headers = build_header_map(ws_val, ROSTER_HEADER_ALIASES)
        if "성명" not in headers:
            return {}

        leave_cols = {k: headers[k] for k in ("발생연차", "사용연차", "잔여연차") if k in headers}
        wanted = list(ROSTER_HEADER_ALIASES.keys())
        roster: dict[str, dict[str, Any]] = {}
        last_row = min(ws_val.max_row, 5000)
        for r in range(2, last_row + 1):
            name = ws_val.cell(r, headers["성명"]).value
            if name is None or str(name).strip() == "":
                continue
            emp_no = ws_val.cell(r, headers["사번"]).value if "사번" in headers else None
            rec: dict[str, Any] = {
                "성명": str(name).strip(),
                "_row": r,
                "_main_sheet": ws_val.title,
            }
            for k in wanted:
                if k == "성명" or k not in headers:
                    continue
                v = ws_val.cell(r, headers[k]).value
                if k == "잔여연차":
                    rec["잔여연차_raw"] = v
                    rec["_잔여연차_초과"] = is_leave_deficit_display(v)
                    bal = read_leave_balance(v)
                    rec[k] = bal if bal is not None else 0.0
                elif k in _NUMERIC_ROSTER_FIELDS:
                    rec[k] = sanitize_roster_number(v)
                else:
                    rec[k] = v
            for col_name, col_idx in leave_cols.items():
                fcell = ws_formula.cell(r, col_idx)
                rec[f"_{col_name}_수식"] = isinstance(fcell.value, str) and str(fcell.value).startswith("=")
            snap = leave_snaps.get(norm_name_key(str(name).strip()))
            hint = period_hint or (
                latest_period_hint_from_snapshot(snap) if snap is not None else ""
            )
            if hint or snap is not None:
                merge_leave_into_roster_record(
                    rec,
                    snap,
                    hint or period_hint or "2099-12",
                )
            elif period_hint:
                merge_leave_into_roster_record(rec, None, period_hint)
            n = norm_name_key(name)
            e = norm_name_key(emp_no)
            if n:
                roster[n] = rec
            if e:
                roster[e] = rec
        return roster
    finally:
        wb_values.close()
        wb_formula.close()


@dataclass
class LeaveLedgerEntry:
    """연차사용대장에 추가할 1건 (연차 또는 결근/무급)."""

    main_row: int
    name: str
    emp_no: Any
    period_label: str
    usage_month: str
    days: float
    usage_memo: str
    record_kind: str = "연차"  # "연차" | "결근"
    accrued: float | None = None
    used_total: float | None = None
    remaining: float | None = None
    occurrence_count: int | None = None


@dataclass
class MonthlyLeaveSummary:
    """월별 인원별 연차·무급/결근 현황 (1인 1행)."""

    name: str
    emp_no: Any
    period_label: str
    accrued: float
    month_leave_used: float
    used_total: float
    remaining: float
    absence_days: float
    absence_count: int
    leave_memo: str = ""
    absence_memo: str = ""


def _ensure_leave_ledger_headers(ws: Worksheet, headers: dict[str, int]) -> dict[str, int]:
    defaults = [
        ("성명", "성명"),
        ("사번", "사번"),
        ("처리월", "처리월"),
        ("사용월", "사용월"),
        ("구분", "구분"),
        ("사용일수", "사용일수"),
        ("사용내역", "사용내역"),
    ]
    for canonical, label in defaults:
        if canonical not in headers:
            col = ws.max_column + 1
            ws.cell(1, col, label)
            headers[canonical] = col
    return headers


def _row_record_kind(
    ws: Worksheet,
    row: int,
    headers: dict[str, int],
) -> str:
    """행의 기록 구분(연차 / 결근)."""
    kind_col = headers.get("구분")
    if kind_col:
        v = ws.cell(row, kind_col).value
        if v is not None and str(v).strip():
            return str(v).strip()
    memo_col = headers.get("사용내역")
    if memo_col:
        memo = str(ws.cell(row, memo_col).value or "")
        if ":결근:" in memo or "결근/무급" in memo:
            return "결근"
    return "연차"


def _find_ledger_rows_for_period(
    ws: Worksheet,
    headers: dict[str, int],
    name: str,
    period_label: str,
    record_kind: str | None = None,
) -> list[int]:
    """같은 급여월·성명(·구분)의 기존 연차사용대장 행(재처리 시 삭제)."""
    name_key = norm_name_key(name)
    name_col = headers.get("성명", 0)
    period_col = headers.get("처리월") or headers.get("사용월")
    memo_col = headers.get("사용내역")
    rows: list[int] = []

    if not name_col:
        return rows

    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, name_col).value
        if norm_name_key(nm) != name_key:
            continue
        period_match = False
        if period_col:
            pv = ws.cell(r, period_col).value
            if pv is not None and str(pv).strip() in (
                period_label,
                period_label.replace("-", "."),
            ):
                period_match = True
        if not period_match and memo_col:
            memo = str(ws.cell(r, memo_col).value or "")
            if f"[{period_label}" in memo or f"[{period_label}:" in memo:
                period_match = True
        if not period_match:
            continue
        if record_kind is not None and _row_record_kind(ws, r, headers) != record_kind:
            continue
        rows.append(r)
    return rows


def roster_dict_to_list(roster: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    """명부 dict(이름·사번 키 중복) → 시트 행 순 목록."""
    by_row: dict[int, dict[str, Any]] = {}
    for rec in roster.values():
        if not isinstance(rec, dict):
            continue
        row = rec.get("_row")
        if not isinstance(row, int) or row < 2:
            continue
        if row not in by_row:
            by_row[row] = rec
    return [by_row[r] for r in sorted(by_row)]


_FORMULA_PROTECTED = frozenset({"발생연차", "사용연차", "잔여연차"})


def _coerce_cell_value(field: str, value: Any) -> Any:
    if value is None or value == "":
        return None
    if field in (
        "기본시급",
        "통상시급",
        "국민연금",
        "건강보험",
        "소득세",
        "수당",
        "예상발생연차",
        "발생연차",
        "사용연차",
        "잔여연차",
    ):
        if field == "잔여연차" and isinstance(value, str) and value.strip() in ("-", "－"):
            return value.strip()
        return sanitize_roster_number(value)
    if isinstance(value, float) and value == int(value):
        return int(value)
    return value


def save_employee_roster_records(
    roster_path: Path,
    records: list[dict[str, Any]],
) -> int:
    """씨엔엘 시트에 명부를 저장합니다. 수식 연차 열은 수식이 있으면 덮어쓰지 않습니다."""
    roster_path.parent.mkdir(parents=True, exist_ok=True)
    if not roster_path.exists():
        _create_minimal_roster(roster_path)

    wb = openpyxl.load_workbook(roster_path)
    ws = find_main_roster_sheet(wb)
    headers = build_header_map(ws, ROSTER_HEADER_ALIASES)

    for canonical, aliases in ROSTER_HEADER_ALIASES.items():
        if canonical in headers:
            continue
        col = ws.max_column + 1
        ws.cell(1, col, aliases[0])
        headers[canonical] = col

    saved = 0
    next_row = max(ws.max_row + 1, 2)

    for rec in records:
        name = str(rec.get("성명") or "").strip()
        if not name:
            continue

        row = rec.get("_row")
        if isinstance(row, int) and row >= 2:
            r = row
        else:
            r = next_row
            next_row += 1
            rec["_row"] = r

        for field, col_idx in headers.items():
            if field == "성명":
                ws.cell(r, col_idx, name)
                continue
            if field in _FORMULA_PROTECTED and _cell_has_formula(ws, r, col_idx):
                continue
            if field not in rec:
                continue
            val = _coerce_cell_value(field, rec[field])
            if val is None:
                ws.cell(r, col_idx, None)
            else:
                ws.cell(r, col_idx, val)

        saved += 1

    from core.file_save import save_workbook

    save_workbook(wb, roster_path)
    wb.close()
    return saved


def _create_minimal_roster(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "씨엔엘"
    for col, (_canonical, aliases) in enumerate(ROSTER_HEADER_ALIASES.items(), start=1):
        ws.cell(1, col, aliases[0])
    from core.file_save import save_workbook

    save_workbook(wb, path)
    wb.close()


def save_leave_ledger_entries(
    roster_path: Path,
    entries: list[LeaveLedgerEntry],
    payroll_period: str | None = None,
) -> int:
    """연차사용대장/ 폴더에 기록 (명부 파일은 수정하지 않음)."""
    from leave_usage_ledger import save_leave_usage_ledger_entries

    if not entries:
        return 0
    period = payroll_period or datetime.now().strftime("%Y-%m")
    info = save_leave_usage_ledger_entries(
        entries,
        period,
        roster_path_for_migration=roster_path if roster_path.exists() else None,
    )
    return int(info.get("written") or 0)
