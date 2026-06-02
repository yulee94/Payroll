"""
payroll_builder.py - 청구서 + 마스터 → 출력용 급여 레코드 생성
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from bank_account import resolve_payment_from_roster
from annual_leave_manager import (
    annual_leave_roster_warnings,
    apply_leave_split_to_invoice,
    apply_leave_usage_decision_to_invoice,
    apply_monthly_annual_leave_to_invoice,
    calc_absence_occurrence_count,
    calc_base_deduction_amount_with_leave,
    format_absence_usage_memo,
    format_usage_month_label,
    roster_has_any_leave_column,
)
from roster_constants import (
    ROSTER_HEADER_ALIASES,
    build_header_map,
    find_fuzzy_name_key,
    norm_name_key,
)
from roster_workbook import load_employee_roster_from_workbook
from core.org_config import get_default_affiliate
from core.payroll_calc_rules import resolve_social_insurance
from tax import calculate_tax
from utils import format_hours_or_days, round_won, round_won_tens, safe_number

from core.paths import dev_root, employees_dir, templates_dir

BASE_DIR = dev_root()
TEMPLATES_DIR = templates_dir()
EMPLOYEES_DIR = employees_dir()
ROSTER_FILENAME = "근로자명부.xlsx"

def get_templates_roster_path() -> Path | None:
    """templates/근로자명부.xlsx 경로 (없으면 근로자명부*.xlsx 탐색)."""
    exact = TEMPLATES_DIR / ROSTER_FILENAME
    if exact.exists():
        return exact
    matches = sorted(TEMPLATES_DIR.glob("근로자명부*.xlsx"))
    return matches[0] if matches else None


def load_payment_master(template_path: Path) -> dict[str, dict[str, Any]]:
    """
    지급내역양식.xlsx 에서 이름→계좌·은행 정보를 읽습니다.
    (양식에 등록된 직원 기준)
    """
    master: dict[str, dict[str, Any]] = {}
    if not template_path.exists():
        return master

    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb.active
    for row in range(5, ws.max_row + 1):
        name = ws.cell(row, 2).value
        if not name:
            continue
        key = str(name).strip().replace(" ", "")
        master[key] = {
            "name": str(name).strip(),
            "workplace": ws.cell(row, 3).value or "한국앰코생산",
            "bank_code": ws.cell(row, 4).value or "",
            "bank_name": ws.cell(row, 5).value or "",
            "account": ws.cell(row, 6).value or "",
            "holder": ws.cell(row, 10).value or name,
        }
    wb.close()
    return master


def load_payslip_master(template_path: Path) -> dict[str, dict[str, Any]]:
    """급여명세서양식.xlsx 에서 전화번호·생년월일 등을 읽습니다."""
    master: dict[str, dict[str, Any]] = {}
    if not template_path.exists():
        return master

    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb.active
    for row in range(3, ws.max_row + 1):
        name = ws.cell(row, 1).value
        if not name:
            continue
        key = str(name).strip().replace(" ", "")
        master[key] = {
            "phone": ws.cell(row, 2).value or "",
            "birth": ws.cell(row, 3).value or "",
            "pay_day": ws.cell(row, 4).value or "25일",
        }
    wb.close()
    return master


def _norm_name_key(name: Any) -> str:
    return norm_name_key(name)


def _build_roster_header_map(ws) -> dict[str, int]:
    return build_header_map(ws, ROSTER_HEADER_ALIASES)


def _score_roster_file(roster_path: Path) -> int:
    """명부 파일 품질 점수(인원·기본시급 보유 여부) — 자동 선택용."""
    roster = load_employee_roster(roster_path)
    if not roster:
        return 0
    names: set[str] = set()
    with_base = 0
    for rec in roster.values():
        if not isinstance(rec, dict):
            continue
        nm = rec.get("성명")
        if nm:
            names.add(_norm_name_key(nm))
        if safe_number(rec.get("기본시급"), 0.0) > 0:
            with_base += 1
    return len(names) * 10 + with_base


def resolve_roster_path(
    invoice_path: Path,
    explicit: Path | None = None,
    *,
    period: str = "",
) -> tuple[Path | None, str]:
    """
    근로자 명부 경로 — 급여월별 스냅샷(2026-01~05) 또는 templates/근로자명부.xlsx.
    """
    _ = invoice_path, explicit
    try:
        from services.employee_roster_store import (
            canonical_roster_path,
            roster_path_for_period,
            roster_updated_display,
        )

        path = roster_path_for_period(period)
        if path.is_file():
            if period and path != canonical_roster_path():
                return path, f"templates/{path.name} ({period} 기준)"
            return path, f"templates/근로자명부 (갱신 {roster_updated_display()})"
    except ImportError:
        pass
    roster = get_templates_roster_path()
    if roster is not None:
        return roster, "templates/근로자명부"
    return None, ""


def load_employee_roster(
    roster_path: Path,
    *,
    period_hint: str = "",
) -> dict[str, dict[str, Any]]:
    """
    templates/근로자명부.xlsx 의 씨엔엘 시트에서 직원 명부를 읽습니다.

    발생/사용/잔여연차는 수식 결과(#N/A 포함)를 숫자로 읽으며 오류는 0입니다.
    """
    try:
        from services.employee_roster_store import (
            _period_hint_from_path,
            get_cached_roster_dict,
        )

        hint = period_hint or _period_hint_from_path(roster_path)
        cached = get_cached_roster_dict(roster_path, period_hint=hint)
        if cached is not None:
            return cached
    except ImportError:
        pass
    hint = period_hint
    try:
        from services.employee_roster_store import _period_hint_from_path

        hint = hint or _period_hint_from_path(roster_path)
    except ImportError:
        pass
    return load_employee_roster_from_workbook(
        roster_path, detect_formulas=False, period_hint=hint
    )


def _roster_lookup_keys(master: dict[str, dict[str, Any]]) -> list[str]:
    keys: list[str] = []
    seen: set[str] = set()
    for dict_key, rec in master.items():
        if not isinstance(rec, dict):
            continue
        for candidate in (dict_key, rec.get("성명"), rec.get("이름"), rec.get("name")):
            nk = _norm_name_key(candidate)
            if nk and nk not in seen:
                seen.add(nk)
                keys.append(nk)
    return keys


def _match_master(name: str, master: dict[str, dict[str, Any]]) -> dict[str, Any]:
    key = _norm_name_key(name)
    if key in master:
        return master[key]
    for v in master.values():
        if not isinstance(v, dict):
            continue
        for field in ("성명", "이름", "name"):
            nm = v.get(field)
            if nm and _norm_name_key(nm) == key:
                return v
    fuzzy = find_fuzzy_name_key(key, _roster_lookup_keys(master))
    if fuzzy:
        if fuzzy in master:
            return master[fuzzy]
        for v in master.values():
            if not isinstance(v, dict):
                continue
            for field in ("성명", "이름", "name"):
                nm = v.get(field)
                if nm and _norm_name_key(nm) == fuzzy:
                    return v
    return {}


def roster_has_annual_leave_fields(emp_roster: dict[str, Any]) -> bool:
    """명부에 연차 관련 열이 있는지 확인합니다."""
    for key in ("잔여연차", "발생연차", "사용연차"):
        if key in emp_roster:
            return True
    return False


def format_ledger_leave_usage_display(inv: dict[str, Any]) -> str:
    """급여대장 6행(짝수) H열 — 연차사용내역 / 결근·무급 표기."""
    parts: list[str] = []

    leave_days = safe_number(inv.get("leave_days"), 0.0)
    if leave_days > 0:
        state = inv.get("_annual_leave_state")
        d = format_hours_or_days(leave_days)
        if state and getattr(state, "usage_memo_entry", ""):
            parts.append(f"{d}일 ({state.usage_memo_entry})")
        else:
            parts.append(f"{d}일")

    shutdown_days = safe_number(inv.get("shutdown_leave_days"), 0.0)
    if shutdown_days > 0:
        pct = safe_number(inv.get("_shutdown_pay_percent"), 0.0)
        sd = int(shutdown_days) if shutdown_days == int(shutdown_days) else shutdown_days
        if pct > 0:
            parts.append(f"휴업 {sd}일({pct:g}%)")
        else:
            parts.append(f"휴업 {sd}일")

    from shutdown_leave import pure_unpaid_days

    unpaid_only = pure_unpaid_days(inv)
    if unpaid_only > 0:
        u = int(unpaid_only) if unpaid_only == int(unpaid_only) else unpaid_only
        parts.append(f"결근/무급 {u}일")
    elif not parts:
        unpaid = safe_number(inv.get("unpaid_days"), 0.0)
        dept = str(inv.get("dept") or "")
        if unpaid > 0 and ("결근" in dept or "무급" in dept):
            u = int(unpaid) if unpaid == int(unpaid) else unpaid
            parts.append(f"결근/무급 {u}일")

    return " / ".join(parts)


def ledger_base_deduction_for_excel(r: dict[str, Any]) -> int:
    """급여대장 기본공제 — 양식 합계(급여총액)에 맞게 음수로 기록."""
    v = safe_number(r.get("base_deduction"), 0.0)
    if v == 0:
        return 0
    if v > 0:
        return -round_won(v)
    return round_won(v)


def _clear_roster_controlled_allowances(inv: dict[str, Any]) -> None:
    """명부 매칭 시 청구서 수당 잔존값 제거(재계산 전 초기화)."""
    inv["position_pay"] = 0
    inv["ot_pay"] = 0
    inv["shift_pay"] = 0
    inv["night_pay"] = 0
    inv["special_pay"] = 0
    inv["special_ext_pay"] = 0
    inv["annual_pay"] = 0


def apply_roster_allowance_policy(
    inv: dict[str, Any],
    emp_roster: dict[str, Any],
) -> list[str]:
    """
    명부에 정의된 항목만 급여대장에 반영.

    - 고정수당: 명부 '수당' 열만 (청구서 직책수당 Z열 무시)
    - 연장·교대·심야·특근: 명부 '통상시급' 있을 때만, 청구서 시간×시급으로 산출
    """
    warnings: list[str] = []
    name = str(inv.get("name") or emp_roster.get("성명") or "")

    invoice_position = safe_number(inv.get("position_pay"), 0.0)
    invoice_ot = safe_number(inv.get("ot_pay"), 0.0)
    invoice_shift = safe_number(inv.get("shift_pay"), 0.0)
    invoice_night = safe_number(inv.get("night_pay"), 0.0)
    invoice_special = safe_number(inv.get("special_pay"), 0.0) + safe_number(
        inv.get("special_ext_pay"), 0.0
    )

    _clear_roster_controlled_allowances(inv)

    roster_fixed = safe_number(emp_roster.get("수당"), 0.0)
    inv["position_pay"] = round_won(roster_fixed)
    if invoice_position > 0 and inv["position_pay"] == 0:
        warnings.append(
            f"{name}: 청구서 직책/고정수당 {int(invoice_position):,}원 → "
            "명부 '수당' 없음(0)으로 미지급"
        )

    ordinary_hourly = safe_number(emp_roster.get("통상시급"), 0.0)

    if ordinary_hourly > 0:
        inv["ordinary_hourly"] = ordinary_hourly
        ot_h = safe_number(inv.get("ot_hours"), 0.0)
        shift_h = safe_number(inv.get("shift_hours"), 0.0)
        night_h = safe_number(inv.get("night_hours"), 0.0)
        special_h = safe_number(inv.get("special_hours"), 0.0)
        special_ext_h = safe_number(inv.get("special_ext_hours"), 0.0)

        inv["ot_pay"] = round_won(ordinary_hourly * ot_h * 1.5) if ot_h > 0 else 0
        inv["shift_pay"] = round_won(ordinary_hourly * shift_h * 1.5) if shift_h > 0 else 0
        inv["night_pay"] = round_won(ordinary_hourly * night_h * 0.5) if night_h > 0 else 0
        inv["special_pay"] = round_won(ordinary_hourly * special_h * 1.5) if special_h > 0 else 0
        inv["special_ext_pay"] = (
            round_won(ordinary_hourly * special_ext_h * 0.5) if special_ext_h > 0 else 0
        )

        if invoice_ot > 0 and inv["ot_pay"] == 0 and ot_h <= 0:
            warnings.append(
                f"{name}: 청구서 연장수당 {int(invoice_ot):,}원 → "
                "명부 통상시급·연장시간 없어 미지급"
            )
        if invoice_shift > 0 and inv["shift_pay"] == 0 and shift_h <= 0:
            warnings.append(
                f"{name}: 청구서 교대수당 {int(invoice_shift):,}원 → "
                "교대시간 없어 미지급"
            )
    else:
        stripped = invoice_ot + invoice_shift + invoice_night + invoice_special
        if stripped > 0:
            warnings.append(
                f"{name}: 청구서 시간수당 합계 {int(stripped):,}원 → "
                "명부 '통상시급' 없어 미지급"
            )

    return warnings


def _apply_roster_hourly_and_recalc(inv: dict[str, Any], emp_roster: dict[str, Any]) -> list[str]:
    """명부 시급·수당 정책 반영. 경고 문구 목록 반환."""
    from core.payroll.fixed_hours import apply_fixed_hours_to_invoice, resolve_employee_fixed_hours
    from services.workplace_hours import apply_monthly_hours_to_invoice

    workplace = str(emp_roster.get("근무지") or inv.get("workplace") or "").strip()
    if workplace:
        inv["workplace"] = workplace

    name = str(inv.get("name") or emp_roster.get("성명") or "")
    fixed_profile = resolve_employee_fixed_hours(
        employee_name=name,
        workplace=workplace,
        emp_roster=emp_roster,
    )
    if fixed_profile and fixed_profile.get("fixed_hours_mode"):
        apply_fixed_hours_to_invoice(inv, fixed_profile, workplace=workplace)

    base_hourly_master = safe_number(emp_roster.get("기본시급"), 0.0)
    ordinary_hourly_master = safe_number(emp_roster.get("통상시급"), 0.0)

    if base_hourly_master > 0:
        inv["base_hourly"] = base_hourly_master
    if ordinary_hourly_master > 0:
        inv["ordinary_hourly"] = ordinary_hourly_master

    allowance_warnings = apply_roster_allowance_policy(inv, emp_roster)

    base_hourly_new = safe_number(inv.get("base_hourly"), 0.0)
    base_days = safe_number(inv.get("base_days"), 0.0)
    leave_state = inv.get("_annual_leave_state")

    if base_hourly_new > 0:
        if not inv.get("_fixed_hours_mode"):
            monthly_h = apply_monthly_hours_to_invoice(inv, workplace)
        else:
            monthly_h = safe_number(inv.get("_monthly_work_hours"), 0.0)
        inv["base_salary"] = round_won(base_hourly_new * monthly_h)
        if base_days > 0 or monthly_h > 0:
            inv["base_deduction"] = calc_base_deduction_amount_with_leave(inv, leave_state)
        else:
            inv["base_deduction"] = 0
    elif leave_state is not None:
        inv["base_deduction"] = calc_base_deduction_amount_with_leave(inv, leave_state)

    _apply_early_leave_deduction(inv)
    return allowance_warnings


def calc_early_leave_deduction(inv: dict[str, Any]) -> int:
    """조퇴공제 금액만 ROUND: -(기본시급 × 조퇴시간). 시간은 소수 허용."""
    hours = safe_number(inv.get("early_leave_hours"), 0.0)
    hourly = safe_number(inv.get("base_hourly"), 0.0)
    if hours <= 0 or hourly <= 0:
        return 0
    return -round_won(hourly * hours)


def _apply_early_leave_deduction(inv: dict[str, Any]) -> None:
    inv["early_leave_deduction"] = calc_early_leave_deduction(inv)


def _unique_roster_names(employee_roster: dict[str, dict[str, Any]]) -> dict[str, str]:
    """명부에서 성명 목록을 추출합니다. {정규화키: 표시이름}."""
    names: dict[str, str] = {}
    for rec in employee_roster.values():
        if not isinstance(rec, dict):
            continue
        display = rec.get("성명") or rec.get("이름")
        if not display:
            continue
        key = _norm_name_key(display)
        if key and key not in names:
            names[key] = str(display).strip()
    return names


def compare_roster_invoice_personnel(
    invoice_rows: list[dict[str, Any]],
    employee_roster: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    """
    명부와 청구서 인원을 비교합니다.

    Returns:
        only_in_invoice: 청구서에만 있는 성명
        only_in_roster: 명부에만 있는 성명
        invoice_count, roster_count, matched_count
        has_mismatch
    """
    roster_names = _unique_roster_names(employee_roster)
    roster_keys = set(roster_names.keys())

    invoice_by_key: dict[str, str] = {}
    for row in invoice_rows:
        display = str(row["name"]).strip()
        key = _norm_name_key(display)
        if key:
            invoice_by_key[key] = display
    invoice_keys = set(invoice_by_key.keys())

    matched_keys = invoice_keys & roster_keys
    only_invoice_keys = set(invoice_keys - matched_keys)
    only_roster_keys = set(roster_keys - matched_keys)
    for inv_key in list(only_invoice_keys):
        fuzzy = find_fuzzy_name_key(inv_key, only_roster_keys)
        if fuzzy:
            matched_keys.add(inv_key)
            only_invoice_keys.discard(inv_key)
            only_roster_keys.discard(fuzzy)
    only_invoice_keys = sorted(only_invoice_keys)
    only_roster_keys = sorted(only_roster_keys)

    return {
        "only_in_invoice": [invoice_by_key[k] for k in only_invoice_keys],
        "only_in_roster": [roster_names[k] for k in only_roster_keys],
        "invoice_count": len(invoice_keys),
        "roster_count": len(roster_keys),
        "matched_count": len(matched_keys),
        "has_mismatch": bool(only_invoice_keys or only_roster_keys),
    }


def format_personnel_mismatch_message(diff: dict[str, Any]) -> str:
    """인원 불일치 알림창용 메시지 본문 (messagebox·전용 창 공용)."""
    only_inv = [
        str(n).strip()
        for n in (diff.get("only_in_invoice") or [])
        if n is not None and str(n).strip()
    ]
    only_rost = [
        str(n).strip()
        for n in (diff.get("only_in_roster") or [])
        if n is not None and str(n).strip()
    ]

    lines = [
        f"청구서 {diff.get('invoice_count', 0)}명, 명부 {diff.get('roster_count', 0)}명 "
        f"(일치 {diff.get('matched_count', 0)}명)",
        "",
    ]
    if only_inv:
        lines.append(f"[청구서에만 있음] {len(only_inv)}명")
        lines.extend(f"  - {name}" for name in only_inv)
        lines.append("")
    if only_rost:
        lines.append(f"[명부에만 있음] {len(only_rost)}명")
        lines.extend(f"  - {name}" for name in only_rost)
        lines.append("")
    lines.append("성명·띄어쓰기가 동일한지 확인해 주세요.")
    return "\n".join(lines).rstrip()


def roster_hourly_apply_stats(
    invoice_rows: list[dict[str, Any]],
    employee_roster: dict[str, dict[str, Any]],
    records: list[dict[str, Any]],
) -> dict[str, Any]:
    """명부 시급 반영 결과 요약(경고 메시지용)."""
    total = len(invoice_rows)
    matched = 0
    hourly_in_roster = 0
    hourly_applied = 0
    not_applied: list[str] = []

    rec_by_name = {r["name"]: r for r in records}

    for row in invoice_rows:
        name = row["name"]
        emp = _match_master(name, employee_roster)
        if not emp:
            continue
        matched += 1
        has_hourly = safe_number(emp.get("기본시급"), 0.0) > 0 or safe_number(emp.get("통상시급"), 0.0) > 0
        if not has_hourly:
            continue
        hourly_in_roster += 1
        out = rec_by_name.get(name)
        if not out:
            continue
        applied = False
        if safe_number(emp.get("기본시급"), 0.0) > 0:
            applied = abs(out["base_hourly"] - emp["기본시급"]) < 0.01
        if safe_number(emp.get("통상시급"), 0.0) > 0:
            applied = applied or abs(out["ordinary_hourly"] - emp["통상시급"]) < 0.01
        if applied:
            hourly_applied += 1
        else:
            not_applied.append(name)

    return {
        "total": total,
        "matched": matched,
        "hourly_in_roster": hourly_in_roster,
        "hourly_applied": hourly_applied,
        "not_applied": not_applied,
    }


def _load_prior_payroll_records_for_year(period: str) -> list[dict[str, Any]]:
    """신원보증보험료 연 1회 공제 확인용 — 당해 이전 월 스냅샷."""
    if len(period) < 7 or period[4] != "-":
        return []
    year = period[:4]
    try:
        month = int(period[5:7])
    except ValueError:
        return []
    if month <= 1:
        return []
    out: list[dict[str, Any]] = []
    try:
        from payroll_archive import load_snapshot_records
        from services.payroll_scope import discover_scopes

        for m in range(1, month):
            p = f"{year}-{m:02d}"
            for scope in discover_scopes():
                if scope.period != p:
                    continue
                for rec in load_snapshot_records(p, scope):
                    if isinstance(rec, dict):
                        out.append(rec)
    except ImportError:
        pass
    return out


def _apply_site_benefits_and_recalc_gross(
    inv: dict[str, Any],
    *,
    workplace: str,
    payroll_period: str,
    prior_records: list[dict[str, Any]] | None = None,
) -> None:
    """사업장 특수 항목 반영 후 소계·총지급 재계산."""
    from core.payroll.site_benefits import apply_site_benefits_to_invoice

    wp = str(workplace or inv.get("workplace") or "").strip()
    apply_site_benefits_to_invoice(
        inv,
        workplace=wp,
        payroll_period=payroll_period,
        prior_records=prior_records,
    )
    workers_day = int(safe_number(inv.get("workers_day_allowance"), 0))
    position_pay = int(safe_number(inv.get("position_pay"), 0.0))
    inv["subtotal"] = (
        int(safe_number(inv.get("base_salary"), 0.0))
        + int(safe_number(inv.get("base_deduction"), 0.0))
        + int(safe_number(inv.get("early_leave_deduction"), 0.0))
        + int(safe_number(inv.get("shutdown_allowance"), 0.0))
        + int(safe_number(inv.get("ot_pay"), 0.0))
        + int(safe_number(inv.get("shift_pay"), 0.0))
        + int(safe_number(inv.get("night_pay"), 0.0))
        + int(safe_number(inv.get("special_pay"), 0.0))
        + int(safe_number(inv.get("special_ext_pay"), 0.0))
        + position_pay
        + workers_day
    )
    transport = int(safe_number(inv.get("transport"), 0.0))
    inv["gross_pay"] = inv["subtotal"] + transport


def build_payroll_records(
    invoice_rows: list[dict[str, Any]],
    payment_master: dict[str, dict[str, Any]],
    payslip_master: dict[str, dict[str, Any]],
    employee_roster: dict[str, dict[str, Any]] | None = None,
    leave_ledger_entries: list | None = None,
    monthly_leave_summaries: list | None = None,
    payroll_period: str | None = None,
    leave_usage_decisions: dict[str, str | dict[str, float]] | None = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    """청구서 데이터 + 마스터를 합쳐 급여대장·명세서·지급내역용 레코드를 만듭니다.

    Returns:
        records, allowance_warnings (명부에 없는 수당을 청구서에서 제외한 내역)
    """
    from roster_workbook import LeaveLedgerEntry, MonthlyLeaveSummary

    records: list[dict[str, Any]] = []
    ledger_entries: list = leave_ledger_entries if leave_ledger_entries is not None else []
    monthly_summaries: list = (
        monthly_leave_summaries if monthly_leave_summaries is not None else []
    )
    period = payroll_period or datetime.now().strftime("%Y-%m")
    leave_decisions = leave_usage_decisions or {}
    allowance_warnings: list[str] = []
    prior_records = _load_prior_payroll_records_for_year(period)

    for seq, inv in enumerate(invoice_rows, 1):
        from shutdown_leave import ensure_reference_gross_pay

        ensure_reference_gross_pay(inv)
        name = inv["name"]
        pay_info = _match_master(name, payment_master)
        slip_info = _match_master(name, payslip_master)
        emp_roster = None
        if employee_roster:
            emp_roster = _match_master(name, employee_roster) or None

        if emp_roster and leave_decisions:
            decision = leave_decisions.get(_norm_name_key(name))
            if isinstance(decision, dict):
                apply_leave_split_to_invoice(
                    inv,
                    float(decision.get("annual_days", 0)),
                    float(decision.get("unpaid_days", 0)),
                    shutdown_days=float(decision.get("shutdown_days", 0)),
                )
            elif isinstance(decision, str) and decision:
                apply_leave_usage_decision_to_invoice(inv, decision)

        identity = (emp_roster or {}).get("주민번호") or slip_info.get("birth")
        income_tax = 0
        local_tax = 0
        net_pay = 0
        gross = inv["gross_pay"]

        # ------------------------------------------------------------
        # 직원정보.xlsx(근로자 명부) 우선 반영
        # ------------------------------------------------------------
        if emp_roster:
            # 전화번호/주민번호/입사일/근무지/계좌는 명부가 우선
            phone = emp_roster.get("휴대폰") or slip_info.get("phone", "")
            birth = emp_roster.get("주민번호") or slip_info.get("birth", "")
            hire_date = emp_roster.get("입사일") or inv.get("hire_date") or ""
            workplace = emp_roster.get("근무지") or pay_info.get("workplace", "한국앰코생산")
            affiliate = str(emp_roster.get("계열사") or "").strip() or get_default_affiliate()
            pay_acct = resolve_payment_from_roster(
                emp_roster, pay_info, employee_name=name
            )
            account = pay_acct["account"]

            # 당월 연차 사용 반영 → 사용연차/잔여연차 갱신(명부 저장용)
            leave_state = apply_monthly_annual_leave_to_invoice(inv, emp_roster, period)
            row_no = emp_roster.get("_row")
            can_write_ledger = isinstance(row_no, int) and row_no >= 2
            from shutdown_leave import pure_unpaid_days

            absence_days = pure_unpaid_days(inv)
            absence_count = calc_absence_occurrence_count(absence_days)
            absence_memo = (
                format_absence_usage_memo(period, absence_days) if absence_days > 0 else ""
            )

            if leave_state is not None and can_write_ledger and leave_state.month_used > 0:
                ledger_entries.append(
                    LeaveLedgerEntry(
                        main_row=row_no,
                        name=name,
                        emp_no=emp_roster.get("사번"),
                        period_label=period,
                        usage_month=format_usage_month_label(period),
                        days=leave_state.month_used,
                        usage_memo=leave_state.usage_memo_entry or "",
                        record_kind="연차",
                        accrued=leave_state.effective_accrued,
                        used_total=leave_state.used_after,
                        remaining=leave_state.remaining_after,
                    )
                )
            if absence_days > 0 and can_write_ledger:
                ledger_entries.append(
                    LeaveLedgerEntry(
                        main_row=row_no,
                        name=name,
                        emp_no=emp_roster.get("사번"),
                        period_label=period,
                        usage_month=format_usage_month_label(period),
                        days=absence_days,
                        usage_memo=absence_memo,
                        record_kind="결근",
                        occurrence_count=absence_count,
                    )
                )

            if can_write_ledger and roster_has_any_leave_column(emp_roster):
                ls = leave_state
                monthly_summaries.append(
                    MonthlyLeaveSummary(
                        name=name,
                        emp_no=emp_roster.get("사번"),
                        period_label=period,
                        accrued=ls.effective_accrued if ls else 0.0,
                        month_leave_used=ls.month_used if ls else 0.0,
                        used_total=ls.used_after if ls else 0.0,
                        remaining=ls.remaining_after if ls else 0.0,
                        absence_days=absence_days,
                        absence_count=absence_count,
                        leave_memo=(ls.usage_memo_entry or "") if ls and ls.month_used > 0 else "",
                        absence_memo=absence_memo,
                    )
                )

            # 기본시급·수당: 명부 기준(청구서 수당 잔존값 제거)
            allowance_warnings.extend(_apply_roster_hourly_and_recalc(inv, emp_roster))

            _apply_site_benefits_and_recalc_gross(
                inv,
                workplace=workplace,
                payroll_period=period,
                prior_records=prior_records,
            )
            gross = inv["gross_pay"]

            slip_info = dict(slip_info)
            pay_info = dict(pay_info)
            slip_info["phone"] = phone
            slip_info["birth"] = birth
            inv["hire_date"] = hire_date
            pay_info["workplace"] = workplace
            pay_info["affiliate"] = affiliate
            pay_info["account"] = account
            pay_info["holder"] = pay_acct["holder"]
            pay_info["bank_name"] = pay_acct["bank_name"]
            pay_info["bank_code"] = pay_acct["bank_code"]
        else:
            _apply_early_leave_deduction(inv)
            workplace = pay_info.get("workplace", "한국앰코생산")
            _apply_site_benefits_and_recalc_gross(
                inv,
                workplace=workplace,
                payroll_period=period,
                prior_records=prior_records,
            )
            gross = inv["gross_pay"]
            pay_info = dict(pay_info)
            pay_info.setdefault("affiliate", get_default_affiliate())

        income_tax_master = (
            safe_number(emp_roster.get("소득세"), 0.0) if emp_roster else 0.0
        )
        from core.session_service import session_tenant_id

        resolve_social_insurance(
            inv,
            identity=identity,
            payroll_period=period,
            emp_roster=emp_roster,
            tenant_id=session_tenant_id(),
        )
        ei_warn = str(inv.get("ei_65_warning") or "").strip()
        if ei_warn:
            allowance_warnings.append(ei_warn)
        gross = inv["gross_pay"]
        insurance = inv["insurance_total"]

        if income_tax_master > 0:
            income_tax = int(income_tax_master)
            local_preset = safe_number(emp_roster.get("지방소득세"), 0.0) if emp_roster else 0.0
            local_tax = int(local_preset) if local_preset > 0 else round_won_tens(income_tax * 0.10)
        else:
            taxable = gross - insurance
            tax = calculate_tax(taxable)
            income_tax = tax.income_tax
            local_tax = tax.local_income_tax

        total_deduction = (
            insurance
            + income_tax
            + local_tax
            + abs(int(safe_number(inv.get("identity_guarantee_insurance_deduction"), 0)))
        )
        net_pay = gross - total_deduction

        records.append({
            "seq": seq,
            "name": name,
            "dept": inv.get("dept") or "",
            "hire_date": inv.get("hire_date") or "",
            "birth": slip_info.get("birth", ""),
            "phone": slip_info.get("phone", ""),
            "pay_day": slip_info.get("pay_day", "25일"),
            "base_hourly": inv["base_hourly"],
            "ordinary_hourly": inv["ordinary_hourly"],
            "base_days": inv["base_days"],
            "work_days": inv["work_days"],
            "_monthly_work_hours": inv.get("_monthly_work_hours"),
            "_monthly_hours_source": inv.get("_monthly_hours_source"),
            "_fixed_hours_mode": bool(inv.get("_fixed_hours_mode")),
            "_fixed_hours_source": inv.get("_fixed_hours_source"),
            "_fixed_hours_job_group": inv.get("_fixed_hours_job_group"),
            "_monthly_hours_source": inv.get("_monthly_hours_source", ""),
            "unpaid_days": inv.get("unpaid_days", 0),
            "shutdown_leave_days": inv.get("shutdown_leave_days", 0),
            "shutdown_allowance": int(safe_number(inv.get("shutdown_allowance"), 0.0)),
            "shutdown_pay_percent": inv.get("_shutdown_pay_percent"),
            "leave_days": inv["leave_days"],
            "leave_usage_display": format_ledger_leave_usage_display(inv),
            "leave_sheet_memo": inv.get("leave_sheet_memo", ""),
            "leave_sheet_absence_memo": inv.get("leave_sheet_absence_memo", ""),
            "leave_sheet_leave_days": inv.get("leave_sheet_leave_days"),
            "leave_sheet_unpaid_days": inv.get("leave_sheet_unpaid_days"),
            "leave_sheet_leave_dates": inv.get("leave_sheet_leave_dates"),
            "leave_sheet_absence_dates": inv.get("leave_sheet_absence_dates"),
            "annual_leave_accrued": (
                inv.get("_annual_leave_state").accrued
                if inv.get("_annual_leave_state")
                else None
            ),
            "annual_leave_used_total": (
                inv.get("_annual_leave_state").used_after
                if inv.get("_annual_leave_state")
                else None
            ),
            "remaining_annual_leave": (
                inv.get("_annual_leave_state").remaining_after
                if inv.get("_annual_leave_state")
                else None
            ),
            "annual_leave_fully_used": (
                inv.get("_annual_leave_state").fully_exhausted
                if inv.get("_annual_leave_state")
                else None
            ),
            "ot_hours": inv["ot_hours"],
            "shift_hours": inv["shift_hours"],
            "night_hours": inv["night_hours"],
            "special_hours": inv["special_hours"],
            "special_ext_hours": inv["special_ext_hours"],
            "early_leave_hours": inv["early_leave_hours"],
            "early_leave_deduction": int(safe_number(inv.get("early_leave_deduction"), 0.0)),
            "base_salary": inv["base_salary"],
            "base_deduction": int(safe_number(inv.get("base_deduction"), 0.0)),
            "ot_pay": inv["ot_pay"],
            "night_pay": inv["night_pay"],
            "special_pay": inv["special_pay"],
            "special_ext_pay": inv["special_ext_pay"],
            "position_pay": inv["position_pay"],
            "shift_pay": inv["shift_pay"],
            "workers_day_allowance": int(safe_number(inv.get("workers_day_allowance"), 0)),
            "identity_guarantee_insurance_deduction": int(
                safe_number(inv.get("identity_guarantee_insurance_deduction"), 0)
            ),
            "annual_pay": inv["annual_pay"],
            "transport": inv["transport"],
            "gross_pay": gross,
            "health_insurance": inv["health_insurance"],
            "long_term_care": inv["long_term_care"],
            "national_pension": inv["national_pension"],
            "employment_insurance": inv["employment_insurance"],
            "insurance_exempt": bool(inv.get("insurance_exempt")),
            "edi_premium_source": bool(inv.get("edi_premium_source")),
            "edi_premium_badge": str(inv.get("edi_premium_badge") or ""),
            "industrial_accident": int(safe_number(inv.get("industrial_accident"), 0)),
            "income_tax": income_tax,
            "local_income_tax": local_tax,
            "total_deduction": total_deduction,
            "net_pay": round_won(net_pay),
            "workplace": pay_info.get("workplace", "한국앰코생산"),
            "affiliate": pay_info.get("affiliate", get_default_affiliate()),
            "bank_code": pay_info.get("bank_code", ""),
            "bank_name": pay_info.get("bank_name", ""),
            "account": pay_info.get("account", ""),
            "holder": pay_info.get("holder", name),
        })
        if emp_roster:
            from core.executive_policy import is_executive_roster_row

            records[-1]["직책"] = str(emp_roster.get("직책") or "")
            records[-1]["is_executive"] = is_executive_roster_row(emp_roster)
        else:
            records[-1]["직책"] = ""
            records[-1]["is_executive"] = False

    return records, allowance_warnings
