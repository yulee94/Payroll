"""
베스텍 재직증명서(2026) 명부 파싱 + 참조 급여대장 명부값 병합.

재직증명서 '근로자명부' 시트: 성명·부서·직책·입사일·사번
참조 급여대장: 기본시급·통상시급·수당·소득세·4대보험
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from roster_constants import norm_name_key
from services.bestec_reference_ledger import (
    BestecLedgerEmployee,
    reference_to_roster_row,
)
from utils import safe_number

CERT_SHEET = "근로자명부"
HEADER_ROW = 3
DATA_START = 4

_CERT_HEADERS = {
    "번호": ("번호", "no"),
    "성명": ("성명", "이름"),
    "부서": ("부서", "소속"),
    "직책": ("직책", "직급", "직위"),
    "입사일": ("입사일자", "입사일", "현재입사일"),
    "퇴사일": ("퇴사일자", "퇴사일"),
    "사번": ("사번", "사원번호"),
    "연락처": ("연락처", "휴대폰"),
    "이메일": ("이메일",),
}


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("\n", " ").replace(" ", "")


def _build_cert_header_map(row_values: tuple[Any, ...]) -> dict[str, int]:
    out: dict[str, int] = {}
    for idx, cell in enumerate(row_values, start=1):
        raw = _norm_header(cell)
        if not raw:
            continue
        for canonical, aliases in _CERT_HEADERS.items():
            if canonical in out:
                continue
            if raw in {_norm_header(a) for a in aliases}:
                out[canonical] = idx
                break
    return out


def _parse_date(value: Any) -> Any:
    if value is None or value == "" or str(value).strip() in ("-", "－"):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%y.%m.%d"):
        try:
            return datetime.strptime(text[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return value


def load_bestec_certificate_roster(path: Path) -> dict[str, dict[str, Any]]:
    """재직증명서 근로자명부 → {이름키: 명부행}."""
    if not path.is_file():
        return {}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if CERT_SHEET not in wb.sheetnames:
            return {}
        ws = wb[CERT_SHEET]
        header_row = None
        headers: dict[str, int] = {}
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1
        ):
            trial = _build_cert_header_map(row or ())
            if "성명" in trial:
                headers = trial
                header_row = row_idx
                break
        if not headers or "성명" not in headers:
            return {}

        roster: dict[str, dict[str, Any]] = {}
        empty_streak = 0
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=(header_row or HEADER_ROW) + 1, values_only=True),
            start=(header_row or HEADER_ROW) + 1,
        ):
            name_col = headers["성명"] - 1
            name = row[name_col] if name_col < len(row) else None
            if name is None or not str(name).strip():
                empty_streak += 1
                if empty_streak >= 20:
                    break
                continue
            empty_streak = 0
            name_s = str(name).strip()

            rec: dict[str, Any] = {
                "성명": name_s,
                "근무지": "밀양공장",
                "계열사": "㈜베스텍",
                "_main_sheet": CERT_SHEET,
                "_row": row_idx,
                "_source": "재직증명서",
            }
            for field, col in headers.items():
                if field in ("성명",):
                    continue
                idx = col - 1
                val = row[idx] if idx < len(row) else None
                if field in ("입사일", "퇴사일"):
                    rec[field] = _parse_date(val)
                elif field == "사번":
                    rec["사번"] = str(val).strip() if val not in (None, "") else ""
                else:
                    rec[field] = val

            exec_titles = ("회장", "부회장", "대표", "사장", "부사장", "전무", "상무", "이사")
            title = str(rec.get("직책") or "")
            if any(t in title for t in exec_titles):
                rec["임원"] = "Y"

            key = norm_name_key(name_s)
            if key:
                roster[key] = rec
            emp_no = rec.get("사번")
            if emp_no:
                roster[norm_name_key(emp_no)] = rec
        return roster
    finally:
        wb.close()


def merge_bestec_roster(
    certificate: dict[str, dict[str, Any]],
    reference: dict[str, BestecLedgerEmployee],
) -> dict[str, dict[str, Any]]:
    """재직증명서 + 참조 급여대장 명부 병합 (참조 시급·보험 우선)."""
    merged: dict[str, dict[str, Any]] = {}

    ref_rows = {
        k: reference_to_roster_row(v)
        for k, v in reference.items()
    }

    for key, ref_rec in ref_rows.items():
        cert_rec = certificate.get(key, {})
        row = dict(ref_rec)
        for field in (
            "성명",
            "부서",
            "직책",
            "입사일",
            "퇴사일",
            "사번",
            "연락처",
            "이메일",
            "임원",
        ):
            val = cert_rec.get(field)
            if val not in (None, ""):
                row[field] = val
        if not row.get("성명") and cert_rec.get("성명"):
            row["성명"] = cert_rec["성명"]
        row.setdefault("근무지", "밀양공장")
        row.setdefault("계열사", "㈜베스텍")
        if cert_rec:
            row["_cert_matched"] = True
        merged[key] = row

    for key, cert_rec in certificate.items():
        if key in merged:
            continue
        if not isinstance(cert_rec, dict) or not cert_rec.get("성명"):
            continue
        merged[key] = dict(cert_rec)

    return merged


def roster_stats(roster: dict[str, dict[str, Any]]) -> dict[str, int]:
    names = {
        k
        for k, rec in roster.items()
        if isinstance(rec, dict) and rec.get("성명") and not str(k).isdigit()
    }
    with_rate = sum(
        1
        for k in names
        if safe_number((roster.get(k) or {}).get("기본시급"), 0) > 0
    )
    with_tax = sum(
        1
        for k in names
        if safe_number((roster.get(k) or {}).get("소득세"), 0) > 0
    )
    cert_matched = sum(
        1 for k in names if (roster.get(k) or {}).get("_cert_matched")
    )
    return {
        "employee_count": len(names),
        "with_hourly_rate": with_rate,
        "with_income_tax": with_tax,
        "cert_matched": cert_matched,
    }
