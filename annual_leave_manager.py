"""
annual_leave_manager.py - 명부 연차(발생/사용/잔여) 자동 관리

- 명부 읽기: templates/근로자명부.xlsx (씨엔엘 시트)
- 연차 사용 기록: 연차사용대장/연차사용대장.xlsx (5년 보관)
- 매월 청구서 L열(휴가/연차) 반영
- 무급/결근 공제: 발생연차 전량 소진 후에만
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from roster_constants import norm_name_key
from roster_workbook import is_leave_deficit_display, read_leave_balance, sanitize_roster_number
from utils import round_won, safe_number

_PERIOD_TAG_RE = re.compile(r"\[(\d{4}-\d{2}):(\d+(?:\.\d+)?)\]")

_LEAVE_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "잔여연차": ("잔여연차", "잔여 연차", "연차잔여", "연차 잔여"),
    "발생연차": ("발생연차", "발생 연차", "연차발생", "연차 발생"),
    "사용연차": ("사용연차", "사용 연차", "연차사용", "연차 사용"),
    "사용월": ("사용월", "연차사용월", "사용 월"),
    "연차사용메모": ("연차사용메모", "연차 사용메모", "연차메모", "연차 사용 메모"),
    "성명": ("성명", "이름", "사원명"),
}

_LEAVE_SAVE_FIELDS = frozenset(
    {"사용연차", "잔여연차", "발생연차", "사용월", "연차사용메모"}
)


def _build_leave_header_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        raw = str(v).strip().replace("\n", " ")
        for canonical, aliases in _LEAVE_HEADER_ALIASES.items():
            if raw in aliases:
                headers[canonical] = c
                break
    return headers


@dataclass
class AnnualLeaveState:
    """한 직원의 당월 연차 처리 결과."""

    accrued: float
    expected_accrual: float
    effective_accrued: float
    uses_expected_accrual: bool
    used_before: float
    month_used: float
    used_after: float
    remaining_after: float
    fully_exhausted: bool
    usage_month_label: str
    usage_memo_entry: str


def format_usage_month_label(period_label: str) -> str:
    """급여월 표기: 2026-05 → 05월"""
    try:
        month = int(period_label.split("-")[1])
        return f"{month:02d}월"
    except (IndexError, ValueError):
        return period_label


def _usage_month_from_period(period_label: str) -> int:
    try:
        return int(period_label.split("-")[1])
    except (IndexError, ValueError):
        return datetime.now().month


def calc_absence_occurrence_count(days: float) -> int:
    """무급/결근 횟수 — 정수 일수는 1일=1회, 반차 등은 1회 추가."""
    days = max(0.0, safe_number(days, 0.0))
    if days <= 1e-9:
        return 0
    whole = int(days)
    frac = days - whole
    if whole >= 1 and frac <= 1e-9:
        return whole
    if whole >= 1:
        return whole + 1
    return 1


def format_absence_usage_memo(period_label: str, days: float) -> str:
    """결근/무급 사용내역 (연차사용대장용). 일수는 소수 허용."""
    days = max(0.0, safe_number(days, 0.0))
    if days <= 0:
        return "결근/무급"
    month = _usage_month_from_period(period_label)
    whole = int(days)
    frac = days - whole
    if whole >= 1 and frac <= 1e-9:
        return ", ".join(
            f"{month:02d}월 {d:02d}일 결근/무급" for d in range(1, whole + 1)
        )
    if whole >= 1 and frac > 1e-9:
        parts = [f"{month:02d}월 {d:02d}일 결근/무급" for d in range(1, whole + 1)]
        parts.append(f"{month:02d}월 반차({frac:g}일) 결근/무급")
        return ", ".join(parts)
    return f"{month:02d}월 결근/무급({days:g}일)"


def format_usage_date_memos(period_label: str, days: float) -> str:
    """
    연차 사용 메모. 일수·반차는 소수 그대로(반올림 없음).

    예) 0.5일 → '05월 반차(0.5일) 사용', 2.5일 → 01·02일 + 반차(0.5일)
    """
    days = max(0.0, safe_number(days, 0.0))
    if days <= 0:
        return ""
    month = _usage_month_from_period(period_label)
    whole = int(days)
    frac = days - whole
    if frac > 1e-9 and whole == 0:
        return f"{month:02d}월 반차({days:g}일) 사용"
    parts: list[str] = []
    if whole >= 1:
        parts.extend(
            f"{month:02d}월 {d:02d}일 사용" for d in range(1, whole + 1)
        )
    if frac > 1e-9:
        parts.append(f"{month:02d}월 반차({frac:g}일) 사용")
    if not parts:
        return f"{month:02d}월 {days:g}일 사용"
    return ", ".join(parts)


def parse_period_leave_from_memo(memo: str, period_label: str) -> float:
    """메모에 기록된 해당 급여월 연차 사용일(재처리 시 차감용)."""
    total = 0.0
    for m in _PERIOD_TAG_RE.finditer(memo or ""):
        if m.group(1) == period_label:
            total += float(m.group(2))
    return total


def append_leave_usage_memo(
    existing: str,
    period_label: str,
    month_used: float,
) -> str:
    """연차사용메모에 [YYYY-MM:N] 태그와 'MM월 DD일 사용' 문구를 추가·갱신합니다."""
    existing = (existing or "").strip()
    tag = f"[{period_label}:{int(month_used) if month_used == int(month_used) else month_used}]"
    usage_text = format_usage_date_memos(period_label, month_used)

    if month_used <= 0:
        parts = [p.strip() for p in existing.split("|") if p.strip()]
        parts = [p for p in parts if not p.startswith(f"[{period_label}:")]
        return " | ".join(parts)

    new_block = f"{tag} {usage_text}" if usage_text else tag

    parts = [p.strip() for p in existing.split("|") if p.strip()]
    parts = [p for p in parts if not p.startswith(f"[{period_label}:")]
    parts.append(new_block)
    return " | ".join(parts)


def is_roster_leave_deficit(emp_roster: dict[str, Any]) -> bool:
    """잔여연차가 '-' 또는 음수 표기인지."""
    return bool(emp_roster.get("_잔여연차_초과")) or is_leave_deficit_display(
        emp_roster.get("잔여연차_raw", emp_roster.get("잔여연차"))
    )


def format_remaining_leave_display(emp_roster: dict[str, Any]) -> str:
    """알림창·로그용 잔여연차 표시."""
    raw = emp_roster.get("잔여연차_raw", emp_roster.get("잔여연차"))
    if is_leave_deficit_display(raw):
        return "-"
    bal = read_leave_balance(raw)
    if bal is None:
        return "-"
    if abs(bal - round(bal)) < 1e-9:
        return str(int(round(bal)))
    return f"{bal:g}"


def _match_roster_record(name: str, roster: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    key = norm_name_key(name)
    if key in roster:
        rec = roster[key]
        return rec if isinstance(rec, dict) else None
    for rec in roster.values():
        if not isinstance(rec, dict):
            continue
        for field in ("성명", "이름", "name"):
            nm = rec.get(field)
            if nm and norm_name_key(nm) == key:
                return rec
    return None


def available_annual_leave_before_month(
    emp_roster: dict[str, Any],
    period_label: str,
) -> float | None:
    """
    당월 연차 반영 전 사용 가능 일수.

    '-'(초과) 표기면 None(잔여 없음).
    """
    if is_roster_leave_deficit(emp_roster):
        return None

    accrued, _expected, effective, _uses_expected = resolve_leave_entitlement(emp_roster)
    used_before = _used_before_this_period(emp_roster, period_label, effective)

    if effective > 0:
        return max(0.0, effective - used_before)

    bal = read_leave_balance(
        emp_roster.get("잔여연차_raw", emp_roster.get("잔여연차"))
    )
    if bal is None:
        return None
    return max(0.0, bal)


def has_sufficient_annual_leave_for_month(
    emp_roster: dict[str, Any],
    period_label: str,
    month_used: float,
) -> bool:
    """잔여연차가 당월 사용일수 이상이면 True → 연차 자동 처리."""
    if month_used <= 0:
        return True
    available = available_annual_leave_before_month(emp_roster, period_label)
    if available is None:
        return False
    return available >= month_used - 1e-9


def find_deficit_leave_usage_cases(
    invoice_rows: list[dict[str, Any]],
    employee_roster: dict[str, dict[str, Any]],
    period_label: str,
) -> list[dict[str, Any]]:
    """
    잔여연차가 없거나 부족한데 당월 연차 사용(L열)이 있는 직원.

    잔여가 충분한 직원은 목록에 넣지 않음(자동 연차 처리).
    """
    cases: list[dict[str, Any]] = []
    seen: set[str] = set()

    for inv in invoice_rows:
        month_used = max(0.0, safe_number(inv.get("leave_days"), 0.0))
        if month_used <= 0:
            continue
        name = str(inv.get("name") or "").strip()
        if not name:
            continue
        emp = _match_roster_record(name, employee_roster)
        if not emp or not roster_has_any_leave_column(emp):
            continue

        if has_sufficient_annual_leave_for_month(emp, period_label, month_used):
            continue

        accrued, expected, effective, _uses_expected = resolve_leave_entitlement(emp)
        used_before = _used_before_this_period(emp, period_label, effective)
        available = available_annual_leave_before_month(emp, period_label)

        name_key = norm_name_key(name)
        if name_key in seen:
            continue
        seen.add(name_key)

        cases.append(
            {
                "name_key": name_key,
                "name": name,
                "workplace": str(inv.get("workplace") or emp.get("근무지") or "").strip(),
                "month_used": month_used,
                "remaining_display": format_remaining_leave_display(emp),
                "accrual_basis": str(emp.get("_leave_accrual_basis") or "").strip(),
                "accrued": accrued,
                "expected_accrual": expected,
                "effective_accrued": effective,
                "used_before": used_before,
                "available": available,
            }
        )
    return cases


def count_auto_annual_leave_usage(
    invoice_rows: list[dict[str, Any]],
    employee_roster: dict[str, dict[str, Any]],
    period_label: str,
) -> int:
    """잔여 연차가 있어 확인 없이 연차 처리되는 인원 수."""
    count = 0
    for inv in invoice_rows:
        month_used = max(0.0, safe_number(inv.get("leave_days"), 0.0))
        if month_used <= 0:
            continue
        name = str(inv.get("name") or "").strip()
        if not name:
            continue
        emp = _match_roster_record(name, employee_roster)
        if not emp or not roster_has_any_leave_column(emp):
            continue
        if has_sufficient_annual_leave_for_month(emp, period_label, month_used):
            count += 1
    return count


def _mark_unpaid_on_invoice(inv: dict[str, Any], unpaid_part: float) -> bool:
    if unpaid_part <= 0:
        return False
    inv["unpaid_days"] = safe_number(inv.get("unpaid_days"), 0.0) + unpaid_part
    existing_dept = str(inv.get("dept") or "").strip()
    if not existing_dept:
        inv["dept"] = "결근/무급"
    elif "결근" not in existing_dept and "무급" not in existing_dept:
        inv["dept"] = f"{existing_dept} 결근/무급"
    return True


def apply_leave_split_to_invoice(
    inv: dict[str, Any],
    annual_days: float,
    unpaid_days: float,
    *,
    shutdown_days: float = 0.0,
    shutdown_pay_percent: float | None = None,
) -> bool:
    """유급 연차·휴업(회사)·무급/결근 일수를 지정하여 청구서 행에 반영."""
    from services.payroll_settings_store import get_shutdown_pay_percent

    annual_days = max(0.0, float(annual_days))
    shutdown_days = max(0.0, float(shutdown_days))
    unpaid_days = max(0.0, float(unpaid_days))
    max_avail = inv.get("_leave_available_before")
    if max_avail is not None:
        cap = max(0.0, float(max_avail))
        if annual_days > cap + 1e-9:
            excess = annual_days - cap
            annual_days = cap
            shutdown_days += excess
    inv["leave_days"] = annual_days
    inv["shutdown_leave_days"] = shutdown_days
    inv["unpaid_days"] = unpaid_days
    inv["_leave_denied_as_unpaid"] = unpaid_days + shutdown_days
    workplace = str(inv.get("workplace") or "").strip()
    inv["_shutdown_pay_percent"] = (
        float(shutdown_pay_percent)
        if shutdown_pay_percent is not None
        else get_shutdown_pay_percent(workplace or None)
    )
    inv["_leave_split_applied"] = True
    changed = _mark_unpaid_on_invoice(inv, unpaid_days)
    if shutdown_days > 0:
        dept = str(inv.get("dept") or "").strip()
        if dept and "휴업" not in dept:
            inv["dept"] = f"{dept} 휴업"
        elif not dept:
            inv["dept"] = "휴업"
    return changed or annual_days > 0 or shutdown_days > 0


def default_leave_split_for_case(case: dict[str, Any]) -> dict[str, float]:
    """잔여 연차 기준 기본 분할(유급=min(잔여,당월), 나머지 무급·휴업 0)."""
    month_used = max(0.0, safe_number(case.get("month_used"), 0.0))
    available = case.get("available")
    if available is None:
        annual = 0.0
    else:
        annual = min(max(0.0, float(available)), month_used)
    unpaid = max(0.0, month_used - annual)
    return {"annual_days": annual, "shutdown_days": 0.0, "unpaid_days": unpaid}


def apply_leave_usage_decision_to_invoice(inv: dict[str, Any], decision: str) -> bool:
    """
    잔여 부족 연차 사용 처리 (구형 라디오 선택).

    - grant: 청구서 연차 그대로 반영(초과분도 연차 부여)
    - unpaid: 잔여 연차만큼은 연차 처리, 초과분만 결근/무급으로 전환
    """
    if decision == "grant":
        return False

    month_used = max(0.0, safe_number(inv.get("leave_days"), 0.0))
    if month_used <= 0:
        return False

    available_raw = inv.get("_leave_available_before")
    if available_raw is None:
        available_leave = 0.0
    elif isinstance(available_raw, (int, float)) and float(available_raw) < 0:
        available_leave = 0.0
    else:
        available_leave = max(0.0, float(available_raw))

    annual_part = min(available_leave, month_used)
    unpaid_part = month_used - annual_part
    return apply_leave_split_to_invoice(inv, annual_part, unpaid_part)


def roster_has_any_leave_column(emp_roster: dict[str, Any]) -> bool:
    for key in ("잔여연차", "발생연차", "사용연차", "예상발생연차"):
        if key in emp_roster:
            return True
    return False


def get_expected_accrual(emp_roster: dict[str, Any]) -> float:
    """명부의 예상·예정 발생 연차(다음 발생 예정 일수)."""
    if emp_roster.get("_hire_based_leave"):
        return max(0.0, sanitize_roster_number(emp_roster.get("예상발생연차")))
    if "예상발생연차" not in emp_roster:
        return 0.0
    return max(0.0, sanitize_roster_number(emp_roster.get("예상발생연차")))


def resolve_leave_entitlement(emp_roster: dict[str, Any]) -> tuple[float, float, float, bool]:
    """
    연차 한도(발생 기준) 결정.

    입사일 기준 산정(_hire_based_leave)이면 발생연차·예상발생연차를 우선합니다.
    잔여연차가 '-' 또는 음수이면 발생연차를 넘겨 쓴 상태로 보고,
    예상발생연차(있으면)를 한도로 사용합니다.
    """
    accrued = max(0.0, sanitize_roster_number(emp_roster.get("발생연차")))
    if emp_roster.get("_hire_based_leave") and accrued <= 0:
        from annual_leave_accrual import parse_hire_date, compute_leave_accrual

        hire = parse_hire_date(emp_roster.get("입사일"))
        period = str(emp_roster.get("_leave_period_hint") or "")
        if hire and period:
            accrued = compute_leave_accrual(hire, period).accrued
    expected = get_expected_accrual(emp_roster)
    overdrawn = bool(emp_roster.get("_잔여연차_초과")) or is_leave_deficit_display(
        emp_roster.get("잔여연차_raw", emp_roster.get("잔여연차"))
    )

    if overdrawn and expected > 0:
        effective = max(accrued, expected)
        uses_expected = expected >= accrued
    elif overdrawn and accrued > 0:
        effective = accrued
        uses_expected = False
    else:
        effective = accrued if accrued > 0 else expected
        uses_expected = accrued <= 0 and expected > 0

    return accrued, expected, effective, uses_expected


def _used_before_this_period(
    emp_roster: dict[str, Any],
    period_label: str,
    effective_accrued: float,
) -> float:
    """이번 급여월 반영 전 누적 사용연차(같은 달 재처리 시 이중 누적 방지)."""
    from roster_leave_sheet import cumulative_used_before_period

    monthly = emp_roster.get("_monthly_leave_usage")
    if isinstance(monthly, dict) and monthly:
        return cumulative_used_before_period(emp_roster, period_label)

    if "사용연차" in emp_roster:
        used_total = max(0.0, sanitize_roster_number(emp_roster.get("사용연차")))
    elif effective_accrued > 0 and "잔여연차" in emp_roster:
        bal = read_leave_balance(emp_roster.get("잔여연차_raw", emp_roster.get("잔여연차")))
        if bal is not None:
            used_total = max(0.0, effective_accrued - bal)
        else:
            used_total = max(0.0, effective_accrued)
    else:
        used_total = 0.0

    prev_in_period = parse_period_leave_from_memo(
        str(emp_roster.get("연차사용메모") or ""), period_label
    )
    return max(0.0, used_total - prev_in_period)


def compute_annual_leave_state(
    emp_roster: dict[str, Any],
    inv: dict[str, Any],
    period_label: str,
) -> AnnualLeaveState | None:
    if not roster_has_any_leave_column(emp_roster):
        return None

    accrued, expected, effective_accrued, uses_expected = resolve_leave_entitlement(emp_roster)
    month_used = max(0.0, safe_number(inv.get("leave_days"), 0.0))
    used_before = _used_before_this_period(emp_roster, period_label, effective_accrued)

    if effective_accrued > 0:
        used_after = used_before + month_used
        remaining_after = effective_accrued - used_after
        fully_exhausted = used_after >= effective_accrued - 1e-9
    else:
        remaining_before = read_leave_balance(
            emp_roster.get("잔여연차_raw", emp_roster.get("잔여연차"))
        )
        if remaining_before is None:
            remaining_before = 0.0
        prev_applied = parse_period_leave_from_memo(
            str(emp_roster.get("연차사용메모") or ""), period_label
        )
        remaining_after = remaining_before + prev_applied - month_used
        used_after = used_before + month_used
        fully_exhausted = remaining_after <= 1e-9

    usage_month_label = format_usage_month_label(period_label)
    usage_memo_entry = format_usage_date_memos(period_label, month_used)

    return AnnualLeaveState(
        accrued=accrued,
        expected_accrual=expected,
        effective_accrued=effective_accrued,
        uses_expected_accrual=uses_expected,
        used_before=used_before,
        month_used=month_used,
        used_after=used_after,
        remaining_after=remaining_after,
        fully_exhausted=fully_exhausted,
        usage_month_label=usage_month_label,
        usage_memo_entry=usage_memo_entry,
    )


def calc_base_deduction_days_with_leave(
    inv: dict[str, Any],
    leave_state: AnnualLeaveState | None,
) -> float:
    from shutdown_leave import pure_unpaid_days

    unpaid_days = pure_unpaid_days(inv)
    base_days = safe_number(inv.get("base_days"), 0.0)
    work_days = safe_number(inv.get("work_days"), 0.0)
    leave_days = safe_number(inv.get("leave_days"), 0.0)

    if leave_state is not None:
        # 연차 한도 없음(발생·예상 모두 0, 초과표시도 없음): 무급/결근 공제
        if leave_state.effective_accrued <= 0 and not leave_state.uses_expected_accrual:
            if unpaid_days > 0:
                return unpaid_days
            return max(0.0, base_days - work_days - leave_days)

        # 명시적 무급/결근 일수는 잔여 연차 유무와 관계없이 공제
        if unpaid_days > 0:
            return unpaid_days

        if not leave_state.fully_exhausted:
            return 0.0
        return max(0.0, base_days - work_days - leave_days)

    if unpaid_days > 0:
        return unpaid_days
    return max(0.0, base_days - work_days)


def calc_base_deduction_amount_with_leave(
    inv: dict[str, Any],
    leave_state: AnnualLeaveState | None,
) -> int:
    """
    기본공제 금액만 ROUND: -(기본시급 × 공제일수 × 8).

    휴업(회사 사정) 일수는 총급여 비율로 별도 반영(기본공제에 합산).
    """
    from shutdown_leave import calc_shutdown_gross_adjustments

    base_hourly = safe_number(inv.get("base_hourly"), 0.0)
    days = calc_base_deduction_days_with_leave(inv, leave_state)
    hourly_part = 0
    if base_hourly > 0 and days > 0:
        hourly_part = -round_won(base_hourly * 8 * days)

    shutdown_days = max(0.0, safe_number(inv.get("shutdown_leave_days"), 0.0))
    if shutdown_days <= 0:
        return hourly_part

    pct = inv.get("_shutdown_pay_percent")
    try:
        pay_percent = float(pct) if pct is not None else None
    except (TypeError, ValueError):
        pay_percent = None
    _allowance, shutdown_ded = calc_shutdown_gross_adjustments(
        inv, shutdown_days, pay_percent=pay_percent
    )
    inv["shutdown_allowance"] = _allowance
    return hourly_part + shutdown_ded


def apply_monthly_annual_leave_to_invoice(
    inv: dict[str, Any],
    emp_roster: dict[str, Any],
    period_label: str,
) -> AnnualLeaveState | None:
    state = compute_annual_leave_state(emp_roster, inv, period_label)
    if state is None:
        return None
    inv["_annual_leave_state"] = state
    return state


def roster_leave_update_from_state(
    state: AnnualLeaveState,
    period_label: str,
    emp_roster: dict[str, Any],
) -> dict[str, Any]:
    """명부에 저장할 연차·사용월·메모."""
    out: dict[str, Any] = {
        "사용연차": state.used_after,
        "잔여연차": state.remaining_after,
    }
    if state.accrued > 0:
        out["발생연차"] = state.accrued

    existing_memo = str(emp_roster.get("연차사용메모") or "")
    out["연차사용메모"] = append_leave_usage_memo(
        existing_memo, period_label, state.month_used
    )
    if state.month_used > 0 or period_label:
        out["사용월"] = format_usage_month_label(period_label)

    return out


def save_roster_annual_leave_updates(
    roster_path: Path,
    updates_by_row: dict[int, dict[str, Any]],
) -> None:
    if not updates_by_row or not roster_path.exists():
        return

    wb = openpyxl.load_workbook(roster_path)
    ws = wb.active
    headers = _build_leave_header_map(ws)

    def ensure_col(canonical: str) -> int:
        if canonical in headers:
            return headers[canonical]
        label = _LEAVE_HEADER_ALIASES.get(canonical, (canonical,))[0]
        col = ws.max_column + 1
        ws.cell(1, col, label)
        headers[canonical] = col
        return col

    for row, vals in updates_by_row.items():
        if row < 2:
            continue
        for field, value in vals.items():
            if field not in _LEAVE_SAVE_FIELDS:
                continue
            col = ensure_col(field)
            if field in ("사용연차", "잔여연차", "발생연차") and isinstance(value, float):
                ws.cell(row, col, round(value, 2))
            else:
                ws.cell(row, col, value)

    wb.save(roster_path)
    wb.close()


def annual_leave_roster_warnings(
    invoice_rows: list[dict[str, Any]],
    employee_roster: dict[str, dict[str, Any]],
) -> list[str]:
    if not employee_roster:
        return []

    has_accrued_col = any(
        isinstance(rec, dict) and "발생연차" in rec for rec in employee_roster.values()
    )
    needs = any(
        safe_number(r.get("unpaid_days"), 0) > 0
        or safe_number(r.get("leave_days"), 0) > 0
        for r in invoice_rows
    )
    if not needs:
        return []

    if not has_accrued_col:
        return [
            "templates/근로자명부.xlsx에 '발생연차' 열이 없습니다. "
            "연차 자동 관리를 위해 발생연차·사용연차(또는 잔여연차)를 입력해 주세요."
        ]
    return []
