"""
roster_leave_sheet.py - 근로자명부 '연차' 시트에서 발생/사용/잔여 연차 파싱

씨앤엘 시트의 VLOOKUP 수식은 저장 시 계산값이 없으면 None → 0 으로 읽혀
연차 시트 원본을 직접 파싱합니다.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from roster_constants import norm_name_key
from annual_leave_accrual import apply_hire_date_leave_to_record

_MONTH_HEADER_RE = re.compile(r"^(\d{2})\.(\d{2})$")


@dataclass
class AnnualLeaveSnapshot:
    accrued: float
    used_total: float
    remaining: float
    monthly_usage: dict[str, float]
    usage_memo_lines: list[str]


def _parse_usage_cell(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("일", "").replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _last_numeric_in_row(ws: Worksheet, row: int, col_start: int, col_end: int) -> float | None:
    """=LOOKUP(…, 마지막 비어 있지 않은 숫자) 과 동일."""
    last: float | None = None
    for c in range(col_start, col_end + 1):
        v = ws.cell(row, c).value
        if isinstance(v, (int, float)):
            last = float(v)
        elif v is not None and str(v).strip():
            parsed = _parse_usage_cell(v)
            if parsed > 0 or (isinstance(v, (int, float))):
                last = parsed
    return last


def _month_columns(ws: Worksheet, header_row: int = 4) -> dict[str, int]:
    out: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if h is None:
            continue
        text = str(h).strip()
        if _MONTH_HEADER_RE.match(text):
            out[text] = c
    return out


def _year_month_prefix(period_label: str) -> str:
    """2026-05 → '26.' (연차 시트 월 열 접두)."""
    try:
        y, _m = period_label.split("-")
        return f"{int(y) % 100:02d}."
    except (ValueError, AttributeError):
        return "26."


def _months_through(period_label: str, month_cols: dict[str, int]) -> list[str]:
    """YYYY-MM 까지의 연차 시트 월 열(26.MM) 목록."""
    try:
        y, m = period_label.split("-")
        yy = int(y) % 100
        mm = int(m)
    except (ValueError, AttributeError):
        return sorted(month_cols.keys())

    out: list[str] = []
    for label in sorted(month_cols.keys()):
        mobj = _MONTH_HEADER_RE.match(label)
        if not mobj:
            continue
        ly, lm = int(mobj.group(1)), int(mobj.group(2))
        if ly < yy or (ly == yy and lm <= mm):
            out.append(label)
    return out


def parse_leave_annual_sheet(wb: Workbook) -> dict[str, AnnualLeaveSnapshot]:
    """연차 시트 → {이름키: 스냅샷}."""
    if "연차" not in wb.sheetnames:
        return {}

    ws = wb["연차"]
    month_cols = _month_columns(ws)
    # 사용 합계 =SUM(EK6:EP6) → 26.01~26.05 열
    usage_sum_cols = [c for label, c in sorted(month_cols.items()) if label.startswith("26.")]

    result: dict[str, AnnualLeaveSnapshot] = {}
    row = 5
    max_row = min(ws.max_row or 0, 500)

    while row <= max_row:
        name = ws.cell(row, 1).value
        if name is None or not str(name).strip():
            row += 1
            continue

        name_s = str(name).strip()
        block_start = row
        row += 1
        while row <= max_row:
            next_name = ws.cell(row, 1).value
            if next_name is not None and str(next_name).strip():
                break
            row += 1
        block_end = row

        accrued_raw = _last_numeric_in_row(ws, block_start, 6, 147)
        accrued = max(0.0, accrued_raw if accrued_raw is not None else 0.0)

        monthly: dict[str, float] = {}
        usage_memo_lines: list[str] = []
        for r in range(block_start, block_end):
            if str(ws.cell(r, 5).value or "").strip() != "개수":
                continue
            for mlabel, col in month_cols.items():
                days = _parse_usage_cell(ws.cell(r, col).value)
                if days > 0:
                    monthly[mlabel] = days
                    usage_memo_lines.append(f"{mlabel}:{days:g}")

        used_total = sum(
            days
            for label, days in monthly.items()
            if str(label).startswith("26.")
        )

        remaining = accrued - used_total
        if remaining < 0 and accrued > 0:
            pass
        elif accrued <= 0 and used_total <= 0:
            remaining = 0.0

        snap = AnnualLeaveSnapshot(
            accrued=accrued,
            used_total=used_total,
            remaining=max(0.0, remaining) if accrued > 0 else remaining,
            monthly_usage=monthly,
            usage_memo_lines=usage_memo_lines,
        )
        key = norm_name_key(name_s)
        if key:
            result[key] = snap

    return result


def leave_snapshot_for_period(
    snap: AnnualLeaveSnapshot,
    period_label: str,
) -> dict[str, Any]:
    """특정 급여월 시점의 발생/사용/잔여(당해 연도 26.xx 월 열만 합산)."""
    try:
        y, m = period_label.split("-")
        yy, mm = int(y) % 100, int(m)
    except (ValueError, AttributeError):
        yy, mm = 26, 5

    year_prefix = f"{yy:02d}."
    used = 0.0
    monthly_this_year: dict[str, float] = {}
    for label, days in snap.monthly_usage.items():
        if not str(label).startswith(year_prefix):
            continue
        mobj = _MONTH_HEADER_RE.match(label)
        if not mobj:
            continue
        ly, lm = int(mobj.group(1)), int(mobj.group(2))
        if ly < yy or (ly == yy and lm <= mm):
            used += days
            monthly_this_year[label] = days

    remaining = snap.accrued - used if snap.accrued > 0 else snap.remaining
    return {
        "발생연차": snap.accrued,
        "사용연차": used,
        "잔여연차": remaining,
        "잔여연차_raw": remaining if remaining >= 0 else "-",
        "_잔여연차_초과": remaining < 0,
        "_monthly_leave_usage": monthly_this_year,
    }


def merge_leave_into_roster_record(
    rec: dict[str, Any],
    snap: AnnualLeaveSnapshot | None,
    period_label: str,
) -> None:
    """연차 시트 + 입사일 기준 발생을 명부 행에 병합."""
    sheet_leave: dict[str, Any] | None = None
    if snap is not None:
        sheet_leave = leave_snapshot_for_period(snap, period_label)
    if apply_hire_date_leave_to_record(rec, period_label, sheet_leave=sheet_leave):
        return
    if sheet_leave is not None:
        rec.update(sheet_leave)
        for col_name in ("발생연차", "사용연차", "잔여연차"):
            rec[f"_{col_name}_수식"] = False


def latest_period_hint_from_snapshot(snap: AnnualLeaveSnapshot) -> str:
    """연차 시트에 기록된 마지막 월 → '2026-05' (명부 UI 기본)."""
    best_yy, best_mm = 0, 0
    for label in snap.monthly_usage:
        mobj = _MONTH_HEADER_RE.match(str(label))
        if not mobj:
            continue
        ly, lm = int(mobj.group(1)), int(mobj.group(2))
        if ly > best_yy or (ly == best_yy and lm > best_mm):
            best_yy, best_mm = ly, lm
    if best_yy:
        return f"20{best_yy:02d}-{best_mm:02d}"
    return "2026-12"


def cumulative_used_before_period(
    emp_roster: dict[str, Any],
    period_label: str,
) -> float:
    """당월 제외 누적 사용일(연차 시트 월별 데이터 우선)."""
    monthly = emp_roster.get("_monthly_leave_usage")
    if not isinstance(monthly, dict) or not monthly:
        return 0.0

    try:
        y, m = period_label.split("-")
        yy, mm = int(y) % 100, int(m)
    except (ValueError, AttributeError):
        return 0.0

    year_prefix = f"{yy:02d}."
    total = 0.0
    for label, days in monthly.items():
        if not str(label).startswith(year_prefix):
            continue
        mobj = _MONTH_HEADER_RE.match(str(label))
        if not mobj:
            continue
        ly, lm = int(mobj.group(1)), int(mobj.group(2))
        if ly < yy or (ly == yy and lm < mm):
            total += _parse_usage_cell(days)
    return total
