"""
excel_writer.py - 실무 양식 기반 Excel 출력

templates/ 의 급여대장·급여명세서·지급내역 양식을 복사한 뒤
계산 결과를 해당 셀 위치에 기록합니다.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from core.file_save import copy_template, save_workbook
from core.paths import dev_root, output_dir, templates_dir
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from insurance import EMPLOYMENT_INSURANCE_WORKER_RATE
from payroll_builder import ledger_base_deduction_for_excel
from utils import format_hours_or_days, safe_number, write_merged_safe

BASE_DIR = dev_root()
TEMPLATES_DIR = templates_dir()
OUTPUT_DIR = output_dir()

# 급여대장: 직원 1명당 2행 (5~6행이 1번째 직원)
LEDGER_DATA_START = 5
LEDGER_ROWS_PER_EMP = 2

# 급여명세서: 3행부터 1행=1명
PAYSLIP_DATA_START = 3

# 지급내역: 5행부터
PAYMENT_DATA_START = 5


def _set_cell(ws, row: int, col: int, value) -> None:
    """병합셀 포함 모든 셀에 안전하게 값을 씁니다."""
    write_merged_safe(ws, f"{get_column_letter(col)}{row}", value)


def _clear_rows(ws, start_row: int, end_row: int, max_col: int = 30) -> None:
    """기존 샘플 데이터를 지웁니다 (병합셀은 건너뜀)."""
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _find_ledger_section_sum_rows(ws) -> list[int]:
    """급여대장 양식의 구간별 '합계' 행(총합계 제외)을 위에서부터 반환합니다."""
    rows: list[int] = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and "합계" in v and "총합계" not in v:
            rows.append(r)
    return rows


def _find_ledger_totals_start_row(ws) -> int:
    """
    급여대장 양식에서 '합계/소계' 영역이 시작되는 첫 행을 찾습니다.

    왜 필요한가?
    - 양식 하단에는 합계/소계/비고가 있고, 그 영역은 유지해야 합니다.
    - 직원 인원이 늘거나 줄면 직원 데이터 영역(2행/1명)만 늘리고 줄여야 합니다.

    반환:
      합계/소계 시작 행을 찾으면 그 행 번호,
      못 찾으면 ws.max_row+1 (즉, 끝까지 데이터 영역으로 간주)
    """
    section_sums = _find_ledger_section_sum_rows(ws)
    if section_sums:
        return section_sums[0]
    keywords = ("합계", "소계")
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and any(k in v for k in keywords):
                return r
    return ws.max_row + 1


def _find_ledger_employee_row_slots(ws) -> list[int]:
    """
    양식에 정의된 직원 입력 행(홀수 행, 2행/1명) 목록을 반환합니다.

    실무 양식은 '15명 합계' 다음에 '12명 합계'처럼 구간이 나뉘므로,
    첫 합계 행만 기준으로 연속 행(5,7,9,…)에 쓰면 16번째 이후가 누락됩니다.
    """
    sum_rows = _find_ledger_section_sum_rows(ws)
    if not sum_rows:
        totals_start = _find_ledger_totals_start_row(ws)
        return list(range(LEDGER_DATA_START, totals_start, LEDGER_ROWS_PER_EMP))

    block_starts = [LEDGER_DATA_START]
    for sr in sum_rows[:-1]:
        block_starts.append(sr + LEDGER_ROWS_PER_EMP)

    slots: list[int] = []
    for start, end in zip(block_starts, sum_rows):
        for row1 in range(start, end, LEDGER_ROWS_PER_EMP):
            slots.append(row1)
    return slots


def _ensure_ledger_employee_rows(ws, employee_count: int) -> None:
    """
    급여대장 양식에서 직원 영역을 인원수에 맞춰 2행씩 자동 삽입/삭제합니다.

    - 직원 영역: LEDGER_DATA_START부터 (2행/1명)
    - 합계/소계 영역은 _find_ledger_totals_start_row() 기준으로 보호
    """
    totals_start = _find_ledger_totals_start_row(ws)
    if totals_start <= LEDGER_DATA_START:
        # 비정상 템플릿(합계가 너무 위)인 경우: 삽입/삭제를 하지 않고 그대로 진행
        return

    current_rows = max(0, totals_start - LEDGER_DATA_START)  # 현재 직원 영역 총 행 수
    current_emp_capacity = current_rows // LEDGER_ROWS_PER_EMP
    needed_emp_capacity = max(0, employee_count)

    delta_emp = needed_emp_capacity - current_emp_capacity
    if delta_emp == 0:
        return

    if delta_emp > 0:
        # 부족하면: 합계/소계 시작 행 바로 위에 2행씩 삽입
        ws.insert_rows(totals_start, amount=delta_emp * LEDGER_ROWS_PER_EMP)
    else:
        # 남으면: 직원 영역의 마지막부터 2행씩 삭제 (합계/소계 영역은 유지)
        remove_emp = abs(delta_emp)
        for _ in range(remove_emp):
            delete_start = totals_start - LEDGER_ROWS_PER_EMP
            if delete_start < LEDGER_DATA_START:
                break
            ws.delete_rows(delete_start, amount=LEDGER_ROWS_PER_EMP)
            totals_start -= LEDGER_ROWS_PER_EMP


def _remove_all_sheet_comments(ws) -> None:
    """
    시트의 셀 메모(댓글)를 모두 제거합니다.

    급여대장 양식에 달린 메모는 출력 파일에 포함하지 않습니다.
    (병합셀 메모는 Excel '복구' 오류 원인이 되기도 합니다.)
    """
    comments = getattr(ws, "_comments", None)
    if comments is not None and hasattr(comments, "keys"):
        for key in list(comments.keys()):
            del comments[key]

    for row in ws.iter_rows(max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if getattr(cell, "comment", None) is not None:
                cell.comment = None


def _fix_ledger_totals_formulas(ws) -> None:
    """
    직원 행을 insert/delete 하면, 양식 내 '총합계' 수식이 고정 행번호를 참조해
    합계가 깨질 수 있습니다. 따라서 텍스트 위치를 다시 찾아 수식을 재작성합니다.

    템플릿 구조(급여대장양식.xlsx) 기준:
    - "합계 (...)" 행이 2개
    - "총합계 (...)" 행이 1개 (그 아래 줄이 짝행 합계)
    - 총합계 수식은 기본적으로 =F(합계1)+F(합계2) 형태
    """
    sum_rows: list[int] = []
    grand_row: int | None = None

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if not isinstance(a, str):
            continue
        if "총합계" in a:
            grand_row = r
        elif "합계" in a:
            sum_rows.append(r)

    # 합계가 2개가 아니더라도, 가능한 범위에서 동작하게 처리
    if grand_row is None or len(sum_rows) == 0:
        return

    sum_rows = sorted(sum_rows)[:2]
    r1 = sum_rows[0]
    r2 = sum_rows[1] if len(sum_rows) > 1 else sum_rows[0]

    # 총합계는 2행(홀/짝) 세트로 되어있음
    grand_row2 = grand_row + 1

    # 템플릿에 존재하는 컬럼(F~W)에 대해 수식 재작성
    # (실제 양식에서 총합계 행은 이 구간을 사용함)
    for col in range(6, 24):  # F(6) ~ W(23)
        col_letter = get_column_letter(col)
        write_merged_safe(ws, f"{col_letter}{grand_row}", f"={col_letter}{r1}+{col_letter}{r2}")
        write_merged_safe(ws, f"{col_letter}{grand_row2}", f"={col_letter}{r1+1}+{col_letter}{r2+1}")


def _copy_cell(src_cell, dst_cell) -> None:
    """
    셀 값/서식 복사 (도형/개체/댓글은 복사하지 않음).

    Excel 크래시의 주요 원인이 insert_rows + 개체(anchor) 충돌이므로,
    급여대장은 '행 삽입' 대신 새 시트에 셀을 복제하여 안전하게 생성합니다.
    """
    from copy import copy

    dst_cell.value = src_cell.value
    if src_cell.has_style:
        # StyleProxy는 hash 불가 → copy()로 안전하게 복제
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)
    else:
        dst_cell.number_format = src_cell.number_format
    # 댓글/메모는 병합셀 충돌 및 복구 팝업 원인이 될 수 있어 복사하지 않음
    dst_cell.comment = None


def _shift_range(range_str: str, row_offset: int) -> str:
    """병합 범위 문자열('A1:B2')을 행 오프셋만큼 이동."""
    from openpyxl.utils.cell import range_boundaries, get_column_letter

    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    return (
        f"{get_column_letter(min_col)}{min_row + row_offset}:"
        f"{get_column_letter(max_col)}{max_row + row_offset}"
    )


def _clone_ledger_sheet_safely(src_ws, employee_count: int):
    """
    급여대장 템플릿 시트를 안전하게 복제하여 '행 삽입 없이' 직원 수만큼 행을 확보합니다.

    반환: (새 Workbook, 새 Worksheet)
    """
    totals_start = _find_ledger_totals_start_row(src_ws)
    header_end = LEDGER_DATA_START - 1
    footer_start = totals_start

    # 템플릿이 기본으로 갖고 있는 직원 행 수(2행/1명)
    base_emp_rows = max(0, footer_start - LEDGER_DATA_START)
    base_emp_capacity = max(1, base_emp_rows // LEDGER_ROWS_PER_EMP)

    needed_emp_rows = employee_count * LEDGER_ROWS_PER_EMP
    dst_footer_start = LEDGER_DATA_START + needed_emp_rows
    footer_row_offset = dst_footer_start - footer_start

    dst_wb = Workbook()
    dst_ws = dst_wb.active
    dst_ws.title = src_ws.title

    # 열 너비/서식 복사
    for col_key, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_key].width = dim.width
        dst_ws.column_dimensions[col_key].hidden = dim.hidden
        dst_ws.column_dimensions[col_key].outlineLevel = dim.outlineLevel

    # 행 높이(헤더/푸터는 그대로, 직원 영역은 첫 2행 블록 높이로 복제)
    def copy_row_dim(src_r: int, dst_r: int) -> None:
        sd = src_ws.row_dimensions.get(src_r)
        dd = dst_ws.row_dimensions[dst_r]
        if sd is None:
            return
        dd.height = sd.height
        dd.hidden = sd.hidden
        dd.outlineLevel = sd.outlineLevel

    # 1) 헤더 복사
    for r in range(1, header_end + 1):
        copy_row_dim(r, r)
        for c in range(1, src_ws.max_column + 1):
            _copy_cell(src_ws.cell(r, c), dst_ws.cell(r, c))

    # 2) 직원 2행 블록(템플릿의 첫 직원 블록) 서식 복제
    src_block_r1 = LEDGER_DATA_START
    src_block_r2 = LEDGER_DATA_START + 1
    for i in range(employee_count):
        dst_r1 = LEDGER_DATA_START + i * 2
        dst_r2 = dst_r1 + 1
        copy_row_dim(src_block_r1, dst_r1)
        copy_row_dim(src_block_r2, dst_r2)
        for c in range(1, src_ws.max_column + 1):
            _copy_cell(src_ws.cell(src_block_r1, c), dst_ws.cell(dst_r1, c))
            _copy_cell(src_ws.cell(src_block_r2, c), dst_ws.cell(dst_r2, c))

    # 3) 푸터(합계/총합계/하단문구) 복사 — 아래로 이동
    for r in range(footer_start, src_ws.max_row + 1):
        dst_r = r + footer_row_offset
        copy_row_dim(r, dst_r)
        for c in range(1, src_ws.max_column + 1):
            _copy_cell(src_ws.cell(r, c), dst_ws.cell(dst_r, c))

    # 4) 병합셀 복제 (헤더/직원블록/푸터 각각 처리)
    # - 헤더 병합: 그대로
    # - 직원블록 병합(5~6행 범위 내): 직원 수만큼 행 오프셋 반복 복제
    # - 푸터 병합: footer_row_offset 만큼 이동
    src_merges = [str(m) for m in src_ws.merged_cells.ranges]
    for m in src_merges:
        from openpyxl.utils.cell import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(m)
        # 헤더
        if max_row <= header_end:
            dst_ws.merge_cells(m)
            continue
        # 직원 블록(첫 2행)만 복제
        if min_row >= src_block_r1 and max_row <= src_block_r2:
            for i in range(employee_count):
                offset = (LEDGER_DATA_START + i * 2) - src_block_r1
                dst_ws.merge_cells(_shift_range(m, offset))
            continue
        # 푸터
        if min_row >= footer_start:
            dst_ws.merge_cells(_shift_range(m, footer_row_offset))
            continue
        # 그 외 영역(헤더~직원 사이 등): 그대로 복제
        dst_ws.merge_cells(m)

    return dst_wb, dst_ws


def _ledger_gross_pay_formula(row1: int) -> str:
    """
    급여총액 — SUM(지급·공제 항목). 업무추진비(Q열·17열)는 제외.

    명부·소계에 포함된 항목이므로 17열은 합산하지 않습니다.
    """
    row2 = row1 + 1
    return f"=SUM(I{row2}:P{row2},R{row2},I{row1},P{row1},R{row1},S{row1})"


def _ledger_employment_insurance_formula(row1: int) -> str:
    """
    고용보험(근로자 부담 실업급여) — 급여대장 수식.

    · 산정 기준: 동 행 급여총액(S열·짝수행) = 과세 보수총액과 동일 기준(프로그램 산출)
    · 요율: 보수총액 × 0.9% (고용보험법상 근로자 부담 실업급여, 2024~2026 국고 요율)
    · 원단위: 10원 미만 ROUND (근로기준법·4대보험 실무 관행, round_won_tens와 동일)
    """
    row2 = row1 + 1
    rate = EMPLOYMENT_INSURANCE_WORKER_RATE
    gross_cell = f"S{row2}"
    return f"=IF({gross_cell}=\"\",0,ROUND(MAX(0,{gross_cell})*{rate},-1))"


def _ledger_net_pay_formula(row1: int) -> str:
    """차인지급 — 급여총액(수식) − 공제합계(수식)."""
    row2 = row1 + 1
    return f"=IF(S{row2}=\"\",0,MAX(0,S{row2}-W{row1}))"


def _ledger_total_deduction_formula(row1: int) -> str:
    """공제합계 — 소득세·지방세·건강·국민·고용보험(수식) 합."""
    row2 = row1 + 1
    return f"=SUM(T{row1},U{row1},V{row1},T{row2},U{row2})"


def _fill_ledger_employee(ws, row1: int, r: dict[str, Any]) -> None:
    """급여대장 한 명(2행 블록)을 기록합니다."""
    row2 = row1 + 1
    _set_cell(ws, row1, 1, r["seq"])
    _set_cell(ws, row1, 2, r["name"])
    _set_cell(ws, row1, 4, r.get("birth") or "")
    _set_cell(ws, row1, 5, r.get("hire_date") or "")
    _set_cell(
        ws,
        row1,
        6,
        format_hours_or_days(
            r.get("_monthly_work_hours")
            or r.get("base_days")
            or r.get("work_days")
            or 0
        ),
    )
    _set_cell(ws, row1, 7, r["base_hourly"])
    unpaid = safe_number(r.get("unpaid_days"), 0.0)
    if unpaid:
        _set_cell(ws, row1, 8, format_hours_or_days(unpaid))
    _set_cell(ws, row1, 9, r["base_salary"])
    _set_cell(ws, row1, 10, format_hours_or_days(r.get("ot_hours", 0)))
    _set_cell(ws, row1, 11, format_hours_or_days(r.get("shift_hours", 0)))
    _set_cell(ws, row1, 12, format_hours_or_days(r.get("night_hours", 0)))
    _set_cell(ws, row1, 13, format_hours_or_days(r.get("special_hours", 0)))
    _set_cell(ws, row1, 14, format_hours_or_days(r.get("special_ext_hours", 0)))
    early_h = safe_number(r.get("early_leave_hours"), 0.0)
    if early_h:
        _set_cell(ws, row1, 15, format_hours_or_days(early_h))
    _set_cell(ws, row1, 20, r["income_tax"])
    _set_cell(ws, row1, 21, r["health_insurance"])
    if r.get("insurance_exempt"):
        _set_cell(ws, row1, 22, 0)
    else:
        _set_cell(ws, row1, 22, _ledger_employment_insurance_formula(row1))
    _set_cell(ws, row1, 23, _ledger_total_deduction_formula(row1))

    _set_cell(ws, row2, 2, r.get("dept") or "")
    _set_cell(ws, row2, 6, format_hours_or_days(r.get("work_days", 0)))
    _set_cell(ws, row2, 7, round(r["ordinary_hourly"], 2))
    leave_disp = r.get("leave_usage_display")
    if leave_disp:
        _set_cell(ws, row2, 8, leave_disp)
    elif safe_number(r.get("leave_days"), 0) > 0:
        _set_cell(ws, row2, 8, format_hours_or_days(r["leave_days"]))
    _set_cell(ws, row2, 9, ledger_base_deduction_for_excel(r))
    _set_cell(ws, row2, 10, r["ot_pay"])
    _set_cell(ws, row2, 11, r["shift_pay"])
    _set_cell(ws, row2, 12, r["night_pay"])
    _set_cell(ws, row2, 13, r["special_pay"])
    _set_cell(ws, row2, 14, r["special_ext_pay"])
    early_ded = int(safe_number(r.get("early_leave_deduction"), 0.0))
    if early_ded:
        _set_cell(ws, row2, 15, early_ded)
    _set_cell(ws, row2, 16, r["position_pay"])
    # 17열 업무추진비: 명부·소계에 포함 → 별도 기록 없음
    _set_cell(ws, row2, 19, _ledger_gross_pay_formula(row1))
    _set_cell(ws, row2, 20, r["local_income_tax"])
    _set_cell(ws, row2, 21, r["national_pension"])
    _set_cell(ws, row2, 23, _ledger_net_pay_formula(row1))


def write_ledger(
    records: list[dict[str, Any]],
    template_dir: Path,
    out_dir: Path,
    *,
    output_filename: str = "급여대장.xlsx",
) -> Path:
    """
    급여대장.xlsx — 양식 2행/1인 구조에 맞춰 기록합니다.

    5행(홀수): 순번·성명·…·소득세·건강·고용(수식)·공제합계(수식)
    6행(짝수): …·급여총액(수식)·지방세·국민·차인지급(수식)
    고용보험(V열): =ROUND(MAX(0,S열급여총액)×0.9%,-1) — 근로자 실업급여 요율
    """
    template = template_dir / "급여대장양식.xlsx"
    output = copy_template(template, out_dir / output_filename)
    # IMPORTANT (Excel 크래시 방지):
    # 사용자가 준 급여대장 템플릿은 내부 개체/인쇄설정/수식이 복잡해서,
    # openpyxl로 행 삽입/대량 복제 시 Excel이 '복구' 후 강제 종료되는 경우가 있습니다.
    #
    # 따라서 급여대장은 "템플릿을 그대로 복사"하고,
    # 템플릿이 원래 갖고 있는 직원 영역(현재 15명+12명=27명) 안에서 '값만 채우는' 방식으로 동작합니다.
    # 인원이 27명을 초과하면, 2번째 급여대장 파일을 추가로 생성합니다.
    wb = load_workbook(output, data_only=False)
    ws = wb.active

    row_slots = _find_ledger_employee_row_slots(ws)
    capacity_emp = len(row_slots)
    this_batch = records[:capacity_emp]
    overflow = records[capacity_emp:]

    # 직원 영역만 지우고 다시 씁니다 (합계/소계 영역은 유지)
    if row_slots:
        clear_end = row_slots[-1] + LEDGER_ROWS_PER_EMP - 1
        _clear_rows(ws, row_slots[0], clear_end, 30)

    for row1, r in zip(row_slots, this_batch):
        _fill_ledger_employee(ws, row1, r)

    _remove_all_sheet_comments(ws)
    output = save_workbook(wb, output)
    wb.close()

    # 초과 인원 처리: 동일 템플릿으로 2번째 파일 생성 (Excel 안정성 우선)
    if overflow:
        extra_index = 1
        while (out_dir / f"급여대장_추가{extra_index}.xlsx").exists():
            extra_index += 1
        write_ledger(
            overflow,
            template_dir,
            out_dir,
            output_filename=f"급여대장_추가{extra_index}.xlsx",
        )

    return output


def write_payslip_list(records: list[dict[str, Any]], template_dir: Path, out_dir: Path) -> Path:
    """급여명세서.xlsx — 3행부터 1행=1명."""
    template = template_dir / "급여명세서양식.xlsx"
    output = copy_template(template, out_dir / "급여명세서.xlsx")

    wb = load_workbook(output)
    ws = wb.active

    _clear_rows(ws, PAYSLIP_DATA_START, ws.max_row, 35)

    for i, r in enumerate(records):
        row = PAYSLIP_DATA_START + i
        _set_cell(ws, row, 1, r["name"])
        _set_cell(ws, row, 2, r.get("phone") or "")
        _set_cell(ws, row, 3, r.get("birth") or "")
        _set_cell(ws, row, 4, r.get("pay_day") or "25일")
        _set_cell(ws, row, 5, r["base_hourly"])
        _set_cell(ws, row, 6, round(r["ordinary_hourly"], 2))
        _set_cell(ws, row, 7, r["base_salary"])
        _set_cell(ws, row, 8, ledger_base_deduction_for_excel(r))
        _set_cell(ws, row, 9, r["ot_pay"])
        _set_cell(ws, row, 10, r["shift_pay"])
        _set_cell(ws, row, 11, r["night_pay"])
        _set_cell(ws, row, 12, r["special_pay"])
        _set_cell(ws, row, 13, r["special_ext_pay"])
        _set_cell(ws, row, 15, r["position_pay"])
        # 업무추진비·연차수당 별도 미기록 (명부·소계 반영)
        _set_cell(ws, row, 19, r["income_tax"])
        _set_cell(ws, row, 20, r["local_income_tax"])
        _set_cell(ws, row, 21, r["health_insurance"])
        _set_cell(ws, row, 22, r["national_pension"])
        _set_cell(ws, row, 23, r["employment_insurance"])
        _set_cell(ws, row, 25, format_hours_or_days(r.get("work_days", 0)))
        _set_cell(ws, row, 27, format_hours_or_days(r.get("ot_hours", 0)))
        _set_cell(ws, row, 28, format_hours_or_days(r.get("shift_hours", 0)))
        _set_cell(ws, row, 29, format_hours_or_days(r.get("night_hours", 0)))

    output = save_workbook(wb, output)
    wb.close()
    return output


def write_payment_list(records: list[dict[str, Any]], template_dir: Path, out_dir: Path) -> Path:
    """지급내역.xlsx — 5행부터 1행=1명 (실수령액)."""
    template = template_dir / "지급내역양식.xlsx"
    output = copy_template(template, out_dir / "지급내역.xlsx")

    wb = load_workbook(output)
    ws = wb.active

    _clear_rows(ws, PAYMENT_DATA_START, ws.max_row, 12)

    for i, r in enumerate(records):
        row = PAYMENT_DATA_START + i
        _set_cell(ws, row, 1, i + 1)
        _set_cell(ws, row, 2, r["name"])
        _set_cell(ws, row, 3, r.get("workplace") or "한국앰코생산")
        _set_cell(ws, row, 4, r.get("bank_code") or "")
        _set_cell(ws, row, 5, r.get("bank_name") or "")
        _set_cell(ws, row, 6, r.get("account") or "")
        _set_cell(ws, row, 7, r["net_pay"])
        _set_cell(ws, row, 8, "급여")
        _set_cell(ws, row, 9, "씨앤엘급여")
        _set_cell(ws, row, 10, r.get("holder") or r["name"])

    output = save_workbook(wb, output)
    wb.close()
    return output


def write_all_outputs(
    records: list[dict[str, Any]],
    template_dir: Path,
    out_dir: Path,
    period_label: str | None = None,
) -> dict[str, Path | list[Path]]:
    """급여대장·급여명세서·지급내역 3종 파일을 지정 폴더에 저장합니다."""
    out_dir.mkdir(parents=True, exist_ok=True)

    write_ledger(records, template_dir, out_dir)
    ledger_paths = [out_dir / "급여대장.xlsx"]
    ledger_paths.extend(sorted(out_dir.glob("급여대장_추가*.xlsx")))

    return {
        "ledger": ledger_paths[0],
        "ledger_extra": ledger_paths[1:],
        "payslip": write_payslip_list(records, template_dir, out_dir),
        "payment": write_payment_list(records, template_dir, out_dir),
    }
