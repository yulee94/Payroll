"""
services/excel_diff.py - Excel 수정 전·후 비교 요약
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return f"{value:.4f}".rstrip("0").rstrip(".")
    return str(value).strip()


def summarize_excel_diff(
    before_path: Path,
    after_path: Path,
    *,
    max_rows: int = 400,
    max_cols: int = 40,
    max_details: int = 20,
) -> dict[str, Any]:
    """두 Excel 파일의 셀 값 차이를 요약합니다."""
    wb_before = load_workbook(before_path, data_only=True, read_only=True)
    wb_after = load_workbook(after_path, data_only=True, read_only=True)

    sheets_before = set(wb_before.sheetnames)
    sheets_after = set(wb_after.sheetnames)
    added = sorted(sheets_after - sheets_before)
    removed = sorted(sheets_before - sheets_after)
    common = sorted(sheets_before & sheets_after)

    details: list[str] = []
    cells_changed = 0

    for sheet_name in common:
        ws_b = wb_before[sheet_name]
        ws_a = wb_after[sheet_name]
        max_r = min(max(ws_b.max_row or 0, ws_a.max_row or 0), max_rows)
        max_c = min(max(ws_b.max_column or 0, ws_a.max_column or 0), max_cols)
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                vb = _cell_str(ws_b.cell(r, c).value)
                va = _cell_str(ws_a.cell(r, c).value)
                if vb == va:
                    continue
                cells_changed += 1
                if len(details) < max_details:
                    col = get_column_letter(c)
                    short_b = vb if len(vb) <= 24 else vb[:21] + "…"
                    short_a = va if len(va) <= 24 else va[:21] + "…"
                    details.append(f"{sheet_name} {col}{r}: 「{short_b}」 → 「{short_a}」")

    wb_before.close()
    wb_after.close()

    if added:
        details.insert(0, f"추가된 시트: {', '.join(added)}")
    if removed:
        details.insert(0, f"삭제된 시트: {', '.join(removed)}")

    summary_lines: list[str] = []
    if cells_changed:
        summary_lines.append(f"변경 셀 {cells_changed}곳")
    if added:
        summary_lines.append(f"시트 추가 {len(added)}개")
    if removed:
        summary_lines.append(f"시트 삭제 {len(removed)}개")
    if not summary_lines:
        summary_lines.append("표시 영역에서 값 변경 없음 (서식·수식만 바뀌었을 수 있음)")

    return {
        "cells_changed": cells_changed,
        "sheets_added": added,
        "sheets_removed": removed,
        "summary_text": " · ".join(summary_lines),
        "details": details,
        "truncated": cells_changed > max_details,
    }
